import os
"""
run_all.py  —  台股整合儀表板一鍵執行腳本
═══════════════════════════════════════════════════════════
執行一次，輸出單一 dashboard.json 給儀表板載入。

執行順序：
  Step 1  更新今日法人資料          taiwan_market_flow.py  → market_data/market_flow_all.csv
  Step 2  資金流向分析               analyze_local.py       → analysis_result
  Step 3  多時框雷達                 stock_radar_multi.py   → radar_multi
  Step 4  強勢股篩選（yfinance 下載）tw_strong_screener.py  → strong_stocks
  Step 5  創新高 + 連買+MA           screener_newhigh_insti.py → newhigh + insti_ma
  Step 6  合併輸出                   →  output/dashboard.json

優化：
  Step 4 + Step 5 共用同一份 yfinance 下載，節省 5-8 分鐘

使用：
  python run_all.py           # 完整執行
  python run_all.py --skip-flow   # 跳過法人爬蟲（已有今日資料時）
  python run_all.py --skip-yf     # 跳過 yfinance（測試用）

依賴：
  pip install requests pandas yfinance curl_cffi --break-system-packages
"""

import argparse
import json
import logging
import sys
import time
import urllib.request
import warnings
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import yfinance as yf

try:
    from vix_fetcher import get_vix, get_twvix
except ImportError:
    get_vix = None
    get_twvix = None  # vix_fetcher.py 不在同目錄時，Step 1h/1i 會自動跳過，不影響主流程

warnings.filterwarnings("ignore")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Telegram 通知設定 ─────────────────────────────────────
TG_TOKEN   = "8878076016:AAHqyAdWDN8ig2D8rC0vTr9SlQxvWkO4vxM"
TG_CHAT_ID = "1092403230"
TG_ENABLED = True   # 改成 False 可暫時關閉通知

# ── 路徑設定 ──────────────────────────────────────────────
BASE_DIR    = Path(".")
MARKET_DIR  = BASE_DIR / "market_data"
OUTPUT_DIR  = BASE_DIR / "output"
MARKET_CSV  = BASE_DIR / "market_flow_all.csv"   # 依需求搬到根目錄，跟 run_all.py 同一層（不再放 market_data/ 子資料夾）
FEAR_GREED_CSV = BASE_DIR / "fear_greed_history.csv"  # 恐慌貪婪指數逐日累積的原始數值，用來算60日百分位
INSTI_CSV   = OUTPUT_DIR / "insti_history.csv"
SHARES_CACHE_PATH = OUTPUT_DIR / "shares_cache.json"   # 流通股數快取（供投本比/外本比計算），root目錄優先於此fallback路徑
ENTRIES_HISTORY_CSV = OUTPUT_DIR / "entries_history.csv"   # 各策略每日選股歷史（供出場回測用）
MINERVINI_MIN_PASSED = 8   # Trend Template 完美趨勢：8項條件全過才記錄
CANSLIM_MIN_PASSED   = 6   # CANSLIM：全條件7/7 或 6/7 才記錄

for d in [MARKET_DIR, OUTPUT_DIR]:
    d.mkdir(exist_ok=True)

TODAY     = datetime.today().strftime("%Y%m%d")
TODAY_STR = datetime.today().strftime("%Y-%m-%d")


# ════════════════════════════════════════════════════════════
# STEP 1：更新今日法人資料
# 核心邏輯內嵌自 taiwan_market_flow.py（避免每天執行時依賴外部檔案）。
# taiwan_market_flow.py 本身仍保留，作為獨立的歷史批次補抓工具
# （--year=2023-2025 等模式），跟這裡的每日即時更新是分開的用途。
# ════════════════════════════════════════════════════════════

MF_DELAY = 2.5
MF_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer":    "https://www.twse.com.tw/",
    "Accept":     "application/json, text/plain, */*",
}

MF_INDUSTRY_MAP = {
    "2330":"晶圓代工","5347":"晶圓代工","6770":"晶圓代工",
    "2454":"IC設計","2303":"IC設計","3034":"IC設計","2379":"IC設計",
    "6415":"IC設計","6533":"IC設計","2344":"IC設計","8046":"IC設計",
    "3006":"IC設計","2481":"IC設計","2455":"IC設計","3450":"IC設計",
    "6257":"IC設計","3443":"IC設計","2329":"IC設計","8028":"IC設計",
    "4256":"IC設計","3036":"IC設計","3673":"IC設計",
    "2347":"封測","3711":"封測","6239":"封測","2449":"封測",
    "8150":"封測","2369":"封測","6282":"封測",
    "2337":"記憶體","3260":"記憶體",
    "2327":"被動元件","2492":"被動元件","2351":"被動元件",
    "3481":"面板","2409":"面板",
    "2313":"PCB","2367":"PCB","4958":"PCB","3037":"PCB",
    "3189":"PCB","3715":"PCB","6213":"PCB","2355":"PCB","2402":"PCB","4966":"PCB",
    "2382":"伺服器","6669":"伺服器","3231":"伺服器",
    "2345":"網通","6285":"網通",
    "3017":"散熱","2374":"散熱",
    "2308":"電源管理","6274":"工業電腦","2395":"工業電腦","4977":"光學",
    "2376":"主機板",
    "2357":"電腦品牌","2353":"電腦品牌","2312":"電腦品牌",
    "2301":"EMS","2324":"EMS","4938":"EMS","2356":"EMS",
    "2317":"電子製造","2371":"電機","1504":"電機","3706":"電腦周邊",
    "2603":"貨櫃航運","2615":"貨櫃航運","2609":"貨櫃航運",
    "2610":"航空","2618":"航空",
    "2606":"散裝航運","2605":"散裝航運","2637":"散裝航運",
    "2881":"金控","2882":"金控","2883":"金控","2884":"金控","2885":"金控",
    "2886":"金控","2887":"金控","2888":"金控","2889":"金控","2890":"金控",
    "2891":"金控","2892":"金控","2880":"金控","5876":"金控","5880":"金控",
    "2823":"金控","2801":"金控","2897":"金控",
    "2412":"電信","4904":"電信","3045":"電信",
    "2002":"鋼鐵","2006":"鋼鐵","2007":"鋼鐵","2008":"鋼鐵","2027":"特殊鋼",
    "1301":"石化","1303":"石化","1326":"石化","6505":"石化","1312":"石化",
    "1711":"化工","1717":"化工","1314":"化工",
    "1402":"紡織","1216":"食品","1101":"水泥","1102":"水泥",
    "1802":"玻璃","1605":"電線電纜",
    "6443":"太陽能","3576":"太陽能","8422":"環保","1565":"環保",
    "2207":"汽車","2105":"汽車","8112":"通路",
}

MF_SECTOR_MAP = {
    "晶圓代工":"半導體","IC設計":"半導體","封測":"半導體",
    "記憶體":"半導體","被動元件":"半導體",
    "伺服器":"科技硬體","網通":"科技硬體","散熱":"科技硬體",
    "電源管理":"科技硬體","工業電腦":"科技硬體","光學":"科技硬體","主機板":"科技硬體",
    "面板":"顯示面板","PCB":"PCB",
    "電子製造":"電子製造","EMS":"電子製造","電腦品牌":"電子製造","電腦周邊":"電子製造",
    "電機":"電機傳產","電線電纜":"電機傳產",
    "鋼鐵":"金屬原料","特殊鋼":"金屬原料",
    "石化":"石化原料","化工":"石化原料",
    "貨櫃航運":"航運","散裝航運":"航運","航空":"航運",
    "金控":"金融","電信":"電信",
    "紡織":"傳統產業","食品":"傳統產業","水泥":"傳統產業","玻璃":"傳統產業",
    "太陽能":"綠能","環保":"綠能",
    "汽車":"汽車","通路":"通路",
}

def _mf_get_ind(code): return MF_INDUSTRY_MAP.get(str(code).strip(), "其他")
def _mf_get_sec(ind):  return MF_SECTOR_MAP.get(ind, "其他")

def _mf_is_etf(code, name):
    c = str(code).strip()
    if not (len(c) == 4 and c.isdigit()): return True
    for kw in ["ETF","基金","永續","高息","高股息","台灣50","月配","季配","購","售","認"]:
        if kw in str(name): return True
    return False

def _mf_api_get(url, params):
    import requests
    try:
        r = requests.get(url, params=params, headers=MF_HEADERS, timeout=20, verify=False)
        r.raise_for_status()
        return r.json()
    except Exception:
        return {"stat": "ERROR"}

def _mf_to_k(v):
    try:
        f = float(str(v).replace(",", "").strip())
        return 0 if (f != f) else int(f / 1000)
    except Exception:
        return 0

def _mf_cl(s):
    try: return float(str(s).replace(",", "").strip())
    except Exception: return 0.0

def _mf_fetch_inst(date_str):
    data = _mf_api_get("https://www.twse.com.tw/rwd/zh/fund/T86",
                        {"response": "json", "date": date_str, "selectType": "ALL"})
    if data.get("stat") != "OK": return pd.DataFrame()
    rows = data.get("data", [])
    if not rows: return pd.DataFrame()

    fields = data.get("fields", [])
    df = pd.DataFrame(rows, columns=fields)
    df.columns = df.columns.str.strip()

    result = []
    for _, r in df.iterrows():
        vals = list(r)
        code = str(vals[0]).strip()
        name = str(vals[1]).strip()
        f_net = _mf_to_k(vals[4])  if len(vals) > 4  else 0
        t_net = _mf_to_k(vals[10]) if len(vals) > 10 else 0
        d_net = _mf_to_k(vals[13]) if len(vals) > 13 else 0
        d_hdg = _mf_to_k(vals[16]) if len(vals) > 16 else 0
        i_tot = _mf_to_k(vals[17]) if len(vals) > 17 else (f_net + t_net + d_net)
        ind = _mf_get_ind(code)
        result.append({
            "code": code, "name": name, "industry": ind, "sector": _mf_get_sec(ind),
            "is_etf": _mf_is_etf(code, name),
            "foreign_net": f_net, "trust_net": t_net,
            "dealer_net": d_net, "dealer_hedge": d_hdg, "inst_total": i_tot,
        })
    return pd.DataFrame(result)

def _mf_fetch_price(date_str):
    data = _mf_api_get("https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX",
                        {"response": "json", "date": date_str, "type": "ALL"})
    if data.get("stat") != "OK": return pd.DataFrame()
    for tbl in data.get("tables", []):
        f = tbl.get("fields", [])
        if "收盤價" in f and "證券代號" in f:
            df = pd.DataFrame(tbl["data"], columns=f)
            df.columns = df.columns.str.strip()
            cc    = next((c for c in df.columns if "代號" in c), None)
            clsc  = next((c for c in df.columns if "收盤" in c), None)
            # 證交所這裡是「漲跌價差」（純數字，永遠是正的，只有絕對值）跟「漲跌(+/-)」
            # （只有正負符號，沒有數字）兩個分開的欄位，要兩個一起用才能還原正確的漲跌方向，
            # 之前只抓了「漲跌價差」，導致算出來的漲跌全部變正值（幾乎每檔都被誤判成上漲）
            chg_mag_c = next((c for c in df.columns if "漲跌" in c and "價" in c), None)
            chg_sign_c = next((c for c in df.columns if "漲跌" in c and "價" not in c), None)
            volc  = next((c for c in df.columns if "成交" in c and ("股" in c or "量" in c)), None)
            if not (cc and clsc): continue
            out = pd.DataFrame()
            out["code"]  = df[cc].astype(str).str.strip()
            out["price"] = df[clsc].apply(_mf_cl)
            magnitude = df[chg_mag_c].apply(_mf_cl) if chg_mag_c else 0
            if chg_sign_c is not None:
                def _sign_of(raw):
                    s = str(raw)
                    if "-" in s: return -1
                    if "+" in s: return 1
                    return 0  # 空白／X／不明符號 = 平盤
                sign = df[chg_sign_c].apply(_sign_of)
                out["change"] = magnitude * sign
            else:
                out["change"] = magnitude
            out["volume_k"] = (df[volc].apply(_mf_cl) / 1000).astype(int) if volc else 0
            base = out["price"] - out["change"]
            out["change_pct"] = (out["change"] / base * 100).round(2).where(base != 0, 0)
            return out
    return pd.DataFrame()

def _mf_process(inst_df, price_df, date_label):
    if inst_df.empty: return pd.DataFrame()
    df = inst_df.copy()
    if not price_df.empty:
        df["price"]      = df["code"].map(dict(zip(price_df["code"], price_df["price"]))).fillna(0)
        df["change_pct"] = df["code"].map(dict(zip(price_df["code"], price_df["change_pct"]))).fillna(0)
        df["volume_k"]   = df["code"].map(dict(zip(price_df["code"], price_df["volume_k"]))).fillna(0)
    else:
        df["price"] = df["change_pct"] = df["volume_k"] = 0
    df.insert(0, "date", date_label)
    df = df[df["code"].str.match(r'^\d{4}$')].copy()
    return df

def _mf_save(df, date_str):
    MARKET_DIR.mkdir(exist_ok=True)
    df.to_csv(MARKET_DIR / f"market_{date_str}.csv", index=False, encoding="utf-8-sig")
    if MARKET_CSV.exists():
        ex = pd.read_csv(MARKET_CSV, dtype=str)
        ex = ex[ex["date"] != df["date"].iloc[0]]
        combined = pd.concat([ex, df.astype(str)], ignore_index=True)
    else:
        combined = df.astype(str)
    combined.sort_values("date").reset_index(drop=True).to_csv(MARKET_CSV, index=False, encoding="utf-8-sig")

def market_flow_run_one(d: date, verbose=True):
    """抓單日全市場法人+價格資料並存檔（原 taiwan_market_flow.py 的 run_one）
    回傳 (是否成功: bool, 全市場法人買賣超總額: dict|None)
    """
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    ds = d.strftime("%Y%m%d")
    dl = d.strftime("%Y-%m-%d")
    inst  = _mf_fetch_inst(ds);  time.sleep(MF_DELAY)
    price = _mf_fetch_price(ds); time.sleep(MF_DELAY)
    df = _mf_process(inst, price, dl)
    if df.empty: return False, None
    _mf_save(df, ds)
    inst_summary = None
    if verbose:
        ne = df[~(df["is_etf"].astype(str).str.lower() == "true")]
        f = ne["foreign_net"].sum(); t = ne["trust_net"].sum(); i = ne["inst_total"].sum()
        dl_ = ne["dealer_net"].sum() if "dealer_net" in ne.columns else None
        log.info("  法人資料：%s  外資%+.1f萬  投信%+.1f萬  合計%+.1f萬  (%d檔)",
                  dl, f/10000, t/10000, i/10000, len(ne))
        inst_summary = {
            "date": dl,
            "foreign_net": round(float(f), 0),
            "trust_net": round(float(t), 0),
            "dealer_net": round(float(dl_), 0) if dl_ is not None else None,
            "inst_total": round(float(i), 0),
        }
        # 順便算漲跌家數（不算ETF，跟法人統計口徑一致，只看一般個股廣度）
        if "change_pct" in ne.columns:
            cp = pd.to_numeric(ne["change_pct"], errors="coerce")
            up, down, flat = int((cp > 0).sum()), int((cp < 0).sum()), int((cp == 0).sum())
            log.info("  漲跌家數：%s  上漲%d  下跌%d  平盤%d  (%d檔)", dl, up, down, flat, len(ne))
            inst_summary["breadth"] = {"up": up, "down": down, "flat": flat, "total": int(len(ne))}
    return True, inst_summary


def step1_update_market_flow(skip=False):
    """更新今日全市場法人資料到 market_flow_all.csv
    回傳 (ok: bool, inst_summary: dict|None)"""
    if skip:
        log.info("STEP 1 跳過（--skip-flow）")
        return MARKET_CSV.exists(), None

    try:
        log.info("STEP 1: 更新今日法人資料...")
        ok, inst_summary = market_flow_run_one(date.today(), verbose=True)
        if ok:
            log.info("STEP 1 完成")
        else:
            log.warning("STEP 1: 今日可能是假日或資料尚未更新")
        return (ok or MARKET_CSV.exists()), inst_summary
    except Exception as e:
        log.error("STEP 1 失敗：%s", e)
        return MARKET_CSV.exists(), None


# ════════════════════════════════════════════════════════════
# STEP 1b：大盤月轉折更新（原 stock.py / 台股每日更新.py）
# 資料存放沿用原本的 Excel（日開收盤價_更新.xlsx），run_all.py 執行時
# 自動觸發更新 + 重新產生 K線圖.html，不需要再手動雙擊 .bat 執行。
# 若 Excel 或 openpyxl 不存在，僅記警告、不中斷主流程。
# ════════════════════════════════════════════════════════════

PIVOT_EXCEL_FILE  = BASE_DIR / "日開收盤價_更新.xlsx"
PIVOT_SHEET_DAILY = "每日開收盤價"
PIVOT_SHEET_MONTH = "月轉折計算"
PIVOT_OUTPUT_HTML = BASE_DIR / "K線圖.html"
PIVOT_TWSE_URL = "https://www.twse.com.tw/en/indicesReport/MI_5MINS_HIST"
PIVOT_HEADERS  = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}


def _pv_fetch_month_ohlc(yyyymm):
    import requests
    try:
        r = requests.get(PIVOT_TWSE_URL,
                          params={"response": "json", "date": yyyymm + "01"},
                          headers=PIVOT_HEADERS, timeout=20, verify=False)
        r.raise_for_status(); data = r.json()
    except Exception as e:
        log.warning("  [轉折] %s 抓取失敗：%s", yyyymm, e)
        return {}
    if data.get("stat", "").upper() != "OK":
        return {}
    result = {}
    for row in data.get("data", []):
        try:
            p = str(row[0]).split("/")
            yr = int(p[0]); yr = yr + 1911 if yr < 1911 else yr
            key = f"{yr:04d}-{int(p[1]):02d}-{int(p[2]):02d}"
            f = lambda s: float(str(s).replace(",", ""))
            o, h, l, c = f(row[1]), f(row[2]), f(row[3]), f(row[4])
            result[key] = {"open": o, "high": h, "low": l, "close": c, "change": round(c - o, 2)}
        except Exception:
            continue
    return result


def _pv_calc_monthly_pivots(month_ohlc):
    labels = sorted(month_ohlc.keys())
    result, prev_pivot = {}, None
    for i, lbl in enumerate(labels):
        m = month_ohlc[lbl]
        if i == 0:
            two_high, two_low = m["high"], m["low"]
        else:
            prev = month_ohlc[labels[i - 1]]
            two_high = max(prev["high"], m["high"])
            two_low = min(prev["low"], m["low"])
        pivot = round((two_high + two_low) / 2, 6)
        result[lbl] = {**m, "two_high": round(two_high, 2), "two_low": round(two_low, 2),
                       "pivot": pivot, "ma_pivot": prev_pivot}
        prev_pivot = pivot
    return result


def _pv_copy_style(src, dst):
    if src.has_style:
        import copy as _copy
        for a in ("font", "fill", "border", "alignment", "number_format"):
            try: setattr(dst, a, _copy.copy(getattr(src, a)))
            except Exception: pass


def _pv_update_excel(target):
    import openpyxl
    log.info("  [轉折] 載入 %s", PIVOT_EXCEL_FILE)
    wb = openpyxl.load_workbook(PIVOT_EXCEL_FILE)
    ws_d = wb[PIVOT_SHEET_DAILY]
    ws_m = wb[PIVOT_SHEET_MONTH]

    d2r, last_price_row = {}, 1
    for r in range(2, ws_d.max_row + 1):
        v = ws_d.cell(r, 1).value
        if isinstance(v, datetime):
            d2r[v.date()] = r
            if ws_d.cell(r, 2).value is not None:
                last_price_row = r

    last_date = ws_d.cell(last_price_row, 1).value
    if isinstance(last_date, datetime): last_date = last_date.date()
    log.info("  [轉折] 每日最後日期：%s", last_date)

    if last_date >= target:
        log.info("  [轉折] 每日資料已是最新")
    else:
        months = sorted({(last_date + timedelta(days=i)).strftime("%Y%m")
                          for i in range(1, (target - last_date).days + 2)})
        all_new = {}
        for ym in months:
            all_new.update(_pv_fetch_month_ohlc(ym)); time.sleep(0.8)

        ref_row = last_price_row
        for r in range(ws_d.max_row, 1, -1):
            if ws_d.cell(r, 8).value is not None: ref_row = r; break

        prev_c = ws_d.cell(last_price_row, 5).value
        prev_c = float(prev_c) if prev_c else None
        filled = 0

        for ds, p in sorted(all_new.items()):
            d = datetime.strptime(ds, "%Y-%m-%d").date()
            if d <= last_date or d > target: continue
            if d in d2r: r = d2r[d]
            else:
                r = last_price_row + 1
                ws_d.cell(r, 1).value = datetime(d.year, d.month, d.day)
                ws_d.cell(r, 1).number_format = "YYYY/MM/DD"
                d2r[d] = r; last_price_row = r

            chg = round(p["close"] - prev_c, 2) if prev_c else round(p["close"] - p["open"], 2)
            for col, key in zip([2, 3, 4, 5, 6], ["open", "high", "low", "close"]):
                ws_d.cell(r, col).value = p[key]
            ws_d.cell(r, 6).value = chg

            prev_i = r - 1
            while prev_i > 1 and ws_d.cell(prev_i, 9).value is None: prev_i -= 1
            ch = ws_d.cell(r, 8); ch.value = f"=E{r}-G{r}"; _pv_copy_style(ws_d.cell(ref_row, 8), ch)
            ci = ws_d.cell(r, 9); ci.value = f"=H{r}/1000"; _pv_copy_style(ws_d.cell(ref_row, 9), ci)
            cj = ws_d.cell(r, 10)
            cj.value = f"=AVERAGE(I{prev_i}:I{r})" if prev_i > 1 else f"=I{r}"
            _pv_copy_style(ws_d.cell(ref_row, 10), cj)

            ref_row = r; prev_c = p["close"]; filled += 1
        log.info("  [轉折] 新增 %d 個交易日", filled)

    month_ohlc = {}
    for d, r in sorted(d2r.items()):
        if ws_d.cell(r, 2).value is None: continue
        lbl = f"{str(d.year)[-2:]}M{d.month:02d}"
        o, h, l, c = [float(ws_d.cell(r, i).value) for i in [2, 3, 4, 5]]
        if lbl not in month_ohlc:
            month_ohlc[lbl] = {"open": o, "high": h, "low": l, "close": c, "days": 1}
        else:
            m = month_ohlc[lbl]; m["high"] = max(m["high"], h); m["low"] = min(m["low"], l)
            m["close"] = c; m["days"] += 1

    pivots = _pv_calc_monthly_pivots(month_ohlc)

    months_in_sheet, last_m_row = {}, 1
    for r in range(2, ws_m.max_row + 1):
        v = ws_m.cell(r, 1).value
        if v and isinstance(v, str) and 'M' in v:
            months_in_sheet[v] = r; last_m_row = r

    for lbl in sorted(pivots.keys()):
        p = pivots[lbl]
        r = months_in_sheet.get(lbl) or (last_m_row := last_m_row + 1, last_m_row)[1]
        months_in_sheet[lbl] = r
        ws_m.cell(r, 1).value = lbl
        ws_m.cell(r, 2).value = p["days"]
        for col, key in zip([3, 4, 5, 6, 7, 8, 9], ["open", "high", "low", "close", "two_high", "two_low", "pivot"]):
            ws_m.cell(r, col).value = round(p[key], 3)
        ws_m.cell(r, 10).value = round(p["ma_pivot"], 3) if p["ma_pivot"] else ""

    last_lbl = sorted(pivots.keys())[-1]
    last_pivot = pivots[last_lbl]["pivot"]
    ws_m.cell(last_m_row + 1, 10).value = round(last_pivot, 3)
    log.info("  [轉折] 更新至 %s，轉折點=%s", last_lbl, last_pivot)

    pivot_map = {}
    for lbl, p in pivots.items():
        parts = lbl.split("M")
        ym = f"20{parts[0]}{parts[1]}"
        if p["ma_pivot"] is not None: pivot_map[ym] = p["ma_pivot"]

    updated_g = 0
    for d, r in sorted(d2r.items()):
        if ws_d.cell(r, 2).value is None: continue
        ym = d.strftime("%Y%m"); piv = pivot_map.get(ym)
        if piv:
            ex = ws_d.cell(r, 7).value
            if not isinstance(ex, (int, float)) or abs(float(ex) - piv) > 0.001:
                ws_d.cell(r, 7).value = round(piv, 3); updated_g += 1

    wb.save(PIVOT_EXCEL_FILE)
    log.info("  [轉折] Excel 已存：%s", PIVOT_EXCEL_FILE)
    return pivots


def _pv_build_chart(pivots):
    import openpyxl, json as _json, calendar
    wb = openpyxl.load_workbook(PIVOT_EXCEL_FILE, data_only=True)
    ws_d = wb[PIVOT_SHEET_DAILY]

    daily = []
    for r in range(2, ws_d.max_row + 1):
        d = ws_d.cell(r, 1).value
        o, h, l, c = [ws_d.cell(r, i).value for i in [2, 3, 4, 5]]
        chg = ws_d.cell(r, 6).value
        piv = ws_d.cell(r, 7).value
        if not (isinstance(d, datetime) and o and h and l and c): continue
        daily.append({
            "date": d.strftime("%Y-%m-%d"),
            "open": round(float(o), 2), "high": round(float(h), 2),
            "low": round(float(l), 2), "close": round(float(c), 2),
            "change": round(float(chg), 2) if chg else 0,
            "pivot": round(float(piv), 2) if isinstance(piv, (int, float)) else None,
            "next_piv": None
        })

    if daily:
        last_d = daily[-1]["date"]
        last_ym = f"{last_d[2:4]}M{last_d[5:7]}"
        if last_ym in pivots:
            last_next = round(pivots[last_ym]["pivot"], 2)
            daily[-1]["next_piv"] = last_next
            last_dt = datetime.strptime(last_d, "%Y-%m-%d")
            if last_dt.month == 12:
                ny, nm = last_dt.year + 1, 1
            else:
                ny, nm = last_dt.year, last_dt.month + 1
            days_in_month = calendar.monthrange(ny, nm)[1]
            fake_days = [
                datetime(ny, nm, d).strftime("%Y-%m-%d")
                for d in range(1, days_in_month + 1)
                if datetime(ny, nm, d).weekday() < 5
            ]
            for fd in fake_days:
                daily.append({"date": fd, "open": None, "high": None, "low": None,
                               "close": None, "change": None, "pivot": None, "next_piv": last_next})

    monthly = []
    for lbl in sorted(pivots.keys()):
        p = pivots[lbl]
        monthly.append({
            "ym": lbl,
            "open": round(p["open"], 2), "high": round(p["high"], 2),
            "low": round(p["low"], 2), "close": round(p["close"], 2),
            "pivot": round(p["ma_pivot"], 2) if p["ma_pivot"] else None,
            "next_piv": round(p["pivot"], 2)
        })

    html = _pv_html(daily, monthly)
    with open(PIVOT_OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    log.info("  [轉折] K線圖已輸出：%s", PIVOT_OUTPUT_HTML)
    return daily, monthly


def _pv_html(daily, monthly):
    import json as _json
    dj = _json.dumps(daily, ensure_ascii=False)
    mj = _json.dumps(monthly, ensure_ascii=False)
    return f"""<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>台股加權指數 K 線圖</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',sans-serif;background:#0f1117;color:#e0e0e0;min-height:100vh}}
.hdr{{padding:14px 22px;border-bottom:1px solid #222;display:flex;align-items:center;gap:14px;flex-wrap:wrap}}
h1{{font-size:17px;font-weight:600;color:#fff}}
.stats{{display:flex;gap:10px;flex-wrap:wrap;margin-left:auto}}
.st{{background:#1a1d26;border:1px solid #2a2d3a;border-radius:8px;padding:5px 13px;font-size:12px;color:#999}}
.st b{{display:block;font-size:15px;font-weight:600;color:#fff}}
.up{{color:#e24b4a!important}}.dn{{color:#1d9e75!important}}.am{{color:#e6a817!important}}.pr{{color:#c084fc!important}}
.ctrl{{padding:9px 22px;display:flex;gap:8px;flex-wrap:wrap;align-items:center;border-bottom:1px solid #1a1a2a}}
button{{padding:5px 15px;border-radius:6px;border:1px solid #333;background:#1a1d26;color:#bbb;font-size:13px;cursor:pointer}}
button.on{{background:#2a3050;color:#fff;border-color:#4a5a90}}
select{{padding:5px 9px;border-radius:6px;border:1px solid #333;background:#1a1d26;color:#ccc;font-size:13px}}
.leg{{display:flex;gap:14px;margin-left:auto;font-size:12px;color:#777;align-items:center}}
.lsq{{width:10px;height:10px;border-radius:2px;display:inline-block;margin-right:3px;vertical-align:middle}}
.lso{{width:22px;height:0;display:inline-block;vertical-align:middle;margin-right:3px}}
#cw{{position:relative;padding:14px 18px;height:calc(100vh - 186px);min-height:380px}}
#tip{{display:none;position:fixed;background:#1e2030;border:1px solid #2a2d3a;border-radius:8px;
      padding:9px 13px;font-size:12px;pointer-events:none;z-index:99;line-height:1.9;min-width:190px}}
</style></head><body>
<div class="hdr">
  <h1>🇹🇼 台股加權指數 K 線圖</h1>
  <div class="stats" id="sb"></div>
</div>
<div class="ctrl">
  <button class="on" id="bd" onclick="sw('day',this)">日K</button>
  <button id="bm" onclick="sw('mon',this)">月K</button>
  <select id="rng" onchange="draw()">
    <option value="all">全部</option>
    <option value="2024">2024年</option>
    <option value="2025">2025年</option>
    <option value="2026">2026年</option>
    <option value="r60">近60日</option>
    <option value="r120">近120日</option>
  </select>
  <div class="leg">
    <span><span class="lsq" style="background:#e24b4a"></span>陽線</span>
    <span><span class="lsq" style="background:#1d9e75"></span>陰線</span>
    <span><span class="lso" style="border-top:2.5px solid #e6a817"></span>多空轉折（當月操作用）</span>
    <span><span class="lso" style="border-top:2px dashed #c084fc"></span>預告轉折（下月用）</span>
  </div>
</div>
<div id="cw"><canvas id="ch" role="img" aria-label="台股K線圖"></canvas></div>
<div id="tip"></div>
<script>
const D={dj}, M={mj};
let mode='day', chart=null;

function get(){{
  const r=document.getElementById('rng').value, src=mode==='day'?D:M;
  if(r==='all') return src;
  if(r==='r60') return D.slice(-60);
  if(r==='r120') return D.slice(-120);
  return mode==='day'?D.filter(x=>x.date.startsWith(r)):M.filter(x=>x.ym.startsWith(r));
}}
function lbl(d){{
  return mode==='day'?d.date.slice(5):('20'+d.ym.split('M')[0]+'/'+d.ym.split('M')[1]);
}}

function draw(){{
  const data=get(); if(!data.length) return;
  const n=data.length, cw=Math.max(2,Math.min(mode==='mon'?28:18,Math.floor(580/n-2)));
  const gc='rgba(255,255,255,0.055)', tc='#555';

  const cp={{id:'cp',afterDatasetsDraw(ch){{
    const{{ctx,scales:{{x,y}}}}=ch;
    data.forEach((d,i)=>{{
      if(!d.open||!d.close) return;
      const xc=x.getPixelForValue(i), half=cw/2, bull=d.close>=d.open;
      const col=bull?'#e24b4a':'#1d9e75';
      const yH=y.getPixelForValue(d.high), yL=y.getPixelForValue(d.low);
      const yO=y.getPixelForValue(d.open), yC=y.getPixelForValue(d.close);
      const top=Math.min(yO,yC), bot=Math.max(yO,yC), bh=Math.max(1.5,bot-top);
      ctx.save();
      ctx.strokeStyle=col; ctx.lineWidth=1.5;
      ctx.beginPath(); ctx.moveTo(xc,yH); ctx.lineTo(xc,top); ctx.stroke();
      ctx.beginPath(); ctx.moveTo(xc,bot); ctx.lineTo(xc,yL); ctx.stroke();
      ctx.fillStyle=col; ctx.fillRect(xc-half,top,cw,bh);
      ctx.restore();
    }});
  }}}};

  if(chart) chart.destroy();
  chart=new Chart(document.getElementById('ch'),{{
    type:'line',
    data:{{
      labels: data.map(lbl),
      datasets:[
        {{
          label:'多空轉折（當月）',
          data: data.map(d=>d.pivot),
          borderColor:'#e6a817', borderWidth:2.5,
          borderDash:[], pointRadius:0,
          tension:0, fill:false, spanGaps:true, yAxisID:'y'
        }},
        {{
          label:'預告轉折（下月）',
          data: data.map(d=>d.next_piv),
          borderColor:'#c084fc', borderWidth:1.8,
          borderDash:[5,4], pointRadius:0,
          tension:0, fill:false, spanGaps:true, yAxisID:'y'
        }},
        {{
          label:'_h',
          data: data.map(d=>d.high),
          borderColor:'transparent', backgroundColor:'transparent',
          pointRadius:0, fill:false, yAxisID:'y'
        }}
      ]
    }},
    options:{{
      responsive:true, maintainAspectRatio:false, animation:{{duration:200}},
      interaction:{{mode:'index',intersect:false}},
      plugins:{{legend:{{display:false}}, tooltip:{{enabled:false}}}},
      scales:{{
        x:{{grid:{{color:gc}},ticks:{{color:tc,font:{{size:11}},maxRotation:45,autoSkip:n>30,maxTicksLimit:20}}}},
        y:{{grid:{{color:gc}},ticks:{{color:tc,font:{{size:11}},
          callback:v=>v>=10000?Math.round(v/100)/10+'k':v.toLocaleString()}}}}
      }}
    }},
    plugins:[cp]
  }});

  const cvs=document.getElementById('ch'), tip=document.getElementById('tip');
  cvs.onmousemove=e=>{{
    const rc=cvs.getBoundingClientRect(), mx=e.clientX-rc.left;
    const{{scales:{{x}}}}=chart;
    let ci=-1, md=1e9;
    data.forEach((_,i)=>{{ const px=x.getPixelForValue(i),dv=Math.abs(px-mx); if(dv<md){{md=dv;ci=i;}} }});
    if(ci>=0&&md<cw*3){{
      const d=data[ci];
      if(!d.open){{
        tip.innerHTML=`<b style="color:#c084fc">預告轉折（下月生效）</b><br>
<span style="color:#c084fc">╌</span> ${{d.next_piv?Math.round(d.next_piv).toLocaleString():'—'}}`;
        tip.style.display='block';
        tip.style.left=(e.clientX+200>window.innerWidth?e.clientX-205:e.clientX+12)+'px';
        tip.style.top=(e.clientY-10)+'px';
        return;
      }}
      const bull=d.close>=d.open;
      const chg=d.close-d.open, pct=(chg/d.open*100).toFixed(2);
      const pAbv=d.pivot&&d.close>d.pivot;
      const pLine=d.pivot
        ?`<span style="color:#e6a817">▬</span> 多空轉折：${{Math.round(d.pivot).toLocaleString()}} <span style="color:${{pAbv?'#1d9e75':'#e24b4a'}}">${{pAbv?'▲多':'▼空'}}</span>`
        :'';
      const npLine=d.next_piv
        ?`<br><span style="color:#c084fc">╌</span> 預告轉折（下月）：${{Math.round(d.next_piv).toLocaleString()}}`
        :'';
      tip.innerHTML=`<b style="color:#eee">${{mode==='day'?d.date:lbl(d)}}</b><br>
開：${{d.open.toLocaleString()}} &nbsp; 高：<b style="color:#e24b4a">${{d.high.toLocaleString()}}</b><br>
低：<b style="color:#1d9e75">${{d.low.toLocaleString()}}</b> &nbsp; 收：${{d.close.toLocaleString()}}<br>
漲跌：<span style="color:${{bull?'#e24b4a':'#1d9e75'}}">${{bull?'+':''}}${{chg.toFixed(0)}} (${{bull?'+':''}}${{pct}}%)</span><br>
${{pLine}}${{npLine}}`;
      tip.style.display='block';
      tip.style.left=(e.clientX+200>window.innerWidth?e.clientX-205:e.clientX+12)+'px';
      tip.style.top=(e.clientY-10)+'px';
    }}
  }};
  cvs.onmouseleave=()=>{{tip.style.display='none';}};

  const realData=data.filter(d=>d.open!=null);
  const last=realData[realData.length-1], first=realData[0];
  const hi=Math.max(...realData.map(d=>d.high)), lo=Math.min(...realData.map(d=>d.low));
  const hd=realData[realData.map(d=>d.high).indexOf(hi)], ld=realData[realData.map(d=>d.low).indexOf(lo)];
  const tc2=((last.close-first.open)/first.open*100).toFixed(1), bull2=parseFloat(tc2)>=0;
  const nextPivVal = data.find(d=>d.next_piv!=null);
  const nextPiv = nextPivVal ? nextPivVal.next_piv : null;
  document.getElementById('sb').innerHTML=
    `<div class="st">最後交易日<b>${{mode==='day'?last.date:lbl(last)}}</b></div>`+
    `<div class="st">收盤<b class="${{last.close>=last.open?'up':'dn'}}">${{last.close.toLocaleString()}}</b></div>`+
    `<div class="st">區間高<b class="up">${{hi.toLocaleString()}}<br><small style="font-size:10px;font-weight:400">${{mode==='day'?hd.date:lbl(hd)}}</small></b></div>`+
    `<div class="st">區間低<b class="dn">${{lo.toLocaleString()}}<br><small style="font-size:10px;font-weight:400">${{mode==='day'?ld.date:lbl(ld)}}</small></b></div>`+
    `<div class="st">漲跌<b class="${{bull2?'up':'dn'}}">${{bull2?'+':''}}${{tc2}}%</b></div>`+
    `<div class="st">多空轉折<b class="am">${{last.pivot?Math.round(last.pivot).toLocaleString():'—'}}</b></div>`+
    `<div class="st">預告轉折（下月）<b class="pr">${{nextPiv?Math.round(nextPiv).toLocaleString():'—'}}</b></div>`;
}}

function sw(m,btn){{
  mode=m;
  document.querySelectorAll('.ctrl button').forEach(b=>b.classList.remove('on'));
  btn.classList.add('on');
  const sel=document.getElementById('rng');
  ['r60','r120'].forEach(v=>{{const o=[...sel.options].find(x=>x.value===v);if(o)o.style.display=m==='mon'?'none':'';}});
  if(['r60','r120'].includes(sel.value)&&m==='mon') sel.value='all';
  draw();
}}
draw();
</script></body></html>"""


def compute_recent_highs(daily):
    """從大盤日線資料（step1b_pivot_update回傳的daily清單）算近3個月、近6個月最高點，
    以及目前收盤價相對這兩個高點的回撤百分比。不用額外抓資料，沿用既有的大盤日線。
    回傳 {"latest_close":..., "latest_date":...,
          "high_3m":..., "high_3m_date":..., "drawdown_3m_pct":...,
          "high_6m":..., "high_6m_date":..., "drawdown_6m_pct":...} 或 None
    """
    valid = [d for d in daily if d.get("close") is not None and d.get("high") is not None]
    if not valid:
        return None
    latest = valid[-1]
    latest_close = latest["close"]
    latest_dt = datetime.strptime(latest["date"], "%Y-%m-%d")

    def window(months):
        cutoff = latest_dt - timedelta(days=months * 30)
        rows = [d for d in valid if datetime.strptime(d["date"], "%Y-%m-%d") >= cutoff]
        if not rows:
            return None, None
        best = max(rows, key=lambda d: d["high"])
        return best["high"], best["date"]

    result = {"latest_close": latest_close, "latest_date": latest["date"]}
    h3, h3_date = window(3)
    if h3:
        result["high_3m"] = h3
        result["high_3m_date"] = h3_date
        result["drawdown_3m_pct"] = round((latest_close - h3) / h3 * 100, 2)
    h6, h6_date = window(6)
    if h6:
        result["high_6m"] = h6
        result["high_6m_date"] = h6_date
        result["drawdown_6m_pct"] = round((latest_close - h6) / h6 * 100, 2)

    return result if (h3 or h6) else None


def step1b_pivot_update(skip=False):
    """大盤月轉折更新：更新 日開收盤價_更新.xlsx + 重新產生 K線圖.html
    回傳 {"daily":[...], "monthly":[...]} 供 dashboard.json 使用，失敗/跳過時回傳 None"""
    if skip:
        log.info("STEP 1b 跳過（--skip-pivot）")
        return None
    if not PIVOT_EXCEL_FILE.exists():
        log.warning("STEP 1b: 找不到 %s，跳過大盤轉折更新", PIVOT_EXCEL_FILE)
        return None
    try:
        import openpyxl  # noqa: F401  只是確認套件存在
    except ImportError:
        log.warning("STEP 1b: 未安裝 openpyxl，跳過大盤轉折更新（pip install openpyxl）")
        return None
    try:
        log.info("STEP 1b: 更新大盤月轉折...")
        pivots = _pv_update_excel(date.today())
        daily, monthly = _pv_build_chart(pivots)
        log.info("STEP 1b 完成")
        return {"daily": daily, "monthly": monthly}
    except Exception as e:
        log.error("STEP 1b 失敗：%s", e)
        return None


# ════════════════════════════════════════════════════════════
# STEP 1c：融資餘額變化（TWSE 信用交易統計 MI_MARGN）
# ⚠️ 欄位名稱是依公開資料慣例寫的，實際跑一次確認 fields 是否對得上
# （這台機器連不到 twse.com.tw，沒辦法在這裡先驗證）
# ════════════════════════════════════════════════════════════

def step1d_business_cycle_signal(skip=False):
    """抓國發會「景氣指標查詢系統」的 lightscore API，取最新一期景氣對策信號綜合分數，
    再依官方公開的固定門檻換算成燈號（這支API只回傳分數序列，不含燈號文字，換算門檻
    是國發會公告的官方標準，不是自己猜的）：
      9-16分  藍燈
      17-22分 黃藍燈
      23-31分 綠燈
      32-37分 黃紅燈
      38-45分 紅燈
    來源：https://index.ndc.gov.tw/n/json/lightscore（POST）
    這支API是Laravel框架，需要先GET一次網頁：
      1) 建立 session（cookie，requests.Session()自動處理）
      2) 從網頁HTML裡的 <meta name="csrf-token" content="..."> 標籤抓出CSRF token
         （這個token不是cookie，是頁面渲染時嵌進HTML的，跟cookie裡的XSRF-TOKEN是兩回事）
      3) 把這個token放進 X-Csrf-Token 標頭去POST，才能通過驗證
    回傳 {"date":..., "score":..., "light":...} 或 None（找不到/解析失敗時）。
    """
    if skip:
        log.info("STEP 1d 跳過（--skip-cycle）")
        return None
    try:
        import requests
        import re

        log.info("STEP 1d: 更新景氣對策信號...")

        ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36"
        sess = requests.Session()
        sess.headers.update({"User-Agent": ua})

        # 1) 先GET一次頁面：順便建立session cookie，也從HTML裡抓 <meta name="csrf-token"> 的值
        page_resp = sess.get("https://index.ndc.gov.tw/n/zh_tw/lightscore", timeout=15)
        page_resp.raise_for_status()

        m = re.search(r'name=["\']csrf-token["\']\s+content=["\']([^"\']+)["\']', page_resp.text)
        if not m:
            log.warning("STEP 1d: 頁面HTML裡找不到 csrf-token meta標籤，網站可能改版了")
            return None
        csrf_token = m.group(1)

        # 2) 帶著同一個session的cookie + X-Csrf-Token標頭去POST，跟瀏覽器實際送出的請求一致
        api_headers = {
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest",
            "X-Csrf-Token": csrf_token,
            "Referer": "https://index.ndc.gov.tw/n/zh_tw",
            "Origin": "https://index.ndc.gov.tw",
        }
        resp = sess.post(
            "https://index.ndc.gov.tw/n/json/lightscore",
            headers=api_headers, timeout=15,
        )
        resp.raise_for_status()
        payload = resp.json()
        line = payload.get("line") or []
        if not line:
            log.warning("STEP 1d: lightscore API 回傳沒有 line 資料，可能改版了。回應開頭200字：%s",
                        str(payload)[:200])
            return None

        latest = line[-1]
        score = int(latest["y"])
        raw_date = str(latest["x"])  # 格式 YYYYMM，例如 "202512"
        date_str = f"{raw_date[:4]}-{raw_date[4:]}" if len(raw_date) == 6 else raw_date

        if score >= 38:
            light = "紅燈"
        elif score >= 32:
            light = "黃紅燈"
        elif score >= 23:
            light = "綠燈"
        elif score >= 17:
            light = "黃藍燈"
        else:
            light = "藍燈"

        # 連續同燈號月數：從最新一筆往回數，用同一批歷史分數換算燈號，
        # 遇到不同燈號就停止（不用額外存檔案，反正這支API本身就回傳整段歷史分數）
        streak = 0
        for item in reversed(line):
            s = int(item["y"])
            if s >= 38: lgt = "紅燈"
            elif s >= 32: lgt = "黃紅燈"
            elif s >= 23: lgt = "綠燈"
            elif s >= 17: lgt = "黃藍燈"
            else: lgt = "藍燈"
            if lgt == light:
                streak += 1
            else:
                break

        result = {"date": date_str, "score": score, "light": light, "streak": streak}
        log.info("STEP 1d 完成：景氣對策信號 %s", result)
        return result
    except Exception as e:
        log.warning("STEP 1d 景氣對策信號更新失敗（不影響主流程）：%s", e)
        return None


def step1c_margin_update(skip=False):
    """抓當日全市場融資餘額，計算較前一日的變化比例。
    回傳 {"date":..., "today":..., "prev":..., "change_amt":..., "change_pct":...} 或 None"""
    if skip:
        log.info("STEP 1c 跳過（--skip-margin）")
        return None
    try:
        import requests
        log.info("STEP 1c: 更新融資餘額...")
        ds = date.today().strftime("%Y%m%d")
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://www.twse.com.tw/zh/trading/margin/mi-margn.html",
            "Accept": "application/json, text/plain, */*",
        }
        # 舊網址 /exchangeReport/MI_MARGN 這份報表證交所那邊已經失效了（直接用瀏覽器打開也是
        # 「沒有符合條件的資料」，不是我們程式的問題），改用跟 Step1e 逐股資料同一個新版網址
        # /rwd/zh/afterTrading/MI_MARGN，selectType 改成 MS（市場總表）。
        # 保留舊網址當退路，萬一哪天新網址也出狀況，還有機會透過舊網址撈到資料。
        urls_to_try = [
            "https://www.twse.com.tw/rwd/zh/afterTrading/MI_MARGN",
            "https://www.twse.com.tw/exchangeReport/MI_MARGN",
        ]
        data, used_url = None, None
        for url in urls_to_try:
            r = requests.get(url, params={"response": "json", "date": ds, "selectType": "MS"},
                              headers=headers, timeout=20, verify=False)
            log.info("  [融資] %s  HTTP狀態=%s  Content-Type=%s", url, r.status_code, r.headers.get("Content-Type"))
            try:
                d = r.json()
            except Exception as je:
                log.warning("STEP 1c: %s 回傳內容不是有效JSON（%s），原始內容前300字：%s", url, je, r.text[:300])
                continue
            if d.get("stat") == "OK":
                data, used_url = d, url
                break
            log.warning("STEP 1c: %s 回傳 stat=%s，換下一個網址試試", url, d.get("stat"))

        if data is None:
            log.warning("STEP 1c: 兩個網址都抓不到今日融資餘額（今日可能是假日或資料尚未更新）")
            return None
        log.info("  [融資] 這次成功的網址：%s", used_url)

        fields = data.get("fields", [])
        rows = data.get("data", [])

        # 有些 TWSE 報表用「fields/data」封裝，有些用「tables」(多表格)封裝，
        # 先試 fields/data，抓不到東西再試 tables 裡有沒有符合的表格
        if not rows and data.get("tables"):
            for tbl in data["tables"]:
                tbl_fields = tbl.get("fields", [])
                tbl_data = tbl.get("data", [])
                if any(row and "融資" in str(row[0]) and "交易單位" in str(row[0]) for row in tbl_data):
                    fields, rows = tbl_fields, tbl_data
                    break

        # 找「融資(交易單位)」那一列，欄位依序大致是：買進/賣出/現金償還/前日餘額/今日餘額/限額
        target_row = None
        for row in rows:
            if row and "融資" in str(row[0]) and "交易單位" in str(row[0]):
                target_row = row
                break
        if target_row is None:
            log.warning("STEP 1c: 找不到融資(交易單位)列。頂層keys=%s，完整原始回傳(前1500字)：%s",
                        list(data.keys()), r.text[:1500])
            return None

        def _num(s):
            try: return int(str(s).replace(",", ""))
            except Exception: return None

        prev_idx = fields.index("前日餘額") if "前日餘額" in fields else 4
        today_idx = fields.index("今日餘額") if "今日餘額" in fields else 5
        prev_bal  = _num(target_row[prev_idx])
        today_bal = _num(target_row[today_idx])

        if prev_bal is None or today_bal is None or prev_bal == 0:
            log.warning("STEP 1c: 融資餘額數值解析失敗，原始列：%s", target_row)
            return None

        # 順便抓「融資(金額)」那一列的今日餘額（總金額，單位元），給 step1e 估算維持率用，
        # 這樣不用為了拿這一個數字再多打一次API
        margin_amount_row = None
        for row in rows:
            if row and "融資" in str(row[0]) and "金額" in str(row[0]):
                margin_amount_row = row
                break
        margin_amount_today = None
        if margin_amount_row is not None:
            margin_amount_today = _num(margin_amount_row[today_idx]) if today_idx < len(margin_amount_row) else None
            # 證交所「融資(金額)」這一列的單位是「仟元」，不是「元」，要乘1000才能跟
            # step1e 算出來的擔保品市值（股數×股價，單位是元）對齊，不然算出來的維持率會差1000倍
            if margin_amount_today is not None:
                margin_amount_today *= 1000

        change_amt = today_bal - prev_bal
        change_pct = round(change_amt / prev_bal * 100, 3)

        log.info("STEP 1c 完成：融資今日餘額 %s 張，較前日 %+d 張 (%+.2f%%)",
                  f"{today_bal:,}", change_amt, change_pct)

        return {
            "date": date.today().strftime("%Y-%m-%d"),
            "today": today_bal, "prev": prev_bal,
            "change_amt": change_amt, "change_pct": change_pct,
            "margin_amount_today": margin_amount_today,
        }
    except Exception as e:
        log.error("STEP 1c 失敗：%s", e)
        return None


def step1e_margin_maintenance_ratio(margin_amount_today, skip=False):
    """估算全市場融資維持率。這不是證交所每日直接公布的數字（那個只在股市大跌時才會發
    新聞稿臨時公布一次），這裡是用跟 FinLab 等網站相同的估算方式自己算出來的：
        維持率(%) ≈ 全市場擔保品市值 / 全市場融資金額餘額 × 100
        擔保品市值 ≈ Σ(每檔股票 融資今日餘額(張) × 1000 × 今日收盤價)
    融資今日餘額(逐股)來自 MI_MARGN 的 selectType=ALL，收盤價來自今天已經抓好的 market_flow_all.csv。
    margin_amount_today：全市場融資金額餘額（元），從 step1c_margin_update() 回傳的
    margin_amount_today 傳進來，不用再多打一次 API。

    回傳 {"date":..., "ratio":..., "collateral_value":..., "margin_amount":..., "matched_stocks":...}
    或 None（缺資料/解析失敗時）。這是估算值，不是官方數字，儀表板顯示時要註明。
    """
    if skip:
        log.info("STEP 1e 跳過（--skip-maintenance）")
        return None
    if not margin_amount_today:
        log.warning("STEP 1e 跳過：沒有可用的融資金額餘額（step1c 可能失敗或今天沒有資料）")
        return None
    try:
        import requests
        log.info("STEP 1e: 估算融資維持率...")
        ds = date.today().strftime("%Y%m%d")
        r = requests.get(
            "https://www.twse.com.tw/exchangeReport/MI_MARGN",
            params={"response": "json", "date": ds, "selectType": "ALL"},
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Referer": "https://www.twse.com.tw/zh/trading/margin/mi-margn.html",
                "Accept": "application/json, text/plain, */*",
            },
            timeout=30, verify=False,
        )
        try:
            data = r.json()
        except Exception as je:
            log.warning("STEP 1e: 回傳內容不是有效JSON（%s），原始內容前300字：%s", je, r.text[:300])
            return None
        if data.get("stat") != "OK":
            log.warning("STEP 1e: 逐股融資資料尚未更新（stat=%s）", data.get("stat"))
            return None

        fields = data.get("fields", [])
        rows = data.get("data", [])
        if not rows and data.get("tables"):
            for tbl in data["tables"]:
                if tbl.get("fields") and any("代號" in f for f in tbl["fields"]):
                    fields, rows = tbl["fields"], tbl["data"]
                    break
        if not rows:
            log.warning("STEP 1e: 逐股融資資料是空的。頂層keys=%s", list(data.keys()))
            return None

        code_idx = next((i for i, f in enumerate(fields) if "代號" in f), 0)
        # 實際欄位長這樣（融資/融券兩組欄位共用同樣的名稱，沒有「融資」/「融券」前綴，
        # 只能靠出現順序分辨）：
        #   代號,名稱, 買進,賣出,現金償還,前日餘額,今日餘額,次一營業日限額,  ← 這組是融資
        #   買進,賣出,現金償還,前日餘額,今日餘額,次一營業日限額,  ← 這組是融券
        #   資券互抵,註記
        # 融資的「今日餘額」一定排在融券前面，用 index() 取第一個符合的就是融資那欄
        bal_idx = fields.index("今日餘額") if "今日餘額" in fields else None
        if bal_idx is None:
            log.warning("STEP 1e: 找不到「融資今日餘額」欄位，實際欄位：%s", fields)
            return None

        def _num(s):
            try: return int(str(s).replace(",", ""))
            except Exception: return None

        margin_lots = {}
        for row in rows:
            if len(row) <= max(code_idx, bal_idx):
                continue
            code = str(row[code_idx]).strip()
            lots = _num(row[bal_idx])
            if code and lots:
                margin_lots[code] = lots

        if not margin_lots:
            log.warning("STEP 1e: 逐股融資今日餘額全部解析失敗，範例列：%s", rows[:2])
            return None

        today_str = date.today().strftime("%Y-%m-%d")

        # 收盤價：不透過 market_flow_all.csv（那份存檔時故意濾掉了5碼的ETF），
        # 改直接呼叫同一支 MI_INDEX API，但這裡逐一掃過「每一個」回傳的表格去收集代號:收盤價，
        # 不像 _mf_fetch_price 只取第一個符合的表格就停（TWSE type=ALL 有可能把ETF、特別股
        # 這些放在不同表格裡，只抓第一個表格會漏掉）
        price_map = {}
        try:
            pdata = _mf_api_get("https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX",
                                 {"response": "json", "date": date.today().strftime("%Y%m%d"), "type": "ALL"})
            for tbl in pdata.get("tables", []):
                f = tbl.get("fields", [])
                cc = next((c for c in f if "代號" in c), None)
                clsc = next((c for c in f if "收盤" in c), None)
                if not (cc and clsc):
                    continue
                tdf = pd.DataFrame(tbl.get("data", []), columns=f)
                tdf.columns = tdf.columns.str.strip()
                for code, price in zip(tdf[cc].astype(str).str.strip(), tdf[clsc].apply(_mf_cl)):
                    if code and price:
                        price_map[code] = price
        except Exception as pe:
            log.warning("STEP 1e: 抓收盤價失敗（%s），改用 market_flow_all.csv 當備援（會漏掉ETF）", pe)

        if not price_map:
            # 備援：上面失敗的話，退回用今天已經存好的 market_flow_all.csv（會漏掉ETF，但至少有數字）
            if not MARKET_CSV.exists():
                log.warning("STEP 1e: 找不到收盤價可用（API失敗，market_flow_all.csv也不存在）")
                return None
            mf = pd.read_csv(MARKET_CSV, dtype=str)
            mf_today = mf[mf["date"] == today_str]
            if mf_today.empty:
                log.warning("STEP 1e: market_flow_all.csv 裡沒有今天(%s)的資料，可能 Step1 還沒跑過", today_str)
                return None
            price_map = dict(zip(mf_today["code"], pd.to_numeric(mf_today["price"], errors="coerce")))

        collateral_value = 0.0
        matched = 0
        unmatched_codes = []
        for code, lots in margin_lots.items():
            price = price_map.get(code)
            if price and price > 0:
                collateral_value += lots * 1000 * price
                matched += 1
            else:
                unmatched_codes.append(code)

        if matched == 0 or collateral_value <= 0:
            log.warning("STEP 1e: 逐股融資餘額跟收盤價對不起來，matched=0（代號格式可能不一致）")
            return None

        ratio = round(collateral_value / margin_amount_today * 100, 2)

        # 對不上價格的股票裡，扣掉的融資餘額(張)有多少，讓對不上的比例跟量體都看得到，
        # 判斷是不是主要都是ETF(5碼)/特別股(有字母)這類本來就沒進market_flow_all.csv的標的
        unmatched_lots = sum(margin_lots[c] for c in unmatched_codes)
        total_lots = sum(margin_lots.values())
        n_etf_like = sum(1 for c in unmatched_codes if len(c) != 4 or not c.isdigit())
        log.info("STEP 1e 完成：估算全市場融資維持率 %.2f%%（擔保品市值約%.0f億，融資金額約%.0f億，"
                  "%d/%d檔股票對上收盤價，缺漏%d檔佔融資張數%.1f%%，其中%d檔代號非4碼數字-可能是ETF/特別股）",
                  ratio, collateral_value / 1e8, margin_amount_today / 1e8, matched, len(margin_lots),
                  len(unmatched_codes), unmatched_lots / total_lots * 100 if total_lots else 0, n_etf_like)
        if unmatched_codes:
            log.info("STEP 1e   對不上價格的代號範例（前20個）：%s", unmatched_codes[:20])

        return {
            "date": today_str,
            "ratio": ratio,
            "collateral_value": round(collateral_value, 0),
            "margin_amount": margin_amount_today,
            "matched_stocks": matched,
        }
    except Exception as e:
        log.warning("STEP 1e 融資維持率估算失敗（不影響主流程）：%s", e)
        return None


# ════════════════════════════════════════════════════════════
# STEP 1g：恐慌貪婪指數（自製版，仿CNN Fear & Greed Index）
# ════════════════════════════════════════════════════════════
FG_COMPONENTS = ["margin_ratio", "pivot_pct", "inst_5d", "breadth_ratio", "margin_5d_chg", "twii_chg"]
FG_MIN_DAYS_FOR_PERCENTILE = 60   # 累積滿幾個交易日之後，改用百分位排名（之前用固定門檻頂著）
FG_LOOKBACK_DAYS = 60             # 百分位排名回看的天數

# 資料不夠60天時的固定門檻備援（後面數字對應0分，前面數字對應100分，順序決定分數方向）
FG_FALLBACK_RANGES = {
    "margin_ratio":    (120, 200),   # 融資維持率 120%→0分, 200%→100分
    "pivot_pct":       (-10, 10),    # (收盤-轉折點)/轉折點% ：-10%→0分, +10%→100分
    "inst_5d":         (-500, 500),  # 三大法人近5日累計買賣超(萬元) ：-500萬→0分, +500萬→100分
    "margin_5d_chg":   (-5, 5),      # 融資餘額近5日變化% ：-5%→0分, +5%→100分
    "twii_chg":        (-3, 3),      # 大盤單日漲跌% ：-3%→0分, +3%→100分
    # breadth_ratio 天生就是0~100，不需要fallback範圍
}


def _fg_score_from_fallback(value, key):
    lo, hi = FG_FALLBACK_RANGES[key]
    if value is None:
        return None
    score = (value - lo) / (hi - lo) * 100
    return round(max(0, min(100, score)), 1)


def _fg_score_from_percentile(value, history_series):
    """算 value 在 history_series（含今天）裡排第幾百分位，當作0~100分"""
    if value is None or len(history_series) == 0:
        return None
    n_below = (history_series < value).sum()
    n_equal = (history_series == value).sum()
    pct = (n_below + 0.5 * n_equal) / len(history_series) * 100
    return round(pct, 1)


def step1g_fear_greed_index(market_pivot, margin_data, margin_maintenance, inst_summary, skip=False):
    """自製版恐慌貪婪指數，六項指標等權重平均：
      1. 融資維持率（估算值）
      2. 大盤位置：(收盤-多空轉折點)/轉折點%
      3. 三大法人買賣超：近5個交易日累計金額
      4. 漲跌家數比：上漲家數/(上漲+下跌)*100，天生就是0~100不用額外正規化
      5. 融資餘額變化率：近5個交易日變化%
      6. 大盤單日漲跌幅
    每項先看有沒有累積滿60個交易日的歷史（存在 fear_greed_history.csv），
    滿了就用「今天這個數字在過去60天的百分位排名」當分數（跟CNN Fear&Greed Index同樣做法），
    還沒滿60天的項目，先用一組固定門檻線性內插頂著，累積夠了會自動切換，不用改設定。
    回傳 {"date":..., "score":..., "label":..., "components": {...每項的原始值跟分數...}} 或 None
    """
    if skip:
        log.info("STEP 1g 跳過（--skip-feargreed）")
        return None
    try:
        today_str = date.today().strftime("%Y-%m-%d")

        # 1) 從各來源湊出今天的6個原始數值，缺的就是 None（該項今天先不計分）
        raw = {k: None for k in FG_COMPONENTS}

        mm = margin_maintenance or {}
        raw["margin_ratio"] = mm.get("ratio")

        mp = market_pivot or {}
        daily = mp.get("daily") or []
        real = [d for d in daily if d.get("open") is not None]
        if real:
            last = real[-1]
            pivot = last.get("pivot")
            close = last.get("close")
            if pivot and close:
                raw["pivot_pct"] = round((close - pivot) / pivot * 100, 2)
            chg = last.get("change")
            base = (close - chg) if (close is not None and chg is not None) else None
            if base:
                raw["twii_chg"] = round(chg / base * 100, 2)

        s = inst_summary or {}
        inst_total_today = s.get("inst_total")
        bd = s.get("breadth") or {}
        if bd.get("total"):
            raw["breadth_ratio"] = round(bd.get("up", 0) / bd["total"] * 100, 1)

        md = margin_data or {}
        margin_today = md.get("today")

        # 2) 讀歷史檔，把今天的原始數值（inst_total、margin_today 這兩個用來算5日滾動）先記下來，
        #    再用歷史檔算出 inst_5d（近5日法人買賣超累計）跟 margin_5d_chg（近5日融資餘額變化%）
        if FEAR_GREED_CSV.exists():
            hist = pd.read_csv(FEAR_GREED_CSV, dtype={"date": str})
        else:
            hist = pd.DataFrame(columns=["date", "margin_ratio", "pivot_pct", "twii_chg",
                                          "breadth_ratio", "inst_total_raw", "margin_today_raw",
                                          "inst_5d", "margin_5d_chg"])
        hist = hist[hist["date"] != today_str]  # 避免同一天重複寫入

        recent_inst = pd.to_numeric(hist["inst_total_raw"], errors="coerce").dropna().tolist()
        if inst_total_today is not None:
            recent_inst_5 = (recent_inst[-4:] + [inst_total_today]) if len(recent_inst) >= 4 else ([inst_total_today] + recent_inst)
            raw["inst_5d"] = round(sum(recent_inst_5[-5:]) / 10000, 1)  # 換算成「萬元」單位

        recent_margin = pd.to_numeric(hist["margin_today_raw"], errors="coerce").dropna().tolist()
        if margin_today is not None and len(recent_margin) >= 4:
            base_margin = recent_margin[-4]  # 5個交易日前（今天+前4筆裡最舊的那筆）的餘額
            if base_margin:
                raw["margin_5d_chg"] = round((margin_today - base_margin) / base_margin * 100, 2)

        # 3) 把今天的原始值＋算好的5日滾動值一起存回歷史檔（下次算5日滾動、百分位排名都要用到）
        new_row = pd.DataFrame([{
            "date": today_str,
            "margin_ratio": raw["margin_ratio"],
            "pivot_pct": raw["pivot_pct"],
            "twii_chg": raw["twii_chg"],
            "breadth_ratio": raw["breadth_ratio"],
            "inst_total_raw": inst_total_today,
            "margin_today_raw": margin_today,
            "inst_5d": raw["inst_5d"],
            "margin_5d_chg": raw["margin_5d_chg"],
        }])
        hist = pd.concat([hist, new_row], ignore_index=True).sort_values("date")
        hist.to_csv(FEAR_GREED_CSV, index=False, encoding="utf-8-sig")

        # 4) 逐項計分：優先用百分位排名（近60個交易日，含今天），資料不夠才用固定門檻備援
        scores = {}
        lookback = hist.tail(FG_LOOKBACK_DAYS)
        for key in FG_COMPONENTS:
            val = raw[key]
            if key == "breadth_ratio":
                scores[key] = val  # 天生0~100，直接當分數
                continue
            col_series = pd.to_numeric(lookback[key], errors="coerce").dropna()
            if len(col_series) >= FG_MIN_DAYS_FOR_PERCENTILE and val is not None:
                scores[key] = _fg_score_from_percentile(val, col_series)
            else:
                scores[key] = _fg_score_from_fallback(val, key)

        valid_scores = [v for v in scores.values() if v is not None]
        if not valid_scores:
            log.warning("STEP 1g: 六項指標今天全部缺資料，無法計算恐慌貪婪指數")
            return None
        overall = round(sum(valid_scores) / len(valid_scores), 1)

        if overall < 20: label = "極度恐慌"
        elif overall < 40: label = "恐慌"
        elif overall < 60: label = "中性"
        elif overall < 80: label = "貪婪"
        else: label = "極度貪婪"

        log.info("STEP 1g 完成：恐慌貪婪指數 %.1f 分（%s），組成：%s，計分用%d項（缺%d項）",
                  overall, label, {k: scores[k] for k in FG_COMPONENTS}, len(valid_scores),
                  len(FG_COMPONENTS) - len(valid_scores))

        return {
            "date": today_str,
            "score": overall,
            "label": label,
            "components": {k: {"raw": raw[k], "score": scores.get(k)} for k in FG_COMPONENTS},
            "history_days": len(hist),
        }
    except Exception as e:
        log.warning("STEP 1g 恐慌貪婪指數計算失敗（不影響主流程）：%s", e)
        return None


def step1h_vix(skip=False):
    """VIX恐慌指數（CBOE，美股波動率），單純當外部參考指標，不納入 fear_greed 計分。
    抓取邏輯在 vix_fetcher.py（yfinance ^VIX，本地快取30分鐘）。
    回傳 {"value":..., "change":..., "change_pct":..., "level":..., "date":..., "source":...} 或 None
    """
    if skip:
        log.info("STEP 1h 跳過（--skip-vix）")
        return None
    if get_vix is None:
        log.warning("STEP 1h: 找不到 vix_fetcher.py，跳過（不影響主流程）")
        return None
    try:
        data = get_vix()
        if data.get("value") is None:
            log.warning("STEP 1h: VIX 抓取失敗（%s）", data.get("error", "未知原因"))
            return data
        log.info("STEP 1h 完成：VIX %.2f（%+.2f，%s）",
                  data["value"], data.get("change") or 0, data.get("level"))
        return data
    except Exception as e:
        log.warning("STEP 1h VIX 抓取失敗（不影響主流程）：%s", e)
        return None


def step1i_twvix(skip=False):
    """台指VIX（VIXTWN，台灣期交所編製，經FinMind TaiwanOptionVix資料集），台股原生的恐慌指標，
    跟Step 1h的美股VIX是兩個獨立指標，都當外部參考，不納入 fear_greed 計分。
    回傳 {"value":..., "change":..., "change_pct":..., "level":..., "date":..., "source":...} 或 None
    """
    if skip:
        log.info("STEP 1i 跳過（--skip-twvix）")
        return None
    if get_twvix is None:
        log.warning("STEP 1i: 找不到 vix_fetcher.py，跳過（不影響主流程）")
        return None
    try:
        data = get_twvix()
        if data.get("value") is None:
            log.warning("STEP 1i: 台指VIX 抓取失敗（%s）", data.get("error", "未知原因"))
            return data
        log.info("STEP 1i 完成：台指VIX %.2f（%+.2f，%s）",
                  data["value"], data.get("change") or 0, data.get("level"))
        return data
    except Exception as e:
        log.warning("STEP 1i 台指VIX 抓取失敗（不影響主流程）：%s", e)
        return None


def step1j_market_pe_river(skip=False):
    """大盤本益比河流：讀取 market_pe_history.csv（先用 build_market_pe_history.py
    一次性回補歷史），今天的值用同一套「成交金額加權」邏輯即時算一次，如果CSV裡還
    沒有今天的資料就自動append進去——所以歷史會每天自動累加一筆，不用手動維護。
    回傳 {"history":[{"date":"YYYY-MM-DD","pe":...}, ...], "latest_pe":..., "latest_date":...}
    或 None（連CSV都沒有、也抓不到今天資料時）
    """
    if skip:
        log.info("STEP 1j 跳過（--skip-pe-river）")
        return None
    try:
        from build_market_pe_history import fetch_bwibbu_by_date, fetch_trade_value_by_date, compute_weighted_pe
    except ImportError:
        log.warning("STEP 1j: 找不到 build_market_pe_history.py，跳過（不影響主流程）")
        return None

    import csv as csv_mod
    csv_path = Path("market_pe_history.csv")
    history = []
    if csv_path.exists():
        with open(csv_path, encoding="utf-8") as f:
            for row in csv_mod.DictReader(f):
                try:
                    history.append({"date": row["date"], "pe": float(row["market_pe"])})
                except (KeyError, ValueError, TypeError):
                    continue

    today_str = date.today().strftime("%Y%m%d")
    already_today = any(h["date"] == today_str for h in history)

    if not already_today:
        try:
            per_map = fetch_bwibbu_by_date(today_str)
            value_map = fetch_trade_value_by_date(today_str) if per_map else None
            if per_map and value_map:
                pe_result = compute_weighted_pe(per_map, value_map)
                if pe_result:
                    pe, used = pe_result
                    history.append({"date": today_str, "pe": pe})
                    history.sort(key=lambda h: h["date"])
                    with open(csv_path, "w", encoding="utf-8", newline="") as f:
                        writer = csv_mod.writer(f)
                        writer.writerow(["date", "market_pe"])
                        for h in history:
                            writer.writerow([h["date"], h["pe"]])
                    log.info("STEP 1j 完成：今天大盤本益比 %.2f（%d檔），累積 %d 筆歷史", pe, used, len(history))
                else:
                    log.info("STEP 1j: 今天樣本不足，沿用既有 %d 筆歷史", len(history))
            else:
                log.info("STEP 1j: 今天抓不到資料（可能非交易日），沿用既有 %d 筆歷史", len(history))
        except Exception as e:
            log.warning("STEP 1j 抓取今天本益比失敗（不影響主流程）：%s", e)
    else:
        log.info("STEP 1j: 今天已有資料，共 %d 筆歷史", len(history))

    if not history:
        return None

    formatted = [{"date": f"{h['date'][:4]}-{h['date'][4:6]}-{h['date'][6:]}", "pe": h["pe"]} for h in history]
    return {"history": formatted, "latest_pe": formatted[-1]["pe"], "latest_date": formatted[-1]["date"]}


# ════════════════════════════════════════════════════════════
# STEP 2：資金流向分析（analyze_local.py 核心邏輯）
# ════════════════════════════════════════════════════════════

def step2_analysis() -> dict:
    """讀取 market_flow_all.csv，計算月度/產業/個股資金流向"""
    if not MARKET_CSV.exists():
        log.warning("STEP 2: %s 不存在，跳過", MARKET_CSV)
        return {}

    log.info("STEP 2: 資金流向分析...")
    df = pd.read_csv(MARKET_CSV, dtype=str)
    for c in ['foreign_net','trust_net','dealer_net','inst_total']:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    df['is_etf']   = df.get('is_etf', pd.Series('false', index=df.index)).str.lower() == 'true'
    df['is_stock'] = df['code'].str.match(r'^\d{4}$')
    df['date']     = pd.to_datetime(df['date'])
    df['month']    = df['date'].dt.strftime('%Y-%m')
    df['quarter']  = df['date'].dt.to_period('Q').astype(str)
    df['year']     = df['date'].dt.year

    s = df[df['is_stock'] & ~df['is_etf']].copy()
    months_list = sorted(s['month'].unique().tolist())
    result = {}

    result['summary'] = {
        'date_min':          s['date'].min().strftime('%Y-%m-%d'),
        'date_max':          s['date'].max().strftime('%Y-%m-%d'),
        'trading_days':      int(s['date'].nunique()),
        'stocks':            int(s['code'].nunique()),
        'total_foreign_wan': round(float(s['foreign_net'].sum()/10000), 1),
        'total_trust_wan':   round(float(s['trust_net'].sum()/10000),   1),
    }

    monthly = s.groupby('month').agg(
        foreign=('foreign_net','sum'), trust=('trust_net','sum'),
        dealer=('dealer_net','sum')).reset_index()
    monthly['foreign_w'] = (monthly['foreign']/10000).round(1)
    monthly['trust_w']   = (monthly['trust']  /10000).round(1)
    monthly['dealer_w']  = (monthly['dealer'] /10000).round(1)
    result['monthly'] = monthly[['month','foreign_w','trust_w','dealer_w']].to_dict(orient='records')

    ne_k = s[s.get('industry', pd.Series('其他', index=s.index)).fillna('其他') != '其他'] if 'industry' in s.columns else s
    if 'industry' in s.columns and not ne_k.empty:
        top_inds = ne_k.groupby('industry')['foreign_net'].sum().abs().nlargest(15).index.tolist()
        im = ne_k[ne_k['industry'].isin(top_inds)].groupby(
            ['month','industry'])[['foreign_net','trust_net']].sum().reset_index()
        im_f = im.pivot(index='month', columns='industry', values='foreign_net').reindex(months_list).fillna(0) / 10000
        im_t = im.pivot(index='month', columns='industry', values='trust_net').reindex(months_list).fillna(0)  / 10000
        result['industry_monthly'] = {
            'months': months_list, 'industries': top_inds,
            'foreign': {ind: im_f[ind].round(2).tolist() if ind in im_f.columns else [0]*len(months_list) for ind in top_inds},
            'trust':   {ind: im_t[ind].round(2).tolist() if ind in im_t.columns else [0]*len(months_list) for ind in top_inds},
        }
        # 外資進場訊號
        signals = {}
        for ind in top_inds:
            vals = im_f[ind].tolist() if ind in im_f.columns else []
            sig  = []
            for i in range(1, len(vals)):
                prev = vals[i-1]
                curr = vals[i]
                # 規則1：上個月 < 0，這個月反轉且差異 ≥ 30萬張
                if prev < 0 and curr > 0 and (curr - prev) >= 30:
                    sig.append({
                        'month': months_list[i],
                        'prev': round(prev, 1),
                        'curr': round(curr, 1),
                        'swing': round(curr - prev, 1),
                        'rule': '反轉+30萬'
                    })
                # 規則2：上個月 > 0，這個月持續買超且成長 ≥ 50%
                elif prev > 0 and curr > prev * 1.5:
                    sig.append({
                        'month': months_list[i],
                        'prev': round(prev, 1),
                        'curr': round(curr, 1),
                        'swing': round(curr - prev, 1),
                        'rule': '動能加速+50%'
                    })
            signals[ind] = sig
        result['entry_signals'] = signals

        # 投信進場訊號
        trust_signals = {}
        for ind in top_inds:
            vals = im_t[ind].tolist() if ind in im_t.columns else []
            sig  = []
            for i in range(1, len(vals)):
                prev = vals[i-1]
                curr = vals[i]
                # 規則1：上個月 < 0，這個月反轉且差異 ≥ 10萬張
                if prev < 0 and curr > 0 and (curr - prev) >= 10:
                    sig.append({
                        'month': months_list[i],
                        'prev': round(prev, 1),
                        'curr': round(curr, 1),
                        'swing': round(curr - prev, 1),
                        'rule': '反轉+10萬'
                    })
                # 規則2：上個月 > 0，這個月持續買超且成長 ≥ 50%
                elif prev > 0 and curr > prev * 1.5:
                    sig.append({
                        'month': months_list[i],
                        'prev': round(prev, 1),
                        'curr': round(curr, 1),
                        'swing': round(curr - prev, 1),
                        'rule': '動能加速+50%'
                    })
            trust_signals[ind] = sig
        result['trust_signals'] = trust_signals

        # 跨年持續買超
        sy = s.groupby(['year','code']+(['name','industry'] if 'name' in s.columns else [])).agg(
            foreign=('foreign_net','sum')).reset_index()
        yr_buys = {}
        for yr in [2023,2024,2025]:
            if yr in sy['year'].values:
                yr_buys[str(yr)] = set(sy[(sy['year']==yr)&(sy['foreign']>30000)]['code'])
        consistent = []
        for code in set().union(*yr_buys.values()) if yr_buys else []:
            cnt = sum(1 for yr in yr_buys if code in yr_buys[yr])
            if cnt < 2: continue
            sub = sy[sy['code']==code]
            if sub.empty: continue
            yr_vals = {yr: round(float(sub[sub['year']==int(yr)]['foreign'].values[0])/10000,1)
                       if len(sub[sub['year']==int(yr)]) else 0 for yr in yr_buys}
            consistent.append({
                'code':     code,
                'name':     sub['name'].iloc[0] if 'name' in sub.columns else '',
                'industry': sub['industry'].iloc[0] if 'industry' in sub.columns else '其他',
                'years':    yr_vals,
                'total':    round(sum(yr_vals.values()),1),
                'count':    cnt
            })
        result['consistent_buy'] = sorted(consistent, key=lambda x: -x['total'])
        # 季度輪動
        qtrs = sorted(s['quarter'].unique().tolist())
        iq   = ne_k[ne_k['industry'].isin(top_inds)].groupby(
            ['quarter','industry'])['foreign_net'].sum().reset_index()
        iq_p = iq.pivot(index='quarter', columns='industry',
                        values='foreign_net').reindex(qtrs).fillna(0) / 10000
        result['industry_quarterly'] = {
            'quarters':   qtrs,
            'industries': top_inds,
            'data': {ind: iq_p[ind].round(1).tolist() if ind in iq_p.columns else [0]*len(qtrs) for ind in top_inds},
        }
    else:
        result.update({'industry_monthly': {}, 'entry_signals': {},
                       'consistent_buy': [],
                       'industry_quarterly': {'quarters': [], 'industries': [], 'data': {}}})

    # ── 個股年度排行 + 月度追蹤 ──────────────────────────
    if 'year' in s.columns and 'name' in s.columns:
        sy = s.groupby(['year','code','name'] + (['industry'] if 'industry' in s.columns else [])).agg(
            foreign=('foreign_net','sum'), trust=('trust_net','sum'),
            days=('date','nunique')).reset_index()
        sy['foreign_w'] = (sy['foreign']/10000).round(2)
        sy['trust_w']   = (sy['trust']  /10000).round(2)

        result['yearly_stocks'] = {}
        for yr in sy['year'].unique():
            ydf  = sy[sy['year']==yr]
            cols = [c for c in ['code','name','industry','foreign_w','trust_w','days'] if c in ydf.columns]
            result['yearly_stocks'][str(int(yr))] = {
                'buy':  ydf.nlargest(20,'foreign')[cols].to_dict(orient='records'),
                'sell': ydf.nsmallest(20,'foreign')[cols].to_dict(orient='records'),
            }

        # 外資前50 + 投信前50（聯集）
        top50_f = sy.groupby('code')['foreign'].sum().abs().nlargest(50).index.tolist()
        top50_t = sy.groupby('code')['trust'].sum().abs().nlargest(50).index.tolist()
        top_codes = list(dict.fromkeys(top50_f + top50_t))  # 保持順序去重

        sm_grp = s[s['code'].isin(top_codes)].groupby(
            ['month','code'] + (['name','industry'] if 'industry' in s.columns else ['name'])
        )[['foreign_net','trust_net']].sum().reset_index()
        sm_grp['f_w'] = (sm_grp['foreign_net']/10000).round(2)
        sm_grp['t_w'] = (sm_grp['trust_net']  /10000).round(2)

        sml = []
        for code3 in top_codes:
            sub3 = sm_grp[sm_grp['code']==code3].sort_values('month')
            if sub3.empty: continue
            f_piv = sub3.set_index('month')['f_w'].reindex(months_list).fillna(0)
            t_piv = sub3.set_index('month')['t_w'].reindex(months_list).fillna(0)
            # 年度投信合計
            sub_sy = sy[sy['code']==code3]
            trust_total = round(float(sub_sy['trust_w'].sum()), 1) if not sub_sy.empty else 0.0
            sml.append({
                'code':          code3,
                'name':          sub3['name'].iloc[0] if 'name' in sub3.columns else code3,
                'industry':      sub3['industry'].iloc[0] if 'industry' in sub3.columns else '其他',
                'total':         round(float(f_piv.sum()), 1),
                'monthly':       f_piv.round(2).tolist(),
                'trust_total':   trust_total,
                'trust_monthly': t_piv.round(2).tolist(),
            })
        sml.sort(key=lambda x: -abs(x['total']))
        result['stock_monthly'] = {'months': months_list, 'stocks': sml}
    else:
        result['yearly_stocks'] = {}
        result['stock_monthly'] = {'months': months_list if 'months_list' in dir() else [], 'stocks': []}

    log.info("STEP 2 完成：%s ~ %s", result['summary']['date_min'], result['summary']['date_max'])
    return result


# ════════════════════════════════════════════════════════════
# STEP 3：多時框雷達（stock_radar_multi.py 核心邏輯）
# ════════════════════════════════════════════════════════════

def step3_radar() -> dict:
    """讀取 CSV，計算 5/10/20 天多時框分析"""
    if not MARKET_CSV.exists():
        log.warning("STEP 3: %s 不存在，跳過", MARKET_CSV)
        return {}

    log.info("STEP 3: 多時框雷達分析...")
    WINDOWS = [5, 10, 20]

    df = pd.read_csv(MARKET_CSV, dtype=str)
    for c in ['foreign_net','trust_net','inst_total','change_pct','volume_k']:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    df['is_etf']   = df.get('is_etf', pd.Series('false', index=df.index)).str.lower() == 'true'
    df['is_stock'] = df['code'].str.match(r'^\d{4}$')
    df['date']     = pd.to_datetime(df['date'])

    mkt_ne = df[df['is_stock'] & ~df['is_etf']].copy()
    if mkt_ne.empty:
        return {}

    all_dates  = sorted(mkt_ne['date'].unique())
    today_str  = all_dates[-1].strftime('%Y-%m-%d')

    def analyze_window(days):
        recent  = all_dates[-days:]
        mkt_r   = mkt_ne[mkt_ne['date'].isin(recent)]

        ind_grp = mkt_r[mkt_r.get('industry', pd.Series('其他', index=mkt_r.index)).fillna('其他') != '其他'] if 'industry' in mkt_r.columns else mkt_r
        if 'industry' in ind_grp.columns:
            ind = ind_grp.groupby('industry').agg(
                foreign_w=('foreign_net', lambda x: round(x.sum()/10000, 2)),
                trust_w=('trust_net',   lambda x: round(x.sum()/10000, 2)),
            ).reset_index()
            # 連續買超天數
            def consec(grp):
                vals = grp.sort_values('date')['foreign_net'].values
                n = 0
                for v in reversed(vals):
                    if v > 0: n += 1
                    else: break
                return n
            ind_daily = ind_grp.groupby(['date','industry'])['foreign_net'].sum().reset_index()
            cc = ind_daily.groupby('industry').apply(consec).reset_index()
            cc.columns = ['industry','consec']
            ind = ind.merge(cc, on='industry', how='left')
            ind['consec'] = ind['consec'].fillna(0).astype(int)
            ind_list = ind.sort_values('foreign_w', ascending=False).to_dict(orient='records')
        else:
            ind_list = []

        # ── 外資個股（每日外資前50大買超）────────────────
        def top50_codes(grp):
            return grp.nlargest(50, 'foreign_net')['code']
        t50_codes = mkt_r.groupby('date').apply(top50_codes).reset_index(drop=True).unique()
        t50_r = mkt_r[mkt_r['code'].isin(t50_codes)]

        t50_s = t50_r.groupby('code').agg(
            name        =('name',       lambda x: x.mode()[0] if 'name' in x.index.names or True else ''),
            appear_days =('date',       'nunique'),
            avg_chg     =('change_pct', 'mean'),
            avg_vol_k   =('volume_k',   'mean'),
            t50_foreign =('foreign_net','sum'),
        ).reset_index()

        mkt_s = mkt_r.groupby('code').agg(
            industry    =('industry',   lambda x: x.mode()[0] if 'industry' in mkt_r.columns else '其他'),
            mkt_foreign =('foreign_net','sum'),
        ).reset_index() if 'industry' in mkt_r.columns else mkt_r.groupby('code')['foreign_net'].sum().reset_index().rename(columns={'foreign_net':'mkt_foreign'})
        mkt_s['mkt_f_w'] = (mkt_s['mkt_foreign']/10000).round(2)

        cross = t50_s.merge(mkt_s[['code']+(['industry'] if 'industry' in mkt_s.columns else [])+['mkt_f_w']], on='code', how='left')
        cross['mkt_f_w']  = cross['mkt_f_w'].fillna(0)
        cross['avg_chg']  = cross['avg_chg'].round(2)
        cross['avg_vol_k']= cross['avg_vol_k'].round(0)
        ma = cross['appear_days'].max() or 1
        mv = cross['avg_vol_k'].max()   or 1
        mf = cross['mkt_f_w'].clip(0).max() or 1
        cross['score'] = (cross['appear_days']/ma*30 + cross['avg_vol_k']/mv*20 +
                          cross['mkt_f_w'].clip(0)/mf*35 + cross['avg_chg'].clip(0)/5*15).round(1)
        cross = cross.sort_values('score', ascending=False).reset_index(drop=True)
        # 只保留當框累計外資買超 > 0
        cross = cross[cross['mkt_f_w'] > 0].reset_index(drop=True)

        # ── 投信個股（每日投信前50大買超）────────────────
        def top50_trust_codes(grp):
            return grp.nlargest(50, 'trust_net')['code']
        t50t_codes = mkt_r.groupby('date').apply(top50_trust_codes).reset_index(drop=True).unique()
        t50t_r = mkt_r[mkt_r['code'].isin(t50t_codes)]

        t50t_s = t50t_r.groupby('code').agg(
            name         =('name',      lambda x: x.mode()[0] if True else ''),
            appear_days  =('date',      'nunique'),
            avg_chg      =('change_pct','mean'),
            avg_vol_k    =('volume_k',  'mean'),
            t50_trust    =('trust_net', 'sum'),
        ).reset_index()

        mkt_st = mkt_r.groupby('code').agg(
            industry  =('industry',  lambda x: x.mode()[0] if 'industry' in mkt_r.columns else '其他'),
            mkt_trust =('trust_net', 'sum'),
        ).reset_index() if 'industry' in mkt_r.columns else mkt_r.groupby('code')['trust_net'].sum().reset_index().rename(columns={'trust_net':'mkt_trust'})
        mkt_st['mkt_t_w'] = (mkt_st['mkt_trust']/10000).round(2)

        tcross = t50t_s.merge(mkt_st[['code']+(['industry'] if 'industry' in mkt_st.columns else [])+['mkt_t_w']], on='code', how='left')
        tcross['mkt_t_w']  = tcross['mkt_t_w'].fillna(0)
        tcross['avg_chg']  = tcross['avg_chg'].round(2)
        tcross['avg_vol_k']= tcross['avg_vol_k'].round(0)
        ta = tcross['appear_days'].max() or 1
        tv = tcross['avg_vol_k'].max()   or 1
        tt = tcross['mkt_t_w'].clip(0).max() or 1
        tcross['trust_score'] = (tcross['appear_days']/ta*30 + tcross['avg_vol_k']/tv*20 +
                                  tcross['mkt_t_w'].clip(0)/tt*35 + tcross['avg_chg'].clip(0)/5*15).round(1)
        tcross = tcross.sort_values('trust_score', ascending=False).reset_index(drop=True)
        # 只保留當框累計投信買超 > 0
        tcross = tcross[tcross['mkt_t_w'] > 0].reset_index(drop=True)

        return {
            'industry':     ind_list,
            'stocks':       cross.head(20).to_dict(orient='records'),
            'trust_stocks': tcross.head(20).to_dict(orient='records'),
        }

    results = {}
    for w in WINDOWS:
        if len(all_dates) >= w:
            results[str(w)] = analyze_window(w)

    # 交叉比對
    def get_ind_val(w, ind, field='foreign_w'):
        for r in results.get(str(w),{}).get('industry',[]):
            if r.get('industry') == ind: return r.get(field, 0)
        return None

    def get_stk_score(w, code):
        for r in results.get(str(w),{}).get('stocks',[]):
            if r['code'] == code: return r.get('score',0)
        return None

    def get_trust_stk_score(w, code):
        for r in results.get(str(w),{}).get('trust_stocks',[]):
            if r['code'] == code: return r.get('trust_score',0)
        return None

    def get_trust_stk_val(w, code):
        for r in results.get(str(w),{}).get('trust_stocks',[]):
            if r['code'] == code: return r.get('mkt_t_w',0)
        return None

    def make_verdict(vals_pos, vals_neg, n_windows):
        pos = sum(1 for v in vals_pos if v is not None and v > 0)
        neg = sum(1 for v in vals_neg if v is not None and v < 0)
        return ('⭐ 三框一致看多' if pos==n_windows else
                '✅ 偏多' if pos==2 else
                '❌ 三框一致看空' if neg==n_windows else
                '⚠ 偏空' if neg==2 else '↔ 分歧')

    all_inds  = set(r['industry'] for w in WINDOWS for r in results.get(str(w),{}).get('industry',[]))
    all_codes = set(r['code']     for w in WINDOWS for r in results.get(str(w),{}).get('stocks',[]))

    ind_compare = []
    for ind in all_inds:
        row = {'industry': ind}
        f_vals, t_vals = [], []
        for w in WINDOWS:
            fv = get_ind_val(w, ind, 'foreign_w')
            tv = get_ind_val(w, ind, 'trust_w')
            row[f'd{w}']  = round(fv, 1) if fv is not None else None  # 外資
            row[f't{w}']  = round(tv, 1) if tv is not None else None  # 投信
            if fv is not None: f_vals.append(fv)
            if tv is not None: t_vals.append(tv)
        row['verdict']       = make_verdict([row.get(f'd{w}') for w in WINDOWS],
                                            [row.get(f'd{w}') for w in WINDOWS], len(WINDOWS))
        row['trust_verdict'] = make_verdict([row.get(f't{w}') for w in WINDOWS],
                                            [row.get(f't{w}') for w in WINDOWS], len(WINDOWS))
        ind_compare.append(row)
    ind_compare.sort(key=lambda x: (0 if '三框' in x['verdict'] and '多' in x['verdict'] else
                                     1 if '偏多' in x['verdict'] else
                                     2 if '分歧' in x['verdict'] else
                                     3 if '偏空' in x['verdict'] else 4, -(x.get('d5') or 0)))

    stk_compare = []
    for code in all_codes:
        info = next((r for w in WINDOWS for r in results.get(str(w),{}).get('stocks',[]) if r['code']==code), {})
        if not info: continue
        row = {'code': code, 'name': info.get('name',''), 'industry': info.get('industry','其他')}
        scores = []
        for w in WINDOWS:
            sc_val = get_stk_score(w, code)
            # 只有當框買超 > 0 才計入
            if sc_val is not None:
                row[f's{w}'] = round(sc_val, 0)
                scores.append(sc_val)
            else:
                row[f's{w}'] = None
        row['frames']    = sum(1 for w in WINDOWS if row.get(f's{w}') is not None)
        row['avg_score'] = round(sum(scores)/len(scores), 1) if scores else 0
        row['verdict']   = ('⭐ 三框都有' if row['frames']==3 else
                            '✅ 兩框出現' if row['frames']==2 else '— 單框')
        stk_compare.append(row)
    stk_compare.sort(key=lambda x: (-x['frames'], -x['avg_score']))

    # 投信個股交叉比對
    all_trust_codes = set(r['code'] for w in WINDOWS for r in results.get(str(w),{}).get('trust_stocks',[]))
    trust_stk_compare = []
    for code in all_trust_codes:
        info = next((r for w in WINDOWS for r in results.get(str(w),{}).get('trust_stocks',[]) if r['code']==code), {})
        if not info: continue
        row = {'code': code, 'name': info.get('name',''), 'industry': info.get('industry','其他')}
        scores = []
        for w in WINDOWS:
            sc_val = get_trust_stk_score(w, code)
            # 只有當框買超 > 0 才計入
            if sc_val is not None:
                row[f'ts{w}'] = round(sc_val, 0)
                scores.append(sc_val)
            else:
                row[f'ts{w}'] = None
        row['frames']    = sum(1 for w in WINDOWS if row.get(f'ts{w}') is not None)
        row['avg_score'] = round(sum(scores)/len(scores), 1) if scores else 0
        row['verdict']   = ('⭐ 三框都有' if row['frames']==3 else
                            '✅ 兩框出現' if row['frames']==2 else '— 單框')
        trust_stk_compare.append(row)
    trust_stk_compare.sort(key=lambda x: (-x['frames'], -x['avg_score']))

    log.info("STEP 3 完成：%d 個產業，%d 支個股（外資），%d 支個股（投信）",
             len(ind_compare), len(stk_compare), len(trust_stk_compare))
    return {
        'generated_at':       datetime.now().strftime('%Y-%m-%d %H:%M'),
        'today':              today_str,
        'windows':            WINDOWS,
        'results':            results,
        'ind_compare':        ind_compare,
        'stk_compare':        stk_compare,
        'trust_stk_compare':  trust_stk_compare,
    }


# ════════════════════════════════════════════════════════════
# STEP 4+5：共用 yfinance 下載，同時產出強勢股 + 創新高 + 連買MA
# ════════════════════════════════════════════════════════════

def _fetch_stock_list() -> pd.DataFrame:
    """
    取得全市場個股清單（含產業別）
    使用 curl_cffi 模擬瀏覽器，避免 TWSE 反爬蟲擋掉 requests
    """
    import io
    from curl_cffi import requests as cffi_req

    session = cffi_req.Session(impersonate="chrome124")
    session.verify = False
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Referer": "https://isin.twse.com.tw/",
        "Accept-Language": "zh-TW,zh;q=0.9",
    }

    frames = []
    for mode, suffix, market_label in [("2", ".TW", "TWSE"), ("4", ".TWO", "TPEx")]:
        try:
            url  = f"https://isin.twse.com.tw/isin/C_public.jsp?strMode={mode}"
            resp = session.get(url, headers=headers, timeout=30)
            resp.encoding = "big5hkscs"
            tables = pd.read_html(io.StringIO(resp.text))
            if not tables:
                log.warning("  %s 清單：HTML 無表格", market_label)
                continue
            df = tables[0]
            df.columns = df.iloc[0]
            df = df.iloc[1:].reset_index(drop=True)

            # 欄位名稱可能含空白，做 strip
            df.columns = [str(c).strip() for c in df.columns]

            # 找代號欄（只取 4 碼數字開頭的股票）
            code_col = "有價證券代號及名稱"
            if code_col not in df.columns:
                log.warning("  %s 清單：找不到代號欄，欄位為 %s", market_label, list(df.columns))
                continue

            # 嘗試取得產業欄
            ind_col = None
            for candidate in ["產業別", "產業", "Industry"]:
                if candidate in df.columns:
                    ind_col = candidate
                    break

            rows = []
            for _, row in df.iterrows():
                cell = str(row[code_col]).strip()
                if not cell or not cell[:4].isdigit() or len(cell) < 5:
                    continue
                parts = cell.split(None, 1)
                if len(parts) < 2:
                    continue
                code = parts[0]
                name = parts[1].strip()

                # 只保留純數字 4-5 碼的股票，排除權證/牛熊證（6碼、含英文字母）
                if not code.isdigit() or len(code) > 5:
                    continue

                industry = str(row[ind_col]).strip() if ind_col else "其他"
                if industry in ("nan", "", "—"):
                    industry = "其他"
                rows.append({
                    "code":     code,
                    "name":     name,
                    "market":   market_label,
                    "ticker":   code + suffix,
                    "industry": industry,
                })

            if rows:
                frames.append(pd.DataFrame(rows))
                log.info("  %s：%d 支（含產業資訊）", market_label, len(rows))
            else:
                log.warning("  %s 清單：解析後 0 筆", market_label)

        except Exception as e:
            log.warning("  %s 清單失敗：%s", market_label, e)

    if not frames:
        # 嘗試讀取本地快取
        cache_path = OUTPUT_DIR / "stock_list_cache.json"
        if cache_path.exists():
            try:
                cached = pd.read_json(cache_path, dtype={"code": str})
                if not cached.empty:
                    log.info("  TWSE 抓取失敗，使用本地快取（%d 支）", len(cached))
                    return cached
                else:
                    log.warning("  本地快取是空的，無法使用")
            except Exception as e:
                log.warning("  本地快取讀取失敗：%s", e)
        log.warning("  TWSE 抓取失敗且無有效本地快取，stock_list 為空")
        return pd.DataFrame(columns=["code", "name", "market", "ticker", "industry"])

    out = pd.concat(frames, ignore_index=True)
    out = out.drop_duplicates("code").reset_index(drop=True)

    # 成功時存快取
    cache_path = OUTPUT_DIR / "stock_list_cache.json"
    try:
        out.to_json(cache_path, orient="records", force_ascii=False)
        log.info("  股票清單快取已更新（%d 支）", len(out))
    except Exception as e:
        log.warning("  快取儲存失敗：%s", e)
    return out


def _fetch_shares_outstanding() -> dict:
    """
    取得發行股數（千股）{code: shares_k}
    優先讀本地快取 output/shares_cache.json（由 build_shares_cache.py 建立）
    找不到才嘗試 TWSE API
    """
    # ── 1. 優先讀本地快取 ────────────────────────────
    cache_path = OUTPUT_DIR / "shares_cache.json"
    if cache_path.exists():
        try:
            with open(cache_path, encoding="utf-8") as f:
                cache = json.load(f)
            shares = cache.get("shares", {})
            if shares:
                log.info("  發行股數（本地快取 %s）：%d 支",
                         cache.get("updated_at", "?"), len(shares))
                return {k: int(v) for k, v in shares.items()}
        except Exception as e:
            log.warning("  本地快取讀取失敗：%s", e)

    # ── 2. 嘗試 TWSE API ─────────────────────────────
    log.info("  找不到本地快取，嘗試從 TWSE 抓取...")
    from curl_cffi import requests as cffi_req
    session = cffi_req.Session(impersonate="chrome124")
    session.verify = False
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                             "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
               "Referer": "https://www.twse.com.tw/"}
    shares_map = {}
    today_str = datetime.today().strftime("%Y%m%d")
    try:
        r = session.get(
            "https://www.twse.com.tw/rwd/zh/fund/MI_QFIIS",
            params={"response": "json", "date": today_str, "selectType": "ALLBUT0999"},
            headers=headers, timeout=10
        )
        data = r.json()
        if data.get("stat") == "OK":
            fields = data.get("fields", [])
            code_col  = next((i for i, f in enumerate(fields)
                              if "代號" in str(f) or "代碼" in str(f)), 0)
            share_col = next((i for i, f in enumerate(fields)
                              if "發行" in str(f) and "股" in str(f)), None)
            if share_col is None:
                share_col = next((i for i, f in enumerate(fields) if "股數" in str(f)), None)
            if share_col is not None:
                for row in data.get("data", []):
                    try:
                        code = str(row[code_col]).strip()
                        if not code.isdigit() or len(code) > 5: continue
                        shares_k = int(str(row[share_col]).replace(",", "")) // 1000
                        if shares_k > 0:
                            shares_map[code] = shares_k
                    except Exception:
                        continue
                log.info("  發行股數（TWSE）：%d 支", len(shares_map))
    except Exception as e:
        log.warning("  上市發行股數抓取失敗：%s", e)

    if not shares_map:
        log.warning("  ⚠ 無法取得發行股數，請執行 python build_shares_cache.py")
    return shares_map


def _load_trust_cache() -> dict:
    """
    載入 build_trust_ratio.py 產生的長期投信累積持股 cache。
    回傳 {stock_id: cum_net_shares}，單位：股
    找不到時回傳空 dict（不影響主流程，trust_ratio 會 fallback 到短期連買估算）
    """
    cache_path = BASE_DIR / "trust_ratio_cache.json"
    if not cache_path.exists():
        log.info("  投本比 cache 不存在（%s），trust_ratio 使用短期連買估算", cache_path)
        return {}
    try:
        with open(cache_path, encoding="utf-8") as f:
            raw = json.load(f)
        result = {}
        for k, v in raw.items():
            if k.startswith("_"):
                continue
            if isinstance(v, dict) and "cum_net" in v:
                result[k] = int(v["cum_net"])
        meta = raw.get("_meta", {})
        log.info(
            "  投本比 cache 載入：%d 支，累積期間 %s → %s",
            len(result),
            meta.get("start_date", "?"),
            meta.get("last_date", "?"),
        )
        return result
    except Exception as e:
        log.warning("  投本比 cache 載入失敗：%s", e)
        return {}


def _fetch_revenue(codes: list, cache_path=None) -> dict:
    """
    從 finMind 批次抓取月營收，計算年增率
    回傳 {code: {"rev_yoy": 月營收年增率%, "cum_yoy": 累計營收年增率%}}
    使用本地快取避免重複抓取（每日最多更新一次）
    """
    import json as _json
    from datetime import datetime as _dt, timedelta as _td

    CACHE = Path("output/revenue_cache.json") if cache_path is None else Path(cache_path)
    today_str = _dt.today().strftime("%Y-%m-%d")

    # 讀本地快取
    cached = {}
    if CACHE.exists():
        try:
            with open(CACHE, encoding="utf-8") as f:
                c = _json.load(f)
            if c.get("date") == today_str:
                log.info("  月營收（本地快取 %s）：%d 支", today_str, len(c.get("data", {})))
                return c.get("data", {})
            cached = c.get("data", {})
        except Exception:
            pass

    # 從 finMind 抓取
    import requests as _req
    import urllib3 as _urllib3
    _urllib3.disable_warnings()

    # 起始日：去年初（確保有去年同期資料）
    start = (_dt.today().replace(day=1) - _td(days=365)).strftime("%Y-%m-%d")
    revenue_map = dict(cached)
    batch_size = 50
    ok_count = 0

    for i in range(0, len(codes), batch_size):
        batch = codes[i:i+batch_size]
        for code in batch:
            try:
                r = _req.get(
                    "https://api.finmindtrade.com/api/v4/data",
                    params={"dataset": "TaiwanStockMonthRevenue",
                            "data_id": code, "start_date": start, "token": ""},
                    timeout=10, verify=False,
                    headers={"User-Agent": "Mozilla/5.0"}
                )
                if r.status_code != 200: continue
                rows = r.json().get("data", [])
                if not rows: continue

                rev_dict = {}
                for row in rows:
                    y, m = int(row["revenue_year"]), int(row["revenue_month"])
                    rev_dict[(y, m)] = row["revenue"]

                today = _dt.today()
                cur_y, cur_m = today.year, today.month
                prev_m = cur_m - 1 if cur_m > 1 else 12
                prev_m_y = cur_y if cur_m > 1 else cur_y - 1

                cur_rev  = rev_dict.get((prev_m_y, prev_m))
                last_rev = rev_dict.get((prev_m_y - 1, prev_m))
                rev_yoy = round((cur_rev / last_rev - 1) * 100, 1) \
                          if cur_rev and last_rev and last_rev > 0 else None

                cum_cur  = sum(rev_dict.get((prev_m_y, m), 0) for m in range(1, prev_m + 1))
                cum_last = sum(rev_dict.get((prev_m_y - 1, m), 0) for m in range(1, prev_m + 1))
                cum_yoy = round((cum_cur / cum_last - 1) * 100, 1) \
                          if cum_cur and cum_last and cum_last > 0 else None

                revenue_map[code] = {"rev_yoy": rev_yoy, "cum_yoy": cum_yoy}
                ok_count += 1

            except Exception:
                continue

        # 每批次後顯示進度
        done = min(i + batch_size, len(codes))
        log.info("  營收進度：%d/%d，已取得 %d 支", done, len(codes), ok_count)
        import time as _time; _time.sleep(0.2)  # 避免 rate limit

    log.info("  月營收（finMind）：%d 支", ok_count)

    # 存快取
    try:
        CACHE.parent.mkdir(exist_ok=True)
        with open(CACHE, "w", encoding="utf-8") as f:
            _json.dump({"date": today_str, "data": revenue_map}, f,
                       ensure_ascii=False)
    except Exception as e:
        log.warning("  營收快取儲存失敗：%s", e)

    return revenue_map


def _download_prices(stock_list: pd.DataFrame, need_days=400, batch_size=50, delay=0.3) -> dict:
    """
    批次下載 OHLCV，回傳 {code: {"close","volume","high","low"}}
    共用給 STEP 4 和 STEP 5，只下載一次。
    """
    if stock_list is None or stock_list.empty or "ticker" not in stock_list.columns:
        log.warning("_download_prices: stock_list 是空的或缺少 ticker 欄，跳過下載")
        return {}
    end   = datetime.today()
    start = end - timedelta(days=need_days * 2)
    result = {}
    tickers = stock_list["ticker"].tolist()
    codes   = stock_list["code"].tolist()
    batches = [list(zip(codes[i:i+batch_size], tickers[i:i+batch_size]))
               for i in range(0, len(tickers), batch_size)]
    total   = len(batches)

    for bi, batch in enumerate(batches):
        batch_tickers = [t for _,t in batch]
        raw = None
        for attempt in range(3):
            try:
                raw = yf.download(
                    batch_tickers,
                    start=start.strftime("%Y-%m-%d"),
                    end=(end+timedelta(days=1)).strftime("%Y-%m-%d"),
                    auto_adjust=True, progress=False, threads=True,
                )
                break
            except Exception as e:
                if "Too Many" in str(e) or "429" in str(e):
                    w = 30*(attempt+1)
                    log.warning("yfinance 限速，等待 %ds...", w)
                    time.sleep(w)
                else:
                    log.warning("yfinance batch %d 失敗：%s", bi+1, e); break

        if raw is None or (hasattr(raw,'empty') and raw.empty):
            time.sleep(delay); continue

        try:
            mi = isinstance(raw.columns, pd.MultiIndex)
            def get_col(name):
                if mi:
                    return raw[name] if name in raw.columns.get_level_values(0) else None
                else:
                    if name in raw.columns:
                        r = raw[[name]]; r.columns=[batch_tickers[0]]; return r
                    return None

            closes  = get_col("Close")
            volumes = get_col("Volume")
            highs   = get_col("High")
            lows    = get_col("Low")
            opens   = get_col("Open")

            for code, ticker in batch:
                if closes is not None and ticker in closes.columns:
                    s = closes[ticker].dropna()
                    if len(s) > 10:
                        entry = {"close": s}
                        if volumes is not None and ticker in volumes.columns:
                            entry["volume"] = volumes[ticker].dropna()
                        if highs   is not None and ticker in highs.columns:
                            entry["high"]   = highs[ticker].dropna()
                        if lows    is not None and ticker in lows.columns:
                            entry["low"]    = lows[ticker].dropna()
                        if opens   is not None and ticker in opens.columns:
                            entry["open"]   = opens[ticker].dropna()
                        result[code] = entry
        except Exception as e:
            log.warning("解析失敗 batch %d：%s", bi+1, e)

        if (bi+1) % 10 == 0 or (bi+1) == total:
            log.info("  yfinance %d/%d 批次，%d 支有資料", bi+1, total, len(result))
        time.sleep(delay)

    log.info("yfinance 完成 %d / %d 支", len(result), len(tickers))
    return result


POSITIONS_CSV = OUTPUT_DIR / "positions.csv"
POSITIONS_COLS = ["code", "name", "strategy", "entry_date", "entry_price",
                   "status", "exit_date", "exit_price", "exit_reason", "ret_pct",
                   "current_price", "unrealized_ret_pct"]

# ════════════════════════════════════════════════════════════
# 虛擬持倉－股本版
#   A) 有完整回測驗證、且有明確股本甜蜜點的策略 → 只在該級距內進場：
#      1. 創高量縮       → 中大型 100-500億（n=2144, 平均+4.61%, 獲利因子2.62）
#      2. MA20斜率策略   → 中型   30-100億（n=157,  平均+2.96%, 獲利因子2.11）
#      3. 盤整突破-嚴格版 → 中小型 10-30億（n=118,  平均+2.23%, 獲利因子2.49，
#         但小型股(<10億)那格只有 n=10，樣本太小，這格不採信）
#   B) 處置預警-機會股：Kevin 指定保留。這個策略目前不在回測的11個策略清單裡
#      （屬於「連買+MA/創新高/強勢股/CANSLIM/Minervini」那批只累積了2026-07-02~07-11
#      共約10天即時法人資料的類別，還沒有可用的歷史回測），所以沒有股本甜蜜點可用，
#      不套用股本濾網、比照原版邏輯無條件進場。之後累積到足夠回測資料，再回來補級距設定。
# 若之後重跑 analyze_backtest_by_capital.py 有新結果，(A)的級距這裡手動更新即可。
# ════════════════════════════════════════════════════════════
CAPITAL_STRATEGY_RANGES = {
    "創高量縮":        (100, 500),
    "MA20斜率策略":     (30, 100),
    "盤整突破-嚴格版":   (10, 30),
}
# 沒有回測資料可用、但仍要保留在股本版裡的策略：不套股本濾網，訊號出現就進場
CAPITAL_UNFILTERED_STRATEGIES = {"處置預警-機會股"}
POSITIONS_CAPITAL_CSV = OUTPUT_DIR / "positions_capital.csv"


def _capital_billion(code: str, shares_map: dict) -> float | None:
    """股本(億元) = 流通股數(張) × 10,000 / 1e8"""
    lots = shares_map.get(code)
    if lots is None:
        return None
    return (lots * 10000) / 1e8


def capital_bucket_ok(code: str, strategy: str, shares_map: dict) -> bool:
    """判斷這支股票的股本，是否落在該策略回測表現最好的級距內"""
    rng = CAPITAL_STRATEGY_RANGES.get(strategy)
    if rng is None:
        return False
    cap = _capital_billion(code, shares_map)
    if cap is None:
        return False
    lo, hi = rng
    return lo <= cap < hi


def backfill_positions_capital(shares_map: dict) -> int:
    """把原版 positions.csv 裡「現有」的持倉（不管持有中或已出場），
    只要策略符合股本版規則（前三名策略要股本對、處置預警不設限），
    就補進 positions_capital.csv，不用等它們之後重新出現在每日訊號裡。
    用 (code, strategy, entry_date) 判斷是否已經存在，避免重複補入；只執行一次性補齊，
    之後每天新訊號一樣走 update_positions() 的正常流程。
    回傳這次新補進去的筆數。
    """
    orig = load_positions(POSITIONS_CSV)
    if orig.empty:
        return 0
    cap_df = load_positions(POSITIONS_CAPITAL_CSV)

    existing_keys = set(zip(cap_df.get("code", []), cap_df.get("strategy", []), cap_df.get("entry_date", [])))
    new_rows = []
    for _, row in orig.iterrows():
        strategy = row.get("strategy")
        code = row.get("code")
        if strategy in CAPITAL_UNFILTERED_STRATEGIES:
            eligible = True
        elif strategy in CAPITAL_STRATEGY_RANGES:
            eligible = capital_bucket_ok(code, strategy, shares_map)
        else:
            eligible = False
        if not eligible:
            continue
        key = (code, strategy, row.get("entry_date"))
        if key in existing_keys:
            continue
        new_rows.append(row)
        existing_keys.add(key)

    if not new_rows:
        return 0

    cap_df = pd.concat([cap_df, pd.DataFrame(new_rows)], ignore_index=True)
    for c in ["exit_date", "exit_price", "exit_reason", "ret_pct", "current_price", "unrealized_ret_pct"]:
        if c in cap_df.columns:
            cap_df[c] = cap_df[c].astype(object)
    save_positions(cap_df, POSITIONS_CAPITAL_CSV)
    return len(new_rows)


def load_positions(csv_path: Path = POSITIONS_CSV) -> pd.DataFrame:
    if csv_path.exists():
        try:
            df = pd.read_csv(csv_path, dtype={"code": str})
        except Exception as e:
            log.warning("讀取 %s 失敗，視為空倉：%s", csv_path.name, e)
            df = pd.DataFrame(columns=POSITIONS_COLS)
    else:
        df = pd.DataFrame(columns=POSITIONS_COLS)
    # 強制這幾欄用 object dtype，避免全 None 時被 pandas 推斷成 float64，
    # 之後寫入字串（例如 exit_date）會拋 LossySetitemError
    for c in ["exit_date", "exit_price", "exit_reason", "ret_pct", "current_price", "unrealized_ret_pct"]:
        if c in df.columns:
            df[c] = df[c].astype(object)
    return df


def save_positions(df: pd.DataFrame, csv_path: Path = POSITIONS_CSV):
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")


def update_positions(price_data: dict, briefing_picks: list, today_str: str,
                      csv_path: Path = POSITIONS_CSV) -> tuple:
    """虛擬持倉每日更新：
    1) 對現有「未出場」部位，套用跟回測相同的 exit_logic 規則，檢查今天是否觸發出場
    2) 對今天「每日快報」5個分類新選中、且目前沒有未出場部位的股票，建立新部位（進場價=今日收盤）
    csv_path：預設是主版 positions.csv，股本版傳入 POSITIONS_CAPITAL_CSV 可獨立追蹤，互不干擾。
    回傳 (positions_df, buy_alerts, sell_alerts)
    """
    from exit_logic import simulate_exit

    positions = load_positions(csv_path)
    if positions.empty:
        positions = pd.DataFrame(columns=POSITIONS_COLS)

    buy_alerts, sell_alerts = [], []

    # ── 1) 檢查現有未出場部位是否觸發出場 ──────────────────
    open_mask = positions["status"] == "open"
    for idx in positions[open_mask].index:
        row = positions.loc[idx]
        code = row["code"]
        d = price_data.get(code)
        if not d or "close" not in d:
            continue
        close, low = d["close"], d.get("low")
        # 防呆：股價資料的日期索引有時會帶時區資訊（tz-aware），
        # 跟 entry_date 字串轉出來的 tz-naive Timestamp 比較時會直接拋錯，
        # 導致這支股票的出場檢查被 except 吃掉、整個跳過而不自知。
        # 統一都轉成不帶時區，避免這個問題。
        try:
            if close.index.tz is not None:
                close = close.tz_localize(None)
            if low is not None and low.index.tz is not None:
                low = low.tz_localize(None)
        except (AttributeError, TypeError):
            pass
        try:
            res = simulate_exit(
                row["entry_date"], float(row["entry_price"]), close, low,
                ma_type="MA20", support_lookback=5, stop_pct=0.08, max_hold_days=120,
            )
        except Exception as e:
            log.error("  ⚠️ 部位出場檢查失敗（此部位這次沒被檢查到，需人工確認）%s %s：%s",
                      code, row.get("strategy", ""), e)
            continue

        if res.get("status") == "exited":
            positions.loc[idx, "status"]      = "closed"
            positions.loc[idx, "exit_date"]   = res["exit_date"]
            positions.loc[idx, "exit_price"]  = res["exit_price"]
            positions.loc[idx, "exit_reason"] = res["exit_reason"]
            positions.loc[idx, "ret_pct"]     = res["ret_pct"]
            sell_alerts.append({
                "code": code, "name": row["name"], "strategy": row["strategy"],
                "entry_date": row["entry_date"], "entry_price": row["entry_price"],
                "exit_date": res["exit_date"], "exit_price": res["exit_price"],
                "exit_reason": res["exit_reason"], "ret_pct": res["ret_pct"],
                "delayed": res["exit_date"] != today_str,
            })

    # ── 2) 今天新選中、目前沒有未出場部位的股票 → 建新部位 ──
    still_open_keys = set(
        (r["code"], r["strategy"]) for _, r in positions[positions["status"] == "open"].iterrows()
    )
    new_rows = []
    for pick in briefing_picks:
        code, strategy = pick["code"], pick["strategy"]
        if (code, strategy) in still_open_keys:
            continue  # 已經有未出場部位，不重複建倉
        d = price_data.get(code)
        if not d or "close" not in d or d["close"].empty:
            continue
        entry_price = float(d["close"].iloc[-1])
        new_rows.append({
            "code": code, "name": pick.get("name", ""), "strategy": strategy,
            "entry_date": today_str, "entry_price": round(entry_price, 2),
            "status": "open", "exit_date": None, "exit_price": None,
            "exit_reason": None, "ret_pct": None,
        })
        still_open_keys.add((code, strategy))
        buy_alerts.append({
            "code": code, "name": pick.get("name", ""), "strategy": strategy,
            "entry_date": today_str, "entry_price": round(entry_price, 2),
        })

    if new_rows:
        positions = pd.concat([positions, pd.DataFrame(new_rows)], ignore_index=True)
        for c in ["exit_date", "exit_price", "exit_reason", "ret_pct"]:
            if c in positions.columns:
                positions[c] = positions[c].astype(object)

    # ── 幫「持有中」的部位算今日股價/未實現報酬率，不用等出場才看得到損益 ──
    for c in ["current_price", "unrealized_ret_pct"]:
        if c not in positions.columns:
            positions[c] = None
        positions[c] = positions[c].astype(object)

    for idx in positions[positions["status"] == "open"].index:
        code = positions.loc[idx, "code"]
        entry_price = positions.loc[idx, "entry_price"]
        d = price_data.get(code)
        if not d or "close" not in d or d["close"].empty or entry_price in (None, 0):
            continue
        try:
            current_price = float(d["close"].iloc[-1])
            entry_price = float(entry_price)
            positions.loc[idx, "current_price"] = round(current_price, 2)
            positions.loc[idx, "unrealized_ret_pct"] = round((current_price / entry_price - 1) * 100, 2)
        except Exception:
            continue

    save_positions(positions, csv_path)
    return positions, buy_alerts, sell_alerts


def build_kline_slices(price_data: dict, codes: set, days: int = 90) -> dict:
    """從已下載的 price_data 截取指定股票最近N個交易日的OHLC，供每日快報畫K線用。
    不重新抓取，純粹從記憶體裡的資料切片，執行時間可忽略不計。"""
    out = {}
    for code in codes:
        d = price_data.get(code)
        if not d or "close" not in d:
            continue
        close = d["close"].tail(days)
        if close.empty:
            continue
        idx = close.index

        def _series(key):
            s = d.get(key)
            if s is None:
                return [None] * len(idx)
            s = s.reindex(idx)
            return [None if pd.isna(v) else round(float(v), 2) for v in s]

        out[code] = {
            "dates": [x.strftime("%Y-%m-%d") for x in idx],
            "open":  _series("open"),
            "high":  _series("high"),
            "low":   _series("low"),
            "close": _series("close"),
        }
    return out


def step45_strong_and_screener(stock_list: pd.DataFrame,
                                price_data: dict,
                                skip_yf=False) -> tuple[dict, list, list]:
    """
    同時執行 STEP 4（強勢股）和 STEP 5（創新高 + 連買MA）
    共用同一份 price_data，避免重複下載。
    """
    # ── STEP 4：強勢股 ──────────────────────────────────────
    log.info("STEP 4: 強勢股篩選...")
    import requests as req

    try:
        taiex = yf.download("^TWII", period="6mo", auto_adjust=True, progress=False)
        taiex_close  = taiex["Close"].squeeze().dropna()
        taiex_daily  = taiex_close.pct_change().dropna()
        taiex_ret    = {}
        for p, bars in [("1mo",21),("3mo",63)]:
            if len(taiex_close) >= bars+1:
                taiex_ret[p] = float(taiex_close.iloc[-1]/taiex_close.iloc[-bars-1]-1)
        taiex_avg_drop = float(taiex_daily[taiex_daily<0].mean()) if (taiex_daily<0).any() else -0.01
    except Exception as e:
        log.warning("大盤資料失敗：%s", e)
        taiex_ret = {"1mo": 0.0, "3mo": 0.0}
        taiex_daily = pd.Series(dtype=float)
        taiex_avg_drop = -0.01

    DEFENSE_RATIO = 0.8
    MA_S, MA_M, MA_L = 5, 20, 60
    PERIOD_BARS = {"1mo": 21, "3mo": 63}
    if stock_list is None or stock_list.empty or "code" not in stock_list.columns:
        log.warning("STEP 4/5: stock_list 為空，跳過強勢股/創新高篩選")
        empty_strong = {"stocks": [], "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "note": "TWSE 股票清單抓取失敗"}
        return empty_strong, [], []
    info_map = stock_list.set_index("code").to_dict("index")

    # 從 stock_list 取產業對照（_fetch_stock_list 已從 TWSE 抓取產業別）
    # 若 stock_list 沒有 industry 欄，再 fallback 到 market_flow CSV
    if "industry" in stock_list.columns:
        industry_map = (stock_list.dropna(subset=["industry"])
                                  .set_index("code")["industry"]
                                  .to_dict())
        log.info("  產業對照表（來自股票清單）：%d 筆", len(industry_map))
    else:
        industry_map = {}
        if MARKET_CSV.exists():
            try:
                mf = pd.read_csv(MARKET_CSV, dtype=str, usecols=lambda c: c in ["code","industry"])
                if "industry" in mf.columns:
                    industry_map = (mf.dropna(subset=["industry"])
                                      .drop_duplicates("code")
                                      .set_index("code")["industry"]
                                      .to_dict())
                    log.info("  產業對照表（來自 market_flow CSV）：%d 筆", len(industry_map))
            except Exception as e:
                log.warning("  產業對照表載入失敗：%s", e)

    strong_results = []
    for code, d in price_data.items():
        try:
            s = d["close"]
            if len(s) < MA_L + 5: continue
            latest = float(s.iloc[-1])
            row = {"代號": code+".TW" if info_map.get(code,{}).get("market")=="TWSE" else code+".TWO",
                   "股名": info_map.get(code,{}).get("name",""),
                   "產業": industry_map.get(code, "其他"),
                   "最新收盤": round(latest, 2)}

            # Layer 1: RS
            skip = False
            for p, mkt_r in taiex_ret.items():
                bars = PERIOD_BARS.get(p, 21)
                if len(s) < bars+1: skip=True; break
                stock_r = float(s.iloc[-1]/s.iloc[-bars-1]-1)
                rs = stock_r - mkt_r
                row[f"個股_{p}(%)"] = round(stock_r*100, 2)
                row[f"RS_{p}(%)"]   = round(rs*100, 2)
                if rs <= 0: skip=True; break
            if skip: continue

            # Layer 2: MA
            ma5  = float(s.iloc[-MA_S:].mean())
            ma20 = float(s.iloc[-MA_M:].mean())
            ma60 = float(s.iloc[-MA_L:].mean())
            if not (latest > ma5 > ma20 > ma60): continue
            row.update({"MA5": round(ma5,2), "MA20": round(ma20,2), "MA60": round(ma60,2)})

            # Layer 3: 抗跌
            stock_daily = s.pct_change().dropna()
            if len(taiex_daily) > 0:
                common = taiex_daily.index.intersection(stock_daily.index)
                if len(common) >= 5:
                    ta = taiex_daily.loc[common]; sa = stock_daily.loc[common]
                    down = ta < 0
                    if down.sum() >= 5:
                        td = float(ta[down].mean()); sd = float(sa[down].mean())
                        if td != 0:
                            defense = round(sd/td, 3)
                            if defense > DEFENSE_RATIO: continue
                            row["抗跌比"] = defense

            rs1 = row.get("RS_1mo(%)", 0); rs3 = row.get("RS_3mo(%)", 0)
            row["綜合評分"] = round(rs1*0.7+rs3*0.3, 2)
            strong_results.append(row)
        except Exception:
            pass

    strong_df = pd.DataFrame(strong_results).sort_values("綜合評分", ascending=False).reset_index(drop=True) if strong_results else pd.DataFrame()
    strong_stocks = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "taiex": {k: round(v*100,2) for k,v in taiex_ret.items()},
        "count": len(strong_df),
        "stocks": strong_df.to_dict(orient="records") if not strong_df.empty else [],
    }
    log.info("STEP 4 完成：%d 支強勢股", len(strong_df))

    # ── STEP 5a：創新高 ─────────────────────────────────────
    log.info("STEP 5: 創新高 + 連買MA篩選...")
    NEW_HIGH_LEVELS = [
        (252,"波段強勢","🚀"), (120,"標準強勢","📈"), (60,"標準強勢","📈"),
        (20,"轉強","⚡"), (10,"轉強","⚡"), (5,"轉強","⚡"),
    ]
    VOLATILITY_MONTHS = 3
    VOLATILITY_DAYS   = VOLATILITY_MONTHS * 20
    VOL_SURGE_RATIO   = 1.5

    newhigh = []
    for code, d in price_data.items():
        s = d["close"]
        if len(s) < 5: continue
        latest = float(s.iloc[-1])
        chg    = round((latest/float(s.iloc[-2])-1)*100, 2) if len(s) >= 2 else 0
        mkt    = info_map.get(code,{}).get("market","")
        name   = info_map.get(code,{}).get("name","")

        # 波動率（1個月=20日, 3個月=60日, 6個月=120日）
        h_data = d.get("high"); l_data = d.get("low")
        def calc_vola(days):
            vd = min(days, len(s))
            if h_data is not None and l_data is not None and len(h_data) >= vd:
                ph = float(h_data.tail(vd).max()); pl = float(l_data.tail(vd).min())
            else:
                ph = float(s.tail(vd).max()); pl = float(s.tail(vd).min())
            return round((ph-pl)/pl*100, 1) if pl > 0 else 0
        volatility_1m = calc_vola(20)
        volatility_3m = calc_vola(60)
        volatility_6m = calc_vola(120)
        volatility = volatility_3m  # 預設維持 3個月相容舊欄位

        # 量能
        v_series = d.get("volume")
        if v_series is not None and len(v_series) >= 2:
            tv = float(v_series.iloc[-1])
            av = float(v_series.iloc[-11:-1].mean()) if len(v_series) >= 11 else float(v_series.iloc[:-1].mean())
            vol_ratio = round(tv/av, 2) if av > 0 else 0
        else:
            vol_ratio = 0

        for n, category, icon in NEW_HIGH_LEVELS:
            if len(s) < n: continue
            peak = float(s.tail(n).max())
            if latest >= peak * 0.997:
                vk = d.get("volume")
                volume_k = round(float(vk.iloc[-1])/1000, 0) if vk is not None and not vk.empty else None
                # MA 計算
                ma20 = float(s.tail(20).mean()) if len(s) >= 20 else None
                ma60 = float(s.tail(60).mean()) if len(s) >= 60 else None
                ma_aligned = (ma20 is not None and ma60 is not None
                              and latest > ma20 > ma60)

                newhigh.append({
                    "code": code, "name": name, "price": round(latest,2),
                    "chg_pct": chg, "high_n": n, "peak": round(peak,2),
                    "category": category, "icon": icon, "market": mkt,
                    "volume_k": volume_k, "volatility": volatility,
                    "volatility_1m": volatility_1m,
                    "volatility_3m": volatility_3m,
                    "volatility_6m": volatility_6m,
                    "vol_months": VOLATILITY_MONTHS, "vol_ratio": vol_ratio,
                    "is_vol_surge": vol_ratio >= VOL_SURGE_RATIO,
                    "ma20":      round(ma20, 2) if ma20 else None,
                    "ma60":      round(ma60, 2) if ma60 else None,
                    "ma_aligned": ma_aligned,    # 股價>MA20>MA60
                })
                break
    newhigh.sort(key=lambda x: (["波段強勢","標準強勢","轉強"].index(x["category"]), -x["chg_pct"]))
    log.info("  創新高：%d 支", len(newhigh))

    # ── STEP 5b：連買MA ─────────────────────────────────────
    # 取得發行股數（用於計算投本比）
    shares_map = _fetch_shares_outstanding()
    trust_cache = _load_trust_cache()   # 長期投信累積 cache（build_trust_ratio.py 產生）
    # 今日 T86
    from curl_cffi import requests as cffi_req
    SESSION = cffi_req.Session(impersonate="chrome124"); SESSION.verify=False
    HEADERS = {"User-Agent":"Mozilla/5.0 Chrome/124.0.0.0","Referer":"https://www.twse.com.tw/"}

    def to_int(v):
        try: return int(str(v).replace(",","").replace("+","").strip() or 0)
        except: return 0

    # 今日 T86 上市
    today_t86 = pd.DataFrame()
    try:
        r = SESSION.get("https://www.twse.com.tw/rwd/zh/fund/T86",
                        headers=HEADERS, params={"selectType":"ALL","response":"json"},
                        timeout=20, allow_redirects=True)
        if r.status_code == 200:
            data = r.json()
            if data.get("stat") == "OK":
                rows_t86 = data.get("data",[])
                recs = []
                for row in rows_t86:
                    if len(row) < 11: continue
                    code2 = str(row[0]).strip()
                    recs.append({"code": code2, "date": TODAY,
                                 "foreign_net": (to_int(row[4])+to_int(row[7]))//1000,
                                 "trust_net":   to_int(row[10])//1000})
                today_t86 = pd.DataFrame(recs)
                today_t86 = today_t86[today_t86["code"].str.match(r"^\d{4,5}$", na=False)]
    except Exception as e:
        log.warning("  T86 上市法人資料抓取失敗：%s", e)

    # 今日 TPEx
    today_dt  = datetime.today()
    roc       = f"{today_dt.year-1911}/{today_dt.month:02d}/{today_dt.day:02d}"
    today_tpex = pd.DataFrame()
    try:
        r2 = SESSION.get("https://www.tpex.org.tw/web/stock/3insti/daily_trade/3itrade_hedge_result.php",
                         headers={**HEADERS,"Referer":"https://www.tpex.org.tw/"},
                         params={"l":"zh-tw","se":"EW","t":"D","d":roc,"s":"0,asc"},
                         timeout=15, allow_redirects=True)
        if r2.status_code == 200:
            aa = r2.json().get("aaData",[])
            recs2 = [{"code": str(row[0]).strip(), "date": TODAY,
                      "foreign_net": to_int(row[4]), "trust_net": to_int(row[7])}
                     for row in aa if len(row)>=8]
            today_tpex = pd.DataFrame(recs2) if recs2 else pd.DataFrame()
            if not today_tpex.empty:
                today_tpex = today_tpex[today_tpex["code"].str.match(r"^\d{4,5}$", na=False)]
    except Exception as e:
        log.warning("  TPEx 上櫃法人資料抓取失敗：%s", e)

    # 合併並更新快取
    today_parts = [f for f in [today_t86, today_tpex] if not f.empty and "code" in f.columns]
    today_all   = pd.concat(today_parts, ignore_index=True) if today_parts else pd.DataFrame()

    if not today_all.empty:
        if INSTI_CSV.exists():
            cached = pd.read_csv(INSTI_CSV, dtype={"code":str,"date":str})
            cached = cached[cached["date"] != TODAY]
            combined = pd.concat([cached, today_all], ignore_index=True)
        else:
            combined = today_all
        for c in ["foreign_net","trust_net"]:
            combined[c] = pd.to_numeric(combined[c], errors="coerce").fillna(0).astype(int)
        combined.sort_values(["date","code"], inplace=True)
        combined.to_csv(INSTI_CSV, index=False, encoding="utf-8-sig")

    history = pd.read_csv(INSTI_CSV, dtype={"code":str,"date":str}) if INSTI_CSV.exists() else pd.DataFrame()

    # calc_streak
    STREAK_MIN = 2; MA_FAST = 87; MA_SLOW = 284; HISTORY_DAYS = 30
    insti_ma = []
    if not history.empty:
        all_trade_days = sorted(history["date"].unique(), reverse=True)
        latest_day     = all_trade_days[0]
        today_codes    = history[history["date"]==latest_day]["code"].unique()

        for code2 in today_codes:
            if code2 not in price_data: continue
            s = price_data[code2]["close"]
            lookup = history[history["code"]==code2].set_index("date").to_dict("index")

            def streak_fn(col):
                d, cum = 0, 0
                for dt2 in all_trade_days:
                    net = lookup.get(dt2,{}).get(col,0)
                    if net > 0: d+=1; cum+=net
                    else: break
                return d, cum

            fd, fc = streak_fn("foreign_net")
            td, tc = streak_fn("trust_net")

            if fd < STREAK_MIN and td < STREAK_MIN: continue
            if len(s) < MA_SLOW+5: continue
            ma_f = float(s.tail(MA_FAST).mean())
            ma_s = float(s.tail(MA_SLOW).mean())
            if ma_f <= ma_s: continue

            latest2 = float(s.iloc[-1])
            chg2    = round((latest2/float(s.iloc[-2])-1)*100, 2) if len(s)>=2 else 0
            vk2     = price_data[code2].get("volume")
            vol_k2  = round(float(vk2.iloc[-1])/1000, 0) if vk2 is not None and not vk2.empty else None

            insti_ma.append({
                "code": code2, "name": info_map.get(code2,{}).get("name",""),
                "price": round(latest2,2), "chg_pct": chg2,
                "ma_fast": round(ma_f,2), "ma_slow": round(ma_s,2),
                "ma_diff_pct": round((ma_f/ma_s-1)*100, 2),
                "foreign_streak": fd, "foreign_cum": fc,
                "trust_streak": td, "trust_cum": tc,
                "f_hit_limit": fd>=HISTORY_DAYS, "t_hit_limit": td>=HISTORY_DAYS,
                "market": info_map.get(code2,{}).get("market",""),
                "volume_k": vol_k2,
                # ── 投本比：優先用長期 cache，fallback 用短期連買估算 ──
                "trust_ratio": (
                    round(trust_cache[code2] / (shares_map[code2] * 1000) * 100, 2)
                    if code2 in trust_cache and code2 in shares_map and shares_map[code2] > 0
                    else (
                        round(tc / (shares_map[code2] * 1000) * 100, 2)
                        if code2 in shares_map and shares_map[code2] > 0 else None
                    )
                ),
                "trust_ratio_src": "cache" if code2 in trust_cache else "streak",
                "foreign_ratio": round(fc / (shares_map[code2] * 1000) * 100, 2)
                                 if code2 in shares_map and shares_map[code2] > 0 else None,
            })
        insti_ma.sort(key=lambda x: (-(x["foreign_streak"]+x["trust_streak"]), -x["ma_diff_pct"]))

    log.info("STEP 5 完成：創新高 %d 支，連買MA %d 支", len(newhigh), len(insti_ma))

    # ── STEP 5c：VCP 型態篩選 ────────────────────────────────
    log.info("STEP 5c: VCP 型態篩選...")

    def calc_vcp(code, d):
        """
        計算 VCP 型態，回傳 dict 或 None
        核心邏輯：
        1. 找出近期 2-4 個波段高低點（用滾動窗口找局部高低）
        2. 每次回檔幅度遞減（後 < 前 × 0.75）
        3. 每次回檔量縮（後段平均量 < 前段平均量）
        4. 目前價格距近期高點 < 15%（壓縮在前高附近）
        5. MA20 > MA60（基本趨勢向上）
        """
        try:
            s   = d["close"]
            vol = d.get("volume")
            if len(s) < 60: return None

            latest = float(s.iloc[-1])
            ma20   = float(s.iloc[-20:].mean())
            ma60   = float(s.iloc[-60:].mean())

            # 基本趨勢：MA20 > MA60
            if ma20 <= ma60: return None

            # 只看最近 120 日
            s_w = s.iloc[-120:].reset_index(drop=True)
            v_w = vol.iloc[-120:].reset_index(drop=True) if vol is not None and len(vol) >= 120 else None
            n   = len(s_w)

            # 找局部高點（窗口 5 日）
            WIN = 5
            peaks = []
            for i in range(WIN, n - WIN):
                if float(s_w.iloc[i]) == float(s_w.iloc[i-WIN:i+WIN+1].max()):
                    peaks.append(i)

            # 找局部低點
            troughs = []
            for i in range(WIN, n - WIN):
                if float(s_w.iloc[i]) == float(s_w.iloc[i-WIN:i+WIN+1].min()):
                    troughs.append(i)

            if len(peaks) < 2 or len(troughs) < 2: return None

            # 取最近 3 個高點（至少 2 個）
            recent_peaks = peaks[-3:]

            # 計算各次回檔幅度
            contractions = []
            for i, pk in enumerate(recent_peaks):
                # 找該高點之後最近的低點
                tr_after = [t for t in troughs if t > pk]
                if not tr_after: continue
                tr = tr_after[0]
                peak_price  = float(s_w.iloc[pk])
                trough_price = float(s_w.iloc[tr])
                pullback = (peak_price - trough_price) / peak_price  # 回檔幅度

                # 該段平均量
                avg_vol = None
                if v_w is not None:
                    seg_vol = v_w.iloc[pk:tr+1]
                    avg_vol = float(seg_vol.mean()) if len(seg_vol) > 0 else None

                contractions.append({
                    "peak_idx":   pk,
                    "trough_idx": tr,
                    "peak_price": round(peak_price, 2),
                    "trough_price": round(trough_price, 2),
                    "pullback":   round(pullback * 100, 1),
                    "avg_vol":    avg_vol,
                })

            if len(contractions) < 2: return None

            # 驗證收縮：每次回檔幅度 < 上次 × 0.8
            shrink_count = 0
            for i in range(1, len(contractions)):
                prev = contractions[i-1]["pullback"]
                curr = contractions[i]["pullback"]
                if curr < prev * 0.8:
                    shrink_count += 1

            if shrink_count < 1: return None  # 至少 1 次有效收縮

            # 驗證量縮（有量資料才檢查）
            vol_shrink = True
            if all(c["avg_vol"] for c in contractions):
                for i in range(1, len(contractions)):
                    if contractions[i]["avg_vol"] >= contractions[i-1]["avg_vol"] * 1.1:
                        vol_shrink = False
                        break

            # 最近高點
            last_peak_price = float(s_w.iloc[recent_peaks[-1]])
            dist_from_peak  = round((last_peak_price - latest) / last_peak_price * 100, 1)

            # 距前高 < 15%
            if dist_from_peak > 15: return None

            # 最後一次回檔幅度（越小越好，代表壓縮越緊）
            last_pullback = contractions[-1]["pullback"]

            return {
                "shrink_count": shrink_count,
                "contractions": len(contractions),
                "last_pullback": last_pullback,
                "dist_from_peak": dist_from_peak,
                "vol_shrink": vol_shrink,
                "ma20": round(ma20, 2),
                "ma60": round(ma60, 2),
            }
        except Exception:
            return None

    # 跑所有股票的 VCP
    vcp_results = []
    etf_kw_vcp = ["ETF","基金","永續","高息","高股息","月配","季配","債券","期貨","正2","反1"]
    strong_codes = set(
        str(s.get("代號","")).replace(".TW","").replace(".TWO","")
        for s in strong_stocks.get("stocks", [])
    )

    for code, d in price_data.items():
        # 排除 ETF
        name = info_map.get(code, {}).get("name", "")
        if code.startswith("00") or any(k in name for k in etf_kw_vcp):
            continue

        vcp = calc_vcp(code, d)
        if vcp is None: continue

        s    = d["close"]
        latest = round(float(s.iloc[-1]), 2)
        chg    = round((float(s.iloc[-1])/float(s.iloc[-2])-1)*100, 2) if len(s) >= 2 else 0
        vol_k  = round(float(d["volume"].iloc[-1])/1000, 1) if d.get("volume") is not None and len(d["volume"]) > 0 else 0

        is_strong = code in strong_codes
        tier = "強勢VCP" if is_strong else "廣義VCP"

        vcp_results.append({
            "code":           code,
            "name":           name,
            "market":         info_map.get(code, {}).get("market", ""),
            "industry":       info_map.get(code, {}).get("industry", "其他"),
            "price":          latest,
            "chg_pct":        chg,
            "volume_k":       vol_k,
            "tier":           tier,
            "shrink_count":   vcp["shrink_count"],
            "contractions":   vcp["contractions"],
            "last_pullback":  vcp["last_pullback"],
            "dist_from_peak": vcp["dist_from_peak"],
            "vol_shrink":     vcp["vol_shrink"],
            "ma20":           vcp["ma20"],
            "ma60":           vcp["ma60"],
        })

    # 排序：強勢VCP優先，再依距前高排序（越小越接近突破）
    vcp_results.sort(key=lambda x: (0 if x["tier"]=="強勢VCP" else 1, x["dist_from_peak"]))
    log.info("STEP 5c 完成：VCP %d 支（強勢 %d，廣義 %d）",
             len(vcp_results),
             sum(1 for r in vcp_results if r["tier"]=="強勢VCP"),
             sum(1 for r in vcp_results if r["tier"]=="廣義VCP"))

    # ── STEP 5d：馬克維尼亞 Trend Template ──────────────────
    log.info("STEP 5d_m: 馬克維尼亞 Trend Template 篩選...")
    etf_kw_m = ["ETF","基金","永續","高息","高股息","月配","季配","債券","期貨","正2","反1"]
    minervini_results = []

    for code, d in price_data.items():
        name = info_map.get(code, {}).get("name", "")
        if code.startswith("00") or any(k in name for k in etf_kw_m):
            continue
        try:
            s = d["close"]
            if len(s) < 210: continue
            latest = float(s.iloc[-1])
            ma50   = float(s.iloc[-50:].mean())
            ma150  = float(s.iloc[-150:].mean())
            ma200  = float(s.iloc[-200:].mean())
            ma200_1m = float(s.iloc[-230:-180].mean()) if len(s) >= 230 else None
            s_52   = s.iloc[-252:] if len(s) >= 252 else s
            high_52 = float(s_52.max())
            low_52  = float(s_52.min())
            rs3 = 0
            for ss_item in strong_stocks.get("stocks", []):
                c = str(ss_item.get("代號","")).replace(".TW","").replace(".TWO","")
                if c == code:
                    rs3 = ss_item.get("RS_3mo(%)", 0)
                    break
            c1 = latest > ma150 and latest > ma200
            c2 = ma150 > ma200
            c3 = (ma200_1m is not None and ma200 > ma200_1m)
            c4 = ma50 > ma150 and ma50 > ma200
            c5 = latest > ma50
            c6 = (low_52 > 0 and (latest - low_52) / low_52 >= 0.30)
            c7 = (high_52 > 0 and (high_52 - latest) / high_52 <= 0.25)
            c8 = rs3 >= 10
            passed = sum([c1,c2,c3,c4,c5,c6,c7,c8])
            if passed < 6: continue
            dist_high = round((high_52 - latest) / high_52 * 100, 1) if high_52 > 0 else None
            dist_low  = round((latest - low_52)  / low_52  * 100, 1) if low_52  > 0 else None
            chg = round((latest / float(s.iloc[-2]) - 1) * 100, 2) if len(s) >= 2 else 0
            vol_k = round(float(d["volume"].iloc[-1]) / 1000, 1) if d.get("volume") is not None and len(d["volume"]) > 0 else 0
            minervini_results.append({
                "code": code, "name": name,
                "market": info_map.get(code,{}).get("market",""),
                "industry": info_map.get(code,{}).get("industry","其他"),
                "price": round(latest,2), "chg_pct": chg, "volume_k": vol_k,
                "passed": passed, "ma50": round(ma50,2), "ma150": round(ma150,2),
                "ma200": round(ma200,2), "dist_high": dist_high, "dist_low": dist_low, "rs3": rs3,
                "c1_price_above_ma": c1, "c2_ma150_above_ma200": c2, "c3_ma200_rising": c3,
                "c4_ma50_top": c4, "c5_price_above_ma50": c5, "c6_above_52w_low": c6,
                "c7_near_52w_high": c7, "c8_rs_strong": c8,
            })
        except Exception:
            continue
    minervini_results.sort(key=lambda x: (-x["passed"], x.get("dist_high", 999)))
    log.info("STEP 5d_m 完成：馬克維尼亞 %d 支（全條件 %d 支）",
             len(minervini_results), sum(1 for r in minervini_results if r["passed"]==8))

    # ── STEP 5e：歐尼爾 CANSLIM（N/S/L/I/M + 營收）────────────
    log.info("STEP 5e: 歐尼爾 CANSLIM 篩選（N/S/L/I/M + 營收）...")
    etf_kw_o = ["ETF","基金","永續","高息","高股息","月配","季配","債券","期貨","正2","反1"]

    # 先建立各模組代碼集合（candidate_codes 篩選需要用到）
    newhigh52_codes = set(r["code"] for r in newhigh if r.get("high_n",0) >= 252)
    vol_surge_codes = set(r["code"] for r in newhigh if r.get("is_vol_surge", False))
    insti_codes_set = set(r["code"] for r in insti_ma)
    rs_strong_codes = {}
    for ss_item in strong_stocks.get("stocks",[]):
        c = str(ss_item.get("代號","")).replace(".TW","").replace(".TWO","")
        rs_strong_codes[c] = {"rs1": ss_item.get("RS_1mo(%)",0),
                               "rs3": ss_item.get("RS_3mo(%)",0),
                               "ret1": ss_item.get("個股_1mo(%)",0),
                               "ret3": ss_item.get("個股_3mo(%)",0),
                               "score": ss_item.get("綜合評分",0)}
    market_up = all(v > 0 for v in taiex_ret.values()) if taiex_ret else True

    # 抓月營收（finMind）- 只抓 N/S/L/I 至少通過 2 項的候選股，避免抓太多
    candidate_codes = []
    for code in price_data.keys():
        if code.startswith("00"): continue
        name = info_map.get(code,{}).get("name","")
        if any(k in name for k in etf_kw_o): continue
        pre_score = sum([
            code in newhigh52_codes,
            code in vol_surge_codes,
            code in rs_strong_codes,
            code in insti_codes_set,
        ])
        if pre_score >= 2:
            candidate_codes.append(code)
    log.info("  CANSLIM 候選股：%d 支（預抓營收）", len(candidate_codes))

    revenue_map = {}
    try:
        revenue_map = _fetch_revenue(candidate_codes)
    except Exception as e:
        log.warning("  月營收抓取失敗，C/A 條件將跳過：%s", e)

    canslim_results = []
    for code, d in price_data.items():
        name = info_map.get(code,{}).get("name","")
        if code.startswith("00") or any(k in name for k in etf_kw_o):
            continue
        try:
            s = d["close"]
            if len(s) < 60: continue
            latest = float(s.iloc[-1])
            n_score = code in newhigh52_codes
            s_score = code in vol_surge_codes
            l_score = code in rs_strong_codes
            i_score = code in insti_codes_set
            m_score = market_up
            # C：月營收年增率 > 20%
            rev_info = revenue_map.get(code, {})
            rev_yoy = rev_info.get("rev_yoy")
            cum_yoy = rev_info.get("cum_yoy")
            c_score = rev_yoy is not None and rev_yoy >= 20
            # A：累計營收年增率 > 20%
            a_score = cum_yoy is not None and cum_yoy >= 20
            passed_list = [k for k,v in {
                "C": c_score, "A": a_score,
                "N": n_score, "S": s_score,
                "L": l_score, "I": i_score, "M": m_score
            }.items() if v]
            passed = len(passed_list)
            if passed < 3: continue
            rs_info = rs_strong_codes.get(code, {})
            chg = round((latest / float(s.iloc[-2]) - 1) * 100, 2) if len(s) >= 2 else 0
            vol_k = round(float(d["volume"].iloc[-1]) / 1000, 1) if d.get("volume") is not None and len(d["volume"]) > 0 else 0
            nh_info = next((r for r in newhigh if r["code"]==code), {})
            canslim_results.append({
                "code": code, "name": name,
                "market": info_map.get(code,{}).get("market",""),
                "industry": info_map.get(code,{}).get("industry","其他"),
                "price": round(latest,2), "chg_pct": chg, "volume_k": vol_k,
                "passed": passed, "passed_list": passed_list,
                "rs1": rs_info.get("rs1",0), "rs3": rs_info.get("rs3",0),
                "ret1": rs_info.get("ret1",0), "ret3": rs_info.get("ret3",0),
                "score": rs_info.get("score",0),
                "vol_ratio": nh_info.get("vol_ratio",0), "high_n": nh_info.get("high_n",0),
                "rev_yoy": rev_yoy, "cum_yoy": cum_yoy,
                "C": c_score, "A": a_score,
                "N": n_score, "S": s_score, "L": l_score, "I": i_score, "M": m_score,
            })
        except Exception:
            continue
    canslim_results.sort(key=lambda x: (-x["passed"], -x.get("score",0)))
    log.info("STEP 5e 完成：CANSLIM %d 支（全條件 7/7：%d 支，5+：%d 支）",
             len(canslim_results),
             sum(1 for r in canslim_results if r["passed"]==7),
             sum(1 for r in canslim_results if r["passed"]>=5))

    return strong_stocks, newhigh, insti_ma, vcp_results, minervini_results, canslim_results


# ════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════
# STEP 5d：過熱指標計算
# ════════════════════════════════════════════════════════════

def step_overheat(price_data: dict, strong_stocks: dict, newhigh: list,
                  insti_ma: list, vcp_results: list, stock_list) -> list:
    """
    對所有篩選板塊個股計算過熱指標：
      - RSI(14)
      - KD(9,3,3)
      - MA20 乖離率 / MA60 乖離率 / MA240 乖離率
      - 布林通道位置（距上軌標準差倍數）
      - 過熱分數（0–100）與等級
    涵蓋來源：強勢股、創新高、連買MA、VCP
    """
    log.info("STEP 5d: 過熱指標計算...")

    # ── 1. 收集所有篩選板塊的代碼 ──────────────────────────
    etf_kw = {"ETF","基金","永續","高息","高股息","月配","季配","債券","期貨","正2","反1"}

    def is_etf(code, name=""):
        return code.startswith("00") or any(k in name for k in etf_kw)

    all_codes = set()

    # 強勢股
    for s in strong_stocks.get("stocks", []):
        c = str(s.get("代號","")).replace(".TW","").replace(".TWO","")
        if c: all_codes.add(c)

    # 創新高
    for r in (newhigh or []):
        c = str(r.get("code",""))
        if c: all_codes.add(c)

    # 連買 MA
    for r in (insti_ma or []):
        c = str(r.get("code",""))
        if c: all_codes.add(c)

    # VCP
    for r in (vcp_results or []):
        c = str(r.get("code",""))
        if c: all_codes.add(c)

    log.info("  過熱計算涵蓋 %d 支個股", len(all_codes))

    # ── 2. info_map（名稱/產業）─────────────────────────────
    info_map = {}
    if stock_list is not None and not stock_list.empty and "code" in stock_list.columns:
        info_map = stock_list.set_index("code").to_dict("index")

    # 補充名稱從各板塊
    name_map  = {}
    ind_map   = {}
    for s in strong_stocks.get("stocks", []):
        c = str(s.get("代號","")).replace(".TW","").replace(".TWO","")
        name_map[c]  = s.get("股名","")
        ind_map[c]   = s.get("產業","其他")
    for lst in [newhigh or [], insti_ma or [], vcp_results or []]:
        for r in lst:
            c = str(r.get("code",""))
            if c not in name_map:
                name_map[c] = r.get("name","")
            if c not in ind_map:
                ind_map[c]  = r.get("industry","其他")

    # ── 3. 來源標籤 ──────────────────────────────────────────
    strong_codes = {str(s.get("代號","")).replace(".TW","").replace(".TWO","")
                    for s in strong_stocks.get("stocks",[])}
    newhigh_codes= {str(r.get("code","")) for r in (newhigh or [])}
    insti_codes  = {str(r.get("code","")) for r in (insti_ma or [])}
    vcp_codes    = {str(r.get("code","")) for r in (vcp_results or [])}

    # ── 4. 逐股計算 ──────────────────────────────────────────
    results = []

    for code in all_codes:
        if code not in price_data:
            continue
        d   = price_data[code]
        s   = d.get("close")
        vol = d.get("volume")
        hi  = d.get("high")
        lo  = d.get("low")

        if s is None or len(s) < 20:
            continue

        name    = name_map.get(code, info_map.get(code,{}).get("name",""))
        industry= ind_map.get(code, info_map.get(code,{}).get("industry","其他"))

        if is_etf(code, name):
            continue

        latest = float(s.iloc[-1])
        if latest <= 0:
            continue

        try:
            # ── RSI(14) ──────────────────────────────────────
            delta = s.diff()
            gain  = delta.clip(lower=0).rolling(14).mean()
            loss  = (-delta.clip(upper=0)).rolling(14).mean()
            rs_   = gain / loss.replace(0, float("nan"))
            rsi_s = (100 - 100 / (1 + rs_)).dropna()
            rsi14 = round(float(rsi_s.iloc[-1]), 1) if len(rsi_s) >= 1 else None

            # RSI 頂背離：價格創近30日新高，但 RSI < 前高
            rsi_div = False
            if rsi14 is not None and len(rsi_s) >= 30 and len(s) >= 30:
                px30   = s.iloc[-30:]
                rsi30  = rsi_s.iloc[-30:]
                if float(px30.iloc[-1]) >= float(px30.max()) * 0.98:  # 價格接近高點
                    prev_rsi_max = float(rsi30.iloc[:-3].max()) if len(rsi30) > 3 else float(rsi30.max())
                    if rsi14 < prev_rsi_max - 5:  # RSI 明顯低於前高
                        rsi_div = True

            # ── KD(9,3,3) ────────────────────────────────────
            k_val = d_val = None
            if hi is not None and lo is not None and len(hi) >= 9 and len(lo) >= 9:
                low9  = lo.rolling(9).min()
                high9 = hi.rolling(9).max()
                denom = (high9 - low9).replace(0, float("nan"))
                rsv   = (s - low9) / denom * 100
                k_s   = rsv.ewm(com=2, adjust=False).mean()
                d_s   = k_s.ewm(com=2, adjust=False).mean()
                k_val = round(float(k_s.iloc[-1]), 1)
                d_val = round(float(d_s.iloc[-1]), 1)

            # ── 乖離率 ────────────────────────────────────────
            def bias(n):
                if len(s) < n: return None
                ma = float(s.iloc[-n:].mean())
                return round((latest - ma) / ma * 100, 1) if ma > 0 else None

            bias20  = bias(20)
            bias60  = bias(60)
            bias240 = bias(240) if len(s) >= 240 else None

            ma20_val  = round(float(s.iloc[-20:].mean()), 2) if len(s) >= 20 else None
            ma60_val  = round(float(s.iloc[-60:].mean()), 2) if len(s) >= 60 else None

            # ── 布林通道位置 ──────────────────────────────────
            bb_pos = None
            if len(s) >= 20:
                ma20f = s.iloc[-20:].mean()
                std20 = s.iloc[-20:].std()
                if std20 > 0:
                    bb_pos = round((latest - float(ma20f)) / (2 * float(std20)), 2)

            # ── 量比（今日/10日均量）────────────────────────
            vol_ratio = None
            if vol is not None and len(vol) >= 11:
                avg10 = float(vol.iloc[-11:-1].mean())
                if avg10 > 0:
                    vol_ratio = round(float(vol.iloc[-1]) / avg10, 2)

            # ── 過熱分數（0–100）─────────────────────────────
            score = 0

            # RSI 貢獻（最高 35 分）
            if rsi14 is not None:
                if rsi14 >= 85:   score += 35
                elif rsi14 >= 80: score += 28
                elif rsi14 >= 75: score += 18
                elif rsi14 >= 70: score += 10
            if rsi_div:           score += 8   # 頂背離加分

            # 乖離 MA20 貢獻（最高 25 分）
            if bias20 is not None:
                if bias20 >= 25:   score += 25
                elif bias20 >= 18: score += 20
                elif bias20 >= 12: score += 13
                elif bias20 >= 8:  score += 7

            # KD 貢獻（最高 20 分）
            if k_val is not None and d_val is not None:
                if k_val >= 90 and d_val >= 85:   score += 20
                elif k_val >= 80 and d_val >= 75: score += 13
                elif k_val >= 75:                 score += 7

            # 布林通道貢獻（最高 20 分）
            if bb_pos is not None:
                if bb_pos >= 1.5:   score += 20
                elif bb_pos >= 1.2: score += 14
                elif bb_pos >= 1.0: score += 8
                elif bb_pos >= 0.8: score += 4

            score = min(score, 100)

            level = ("極熱🔴" if score >= 75 else
                     "過熱🟠" if score >= 50 else
                     "偏熱🟡" if score >= 30 else
                     "正常🟢")

            # ── 來源標籤 ──────────────────────────────────────
            sources = []
            if code in strong_codes:  sources.append("強勢股")
            if code in newhigh_codes: sources.append("創新高")
            if code in insti_codes:   sources.append("連買MA")
            if code in vcp_codes:     sources.append("VCP")

            results.append({
                "code":      code,
                "name":      name,
                "industry":  industry,
                "price":     round(latest, 2),
                "rsi14":     rsi14,
                "rsi_div":   rsi_div,
                "k_val":     k_val,
                "d_val":     d_val,
                "bias20":    bias20,
                "bias60":    bias60,
                "bias240":   bias240,
                "bb_pos":    bb_pos,
                "vol_ratio": vol_ratio,
                "ma20":      ma20_val,
                "ma60":      ma60_val,
                "score":     score,
                "level":     level,
                "sources":   sources,
            })

        except Exception as e:
            log.debug("  %s 過熱計算失敗：%s", code, e)
            continue

    results.sort(key=lambda x: -x["score"])
    log.info("STEP 5d 完成：過熱計算 %d 支（極熱 %d，過熱 %d）",
             len(results),
             sum(1 for r in results if "極熱" in r["level"]),
             sum(1 for r in results if "過熱" in r["level"]))

    # ── 注入各板塊：讓每支個股記錄帶上過熱欄位 ────────────
    oh_map = {r["code"]: r for r in results}
    OH_FIELDS = ["rsi14", "k_val", "d_val", "bias20", "bias60", "bb_pos",
                 "vol_ratio", "score", "level", "rsi_div"]

    def _inject(record, code):
        oh = oh_map.get(code)
        if oh:
            for f in OH_FIELDS:
                record[f"oh_{f}"] = oh.get(f)
        return record

    # 強勢股
    for s in strong_stocks.get("stocks", []):
        c = str(s.get("代號", "")).replace(".TW", "").replace(".TWO", "")
        _inject(s, c)

    # 創新高
    for r in (newhigh or []):
        _inject(r, str(r.get("code", "")))

    # 連買MA
    for r in (insti_ma or []):
        _inject(r, str(r.get("code", "")))

    # VCP
    for r in (vcp_results or []):
        _inject(r, str(r.get("code", "")))

    return results


# ════════════════════════════════════════════════════════════
# STEP 5f：處置股 + 注意股預警
# ════════════════════════════════════════════════════════════

def step_disposition(price_data: dict = None) -> dict:
    """
    抓取：
    1. 目前處置中的股票（上市+上櫃）
    2. 注意累計次數異常 → 預判明天可能進處置的股票

    進處置條件（符合其一）：
      - 連續 3 個營業日被列注意股
      - 連續 5 個營業日被列注意股
      - 近 10 個營業日內達 6 次
      - 近 30 個營業日內達 12 次
    """
    log.info("STEP 5f: 處置股 + 注意股預警...")

    HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://www.twse.com.tw/",
    }
    import ssl
    ctx = ssl._create_unverified_context()

    result = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "disposal":     [],   # 目前處置中
        "warning":      [],   # 明天可能進處置
    }

    # ── 1. 目前處置股（上市）─────────────────────────────────
    try:
        url = "https://www.twse.com.tw/rwd/zh/announcement/punish"
        req = urllib.request.Request(url, headers=HEADERS)
        r = urllib.request.urlopen(req, context=ctx, timeout=15)
        data = json.loads(r.read())
        if data.get("stat") == "OK" and data.get("data"):
            for row in data["data"]:
                if len(row) < 4:
                    continue
                code = str(row[2]).strip()
                name = str(row[3]).strip()
                # 只保留普通股（4位數字，非ETF）
                if not _is_common_stock(code):
                    continue
                period = str(row[6]).strip() if len(row) > 6 else ""
                if "～" in period:
                    parts = period.split("～"); start = parts[0].strip(); end = parts[1].strip()
                elif "~" in period:
                    parts = period.split("~"); start = parts[0].strip(); end = parts[1].strip()
                else:
                    start = period; end = ""
                mode     = str(row[5]).strip() if len(row) > 5 else ""
                measures = str(row[7]).strip() if len(row) > 7 else ""
                result["disposal"].append({
                    "code": code, "name": name, "market": "上市",
                    "start": start, "end": end, "mode": mode, "measures": measures,
                })
        log.info("  上市處置股：%d 檔", len([x for x in result["disposal"] if x["market"]=="上市"]))
    except Exception as e:
        log.warning("  上市處置股抓取失敗：%s", e)

    # ── 2. 目前處置股（上櫃）─────────────────────────────────
    try:
        url2 = "https://www.tpex.org.tw/openapi/v1/tpex_disposal_announcement"
        req2 = urllib.request.Request(url2, headers={**HEADERS, "Referer": "https://www.tpex.org.tw/"})
        r2 = urllib.request.urlopen(req2, context=ctx, timeout=15)
        raw2 = r2.read()
        try:
            data2 = json.loads(raw2)
        except Exception as je:
            log.warning("  上櫃處置股：回應不是有效JSON（%s），HTTP狀態=%s，內容開頭200字：%s",
                        je, r2.status, raw2[:200])
            data2 = []
        if isinstance(data2, list):
            for item in data2:
                code = str(item.get("SecuritiesCompanyCode", item.get("code",""))).strip()
                name = str(item.get("CompanyName", item.get("name",""))).strip()
                if not _is_common_stock(code):
                    continue
                start = str(item.get("DisposalStartDate", item.get("StartDate",""))).strip()
                end   = str(item.get("DisposalEndDate",   item.get("EndDate",""))).strip()
                mode  = str(item.get("DisposalReason",    item.get("MatchingMethod",""))).strip()
                result["disposal"].append({
                    "code": code, "name": name, "market": "上櫃",
                    "start": start, "end": end, "mode": mode, "measures": "",
                })
        log.info("  上櫃處置股：%d 檔", len([x for x in result["disposal"] if x["market"]=="上櫃"]))
    except Exception as e:
        log.warning("  上櫃處置股抓取失敗：%s", e)

    # ── 3. 注意累計次數（上市）→ 預判明天進處置 ─────────────
    # TWSE noticeCount API 可能回傳 HTML，改用 notice（注意交易資訊）API
    try:
        # 嘗試另一個端點
        url3 = "https://www.twse.com.tw/rwd/zh/announcement/notice"
        req3 = urllib.request.Request(url3, headers=HEADERS)
        r3 = urllib.request.urlopen(req3, context=ctx, timeout=15)
        raw3 = r3.read()
        try:
            data3 = json.loads(raw3)
        except Exception:
            data3 = {}

        if data3.get("stat") == "OK" and data3.get("data"):
            fields3 = data3.get("fields", [])
            for row in data3["data"]:
                d3 = dict(zip(fields3, row)) if fields3 else {}
                # 找代號、名稱、連續次數、近10日、近30日
                code = str(d3.get("有價證券代號", d3.get("代號", row[0] if row else ""))).strip()
                name = str(d3.get("有價證券名稱", d3.get("名稱", row[1] if len(row)>1 else ""))).strip()
                consec = _safe_int(d3.get("連續次數", d3.get("連續", row[2] if len(row)>2 else 0)))
                cnt10  = _safe_int(d3.get("最近10個營業日", d3.get("10日", row[3] if len(row)>3 else 0)))
                cnt30  = _safe_int(d3.get("最近30個營業日", d3.get("30日", row[4] if len(row)>4 else 0)))

                reasons = []
                if consec == 2: reasons.append("連續2日（再1次→連續3日進處置）")
                if consec == 4: reasons.append("連續4日（再1次→連續5日進處置）")
                if cnt10 == 5:  reasons.append("近10日達5次（再1次→6次進處置）")
                if cnt30 == 11: reasons.append("近30日達11次（再1次→12次進處置）")

                if reasons:
                    result["warning"].append({
                        "code": code, "name": name, "market": "上市",
                        "consec": consec, "cnt10": cnt10, "cnt30": cnt30,
                        "reasons": reasons,
                    })
        log.info("  上市注意預警：%d 檔", len([x for x in result["warning"] if x["market"]=="上市"]))
    except Exception as e:
        log.warning("  上市注意累計次數抓取失敗：%s", e)

    # ── 4. 注意累計次數（上櫃）→ 預判明天進處置 ─────────────
    try:
        url4 = "https://www.tpex.org.tw/openapi/v1/tpex_notice_trade_info"
        req4 = urllib.request.Request(url4, headers={**HEADERS, "Referer": "https://www.tpex.org.tw/"})
        r4 = urllib.request.urlopen(req4, context=ctx, timeout=15)
        raw4 = r4.read()
        try:
            data4 = json.loads(raw4)
        except Exception as je:
            log.warning("  上櫃注意累計次數：回應不是有效JSON（%s），HTTP狀態=%s，內容開頭200字：%s",
                        je, r4.status, raw4[:200])
            data4 = []

        if isinstance(data4, list):
            for item in data4:
                code = str(item.get("SecuritiesCompanyCode", item.get("code",""))).strip()
                name = str(item.get("CompanyName", item.get("name",""))).strip()
                consec = _safe_int(item.get("ContinuousDays", item.get("consec",0)))
                cnt10  = _safe_int(item.get("Last10DaysCount", item.get("cnt10",0)))
                cnt30  = _safe_int(item.get("Last30DaysCount", item.get("cnt30",0)))

                reasons = []
                if consec == 2: reasons.append("連續2日（再1次→連續3日進處置）")
                if consec == 4: reasons.append("連續4日（再1次→連續5日進處置）")
                if cnt10 == 5:  reasons.append("近10日達5次（再1次→6次進處置）")
                if cnt30 == 11: reasons.append("近30日達11次（再1次→12次進處置）")

                if reasons:
                    result["warning"].append({
                        "code": code, "name": name, "market": "上櫃",
                        "consec": consec, "cnt10": cnt10, "cnt30": cnt30,
                        "reasons": reasons,
                    })
        log.info("  上櫃注意預警：%d 檔", len([x for x in result["warning"] if x["market"]=="上櫃"]))
    except Exception as e:
        log.warning("  上櫃注意累計次數抓取失敗：%s", e)

    # ── 5. 補上技術指標（股價、MA20、月線斜率、布林通道）───────
    if price_data:
        for section in ("disposal", "warning"):
            for item in result[section]:
                tech = _calc_disposal_technicals(item.get("code",""), price_data)
                item.update(tech)

    log.info("STEP 5f 完成：處置中 %d 檔，預警 %d 檔",
             len(result["disposal"]), len(result["warning"]))
    return result


def _is_common_stock(code: str) -> bool:
    """只保留4位數純數字的普通股，排除ETF(00開頭)、權證(0開頭非4位)等"""
    c = code.strip()
    if not c.isdigit():       return False  # 含英文字母 → 排除（債券、KY除外先不管）
    if len(c) != 4:           return False  # 非4碼 → 排除
    if c.startswith("00"):    return False  # ETF → 排除
    return True


def _calc_disposal_technicals(code: str, price_data: dict) -> dict:
    """計算處置股技術指標：股價、MA20、MA20斜率、布林通道、布林位階。
    布林位階：MA20=0，上軌=+10，下軌=-10，公式 = (price-MA20)/(上軌-MA20)*10
    """
    empty = {"price": None, "ma20": None, "ma20_slope": None,
             "bb_upper": None, "bb_lower": None, "bb_rank": None}
    try:
        d = price_data.get(code)
        if d is None:
            return empty
        close = d.get("close")
        if close is None or len(close) < 22:
            return empty

        closes   = close.values
        price    = round(float(closes[-1]), 2)
        ma20_val = float(closes[-20:].mean())
        std20    = float(closes[-20:].std())
        bb_upper = round(ma20_val + 2 * std20, 2)
        bb_lower = round(ma20_val - 2 * std20, 2)

        # 布林位階：MA20=0，上軌=+10，下軌=-10
        band = ma20_val - bb_lower  # = 2σ，大於0
        if band > 0:
            bb_rank = round((price - ma20_val) / band * 10, 1)
        else:
            bb_rank = None

        # MA20斜率（最近6個MA20點，取首尾差/5/MA20*100）
        if len(closes) >= 25:
            pts   = [closes[-(25-i):-(5-i) if (5-i)>0 else None].mean() for i in range(6)]
            slope = round((pts[-1] - pts[0]) / 5 / pts[0] * 100, 1)
        else:
            slope = None

        return {
            "price":      price,
            "ma20":       round(ma20_val, 2),
            "ma20_slope": slope,
            "bb_upper":   bb_upper,
            "bb_lower":   bb_lower,
            "bb_rank":    bb_rank,   # -10 ~ +10，MA20=0
        }
    except:
        return empty


def _safe_int(val) -> int:
    try: return int(str(val).replace(",","").strip())
    except: return 0



# 多組參數設定，跟回測腳本(compare_versions.py)的版本對齊
BREAKOUT_VERSIONS = [
    {"name": "原始版", "consol_days": 20, "max_range": 10.0, "vol_multiple": 1.2},
    {"name": "嚴格版", "consol_days": 20, "max_range": 5.0,  "vol_multiple": 1.5},
    {"name": "寬鬆版", "consol_days": 15, "max_range": 12.0, "vol_multiple": 1.2},
]
BREAKOUT_LOOKBACK_DAYS  = 5     # 往回找最近幾個交易日內曾出現的突破訊號
BREAKOUT_AVG_VOL_DAYS   = 20    # 均量計算天數
BREAKOUT_MIN_AVG_VOL    = 500   # 最低日均量（張），過濾殭屍股
BREAKOUT_MAX_EXTEND_PCT = 8.0   # 進場緩衝上限：股價距支撐位不可超過此%，避免追高

# ETF 判斷關鍵字（跟 VCP / 選股模型 / 過熱警示模組的判斷邏輯一致）
BREAKOUT_ETF_KW = {"ETF","基金","永續","高息","高股息","月配","季配","債券","期貨","正2","反1"}

def _is_etf(code: str, name: str = "") -> bool:
    return code.startswith("00") or any(k in name for k in BREAKOUT_ETF_KW)


def _scan_breakout_one(df: pd.DataFrame, ver: dict):
    """
    對單一股票的 OHLCV DataFrame（欄位：high, low, close, volume，依日期排序）
    用單一參數版本掃描，回傳最近一次「仍有效」的突破訊號 dict，或 None。
    """
    consol = ver["consol_days"]
    maxrng = ver["max_range"]
    volmul = ver["vol_multiple"]

    n = len(df)
    if n < consol + BREAKOUT_AVG_VOL_DAYS + BREAKOUT_LOOKBACK_DAYS + 2:
        return None

    avg_vol_recent = df["volume"].tail(60).mean()
    if pd.isna(avg_vol_recent) or avg_vol_recent < BREAKOUT_MIN_AVG_VOL * 1000:
        return None

    today_idx = n - 1
    found = None

    for offset in range(0, BREAKOUT_LOOKBACK_DAYS):
        i = today_idx - offset
        if i - consol - BREAKOUT_AVG_VOL_DAYS < 0:
            break

        w   = df.iloc[i - consol:i]
        tod = df.iloc[i]
        hh  = w["high"].max()
        ll  = w["low"].min()
        if ll <= 0 or pd.isna(hh) or pd.isna(ll):
            continue
        rng = (hh - ll) / ll * 100
        avg_v = df["volume"].iloc[i - BREAKOUT_AVG_VOL_DAYS:i].mean()
        vr  = tod["volume"] / avg_v if avg_v and avg_v > 0 else 0

        if rng <= maxrng and tod["close"] > hh and vr >= volmul:
            found = {
                "breakout_idx":   i,
                "breakout_date":  df.index[i].strftime("%Y-%m-%d"),
                "breakout_close": round(float(tod["close"]), 2),
                "support":        round(float(hh), 2),
                "vol_ratio":      round(float(vr), 2),
                "range_pct":      round(float(rng), 2),
                "days_ago":       offset,
            }
            break

    if found is None:
        return None

    support = found["support"]
    path = df["close"].iloc[found["breakout_idx"]:today_idx + 1]
    if (path < support).any():
        return None  # 已跌破支撐 = 假突破，排除

    current_close = float(df["close"].iloc[today_idx])
    extend_pct = round((current_close - support) / support * 100, 2)
    if extend_pct > BREAKOUT_MAX_EXTEND_PCT:
        return None  # 延伸過大 = 追高風險，排除

    found["current_close"] = round(current_close, 2)
    found["extend_pct"]    = extend_pct
    found["is_today"]      = (found["days_ago"] == 0)
    return found


# ════════════════════════════════════════════════════════════
# 近5日投本比／外本比變化量（供多個分頁共用）
# 定義：近5日投信/外資買賣超張數合計 ÷ 流通股數(張) × 100
# ════════════════════════════════════════════════════════════
RATIO5D_WINDOW_DAYS = 5


def load_shares_cache() -> dict:
    """讀 shares_cache.json，回傳 {code: 流通股數(張)}。root目錄優先，找不到才用output/"""
    path = BASE_DIR / "shares_cache.json"
    if not path.exists():
        path = SHARES_CACHE_PATH
    if not path.exists():
        log.warning("找不到 shares_cache.json，投本比/外本比欄位將無法計算")
        return {}
    try:
        d = json.loads(path.read_text(encoding="utf-8"))
        shares = d.get("shares", d) if isinstance(d, dict) else {}
        return {str(k): float(v) for k, v in shares.items() if isinstance(v, (int, float))}
    except Exception as e:
        log.warning("讀取 shares_cache.json 失敗：%s", e)
        return {}


def compute_ratio5d_map(insti_history_df: pd.DataFrame, shares_map: dict) -> dict:
    """
    計算每支股票「近5日投本比變化量」跟「近5日外本比變化量」(百分點)。
    回傳 {code: {"trust_ratio_5d_chg": x, "foreign_ratio_5d_chg": y}}
    """
    result = {}
    if insti_history_df is None or insti_history_df.empty or not shares_map:
        return result
    df = insti_history_df.copy()
    df["date"] = df["date"].astype(str)
    df["foreign_net"] = pd.to_numeric(df["foreign_net"], errors="coerce").fillna(0)
    df["trust_net"] = pd.to_numeric(df["trust_net"], errors="coerce").fillna(0)
    for code, grp in df.groupby("code"):
        shares = shares_map.get(str(code))
        if not shares or shares <= 0:
            continue
        grp = grp.sort_values("date").tail(RATIO5D_WINDOW_DAYS)
        trust_sum = float(grp["trust_net"].sum())
        foreign_sum = float(grp["foreign_net"].sum())
        result[str(code)] = {
            "trust_ratio_5d_chg": round(trust_sum / shares * 100, 3),
            "foreign_ratio_5d_chg": round(foreign_sum / shares * 100, 3),
        }
    return result


def _enrich_ratio5d(items: list, ratio_map: dict):
    """把 trust_ratio_5d_chg / foreign_ratio_5d_chg 兩個欄位掛到清單裡每個項目上（就地修改）"""
    if not ratio_map:
        for it in items:
            it["trust_ratio_5d_chg"] = None
            it["foreign_ratio_5d_chg"] = None
        return
    for it in items:
        code = it.get("code")
        info = ratio_map.get(code, {})
        it["trust_ratio_5d_chg"] = info.get("trust_ratio_5d_chg")
        it["foreign_ratio_5d_chg"] = info.get("foreign_ratio_5d_chg")


def step_v8_scan(price_data: dict, insti_history_df: pd.DataFrame,
                  market_above_ma60: bool = True,
                  info_map: dict = None) -> list:
    """
    v8 策略選股掃描（每日盤後執行）
    ────────────────────────────────────────────────────────
    進場條件：
      1. MA20 斜率 > 1%/日（月線明顯向上）
      2. 布林位階 < 4（未追高）
      3. chip_60d 在 1.4~2.0 OR 3.0~3.6（法人支撐，排除中間差區間）
      4. market_above_ma60=True（大盤 ^TWII > MA60）

    出場參考（僅顯示，不在掃描中判斷）：
      跌破 MA10（進場後 10 日緩衝）/ 保底 -8% / 超時 120 日

    回傳：符合條件的股票清單（list of dict）
    """
    import numpy as np
    if info_map is None:
        info_map = {}

    MA_PERIOD    = 20
    MA10_PERIOD  = 10
    BB_STD       = 2.0
    SLOPE_MIN    = 1.0        # %/日
    BB_MAX       = 4.0        # 布林位階上限
    CHIP_WINDOWS = [1, 5, 10, 20, 60]

    # chip_60d 有效區間（排除 2.0~3.0）
    def chip_ok(c60):
        if c60 is None or (isinstance(c60, float) and np.isnan(c60)):
            return False
        return (1.4 <= c60 <= 2.0) or (3.0 <= c60 <= 3.6)

    if not market_above_ma60:
        log.info("  v8掃描：大盤在MA60之下，跳過（market_above_ma60=False）")
        return []

    # ── 預建法人 lookup ──
    insti_lookup = {}
    if not insti_history_df.empty:
        df_i = insti_history_df.copy()
        df_i["date"]        = pd.to_datetime(df_i["date"], format="%Y%m%d", errors="coerce")
        df_i["foreign_net"] = pd.to_numeric(df_i["foreign_net"], errors="coerce").fillna(0)
        df_i["trust_net"]   = pd.to_numeric(df_i["trust_net"],   errors="coerce").fillna(0)
        df_i["insti_net"]   = df_i["foreign_net"] + df_i["trust_net"]
        for code, grp in df_i.groupby("code"):
            s = grp.set_index("date")["insti_net"]
            insti_lookup[code] = s[~s.index.duplicated(keep="last")]

    results = []

    for code, d in price_data.items():
        try:
            prices = d.get("close") if isinstance(d, dict) else d.get("Close")
            volume = d.get("volume") if isinstance(d, dict) else d.get("Volume")
            if prices is None or len(prices) < MA_PERIOD + max(CHIP_WINDOWS) + 5:
                continue

            # ── 技術指標 ──
            ma20  = prices.rolling(MA_PERIOD).mean()
            std20 = prices.rolling(MA_PERIOD).std()
            upper = ma20 + BB_STD * std20
            lower = ma20 - BB_STD * std20

            # MA20 斜率（%/日）
            slope = ((ma20 - ma20.shift(1)) / ma20.shift(1) * 100).iloc[-1]

            # 布林位階
            close_last = float(prices.iloc[-1])
            ma20_last  = float(ma20.iloc[-1])
            band_u = float((upper - ma20).iloc[-1])
            band_l = float((ma20 - lower).iloc[-1])
            if close_last >= ma20_last:
                bb_rank = (close_last - ma20_last) / band_u * 10 if band_u > 0 else None
            else:
                bb_rank = (close_last - ma20_last) / band_l * 10 if band_l > 0 else None

            # MA10（出場參考）
            ma10_last = float(prices.rolling(MA10_PERIOD).mean().iloc[-1])

            # ── 條件1+2 ──
            if pd.isna(slope) or slope <= SLOPE_MIN:
                continue
            if bb_rank is None or pd.isna(bb_rank) or bb_rank >= BB_MAX:
                continue

            # ── 籌碼計算 ──
            chip_60d = None
            if code in insti_lookup and volume is not None and not volume.empty:
                insti_s = insti_lookup[code]
                vol_lot = volume / 1000
                aligned = pd.DataFrame({"insti": insti_s, "vol": vol_lot})
                aligned = aligned.reindex(prices.index)
                aligned["insti"] = aligned["insti"].fillna(0)
                chip_1d = aligned["insti"] / aligned["vol"].replace(0, np.nan)
                chip_60d_s = chip_1d.rolling(60, min_periods=60).sum()
                if not chip_60d_s.empty and not pd.isna(chip_60d_s.iloc[-1]):
                    chip_60d = float(chip_60d_s.iloc[-1])

            # ── 條件3 ──
            if not chip_ok(chip_60d):
                continue

            # ── 通過所有條件，加入結果 ──
            chg_pct = round((close_last / float(prices.iloc[-2]) - 1) * 100, 2) \
                      if len(prices) >= 2 else None
            vol_k   = round(float(volume.iloc[-1]) / 1000, 0) \
                      if volume is not None and not volume.empty else None

            results.append({
                "code":       code,
                "name":       info_map.get(code, {}).get("name", ""),
                "price":      round(close_last, 2),
                "chg_pct":    chg_pct,
                "ma20":       round(ma20_last, 2),
                "ma20_slope": round(float(slope), 2),
                "bb_rank":    round(float(bb_rank), 2),
                "chip_60d":   round(chip_60d, 3),
                "ma10":       round(ma10_last, 2),
                "volume_k":   vol_k,
                "chip_zone":  "低位(1.4~2.0)" if chip_60d <= 2.0 else "高位(3.0~3.6)",
            })

        except Exception as e:
            log.debug("v8掃描 %s 錯誤：%s", code, e)

    # 排序：斜率大的優先
    results.sort(key=lambda x: -x["ma20_slope"])
    log.info("STEP v8掃描完成：%d 支符合條件", len(results))
    return results



# ════════════════════════════════════════════════════════════
# 創高量縮 選股策略
# 條件：①近2週內曾創一年新高 ②股價>MA10>MA20(多頭排列) ③近一個月均量>300張(流動性)
#      ④今日成交量 < 昨日成交量×0.7(單日急速量縮)
# ════════════════════════════════════════════════════════════
NH_MA_FAST, NH_MA_SLOW = 10, 20
NH_BB_STD = 2.0   # 布林通道標準差倍數（以MA10為中軌：MA10=0，上軌=+10，下軌=-10）
NH_NEWHIGH_LOOKBACK_DAYS = 252
NH_NEWHIGH_WINDOW_DAYS = 10
NH_NEWHIGH_TOLERANCE = 0.997
NH_MIN_AVG_VOLUME_LOTS = 300
NH_VOL_BASE_DAYS = 20
NH_DAY_CONTRACT_RATIO = 0.7


def step_newhigh_contraction_scan(price_data: dict, info_map: dict) -> list:
    """創高量縮選股：近2週曾創一年新高 + 股價多頭排列(股價>MA10>MA20) + 流動性 + 單日急速量縮"""
    results = []
    for code, d in price_data.items():
        try:
            name = info_map.get(code, {}).get("name", "")
            if _is_etf(code, name):
                continue

            prices = d.get("close")
            volume = d.get("volume")
            min_len_needed = max(NH_NEWHIGH_LOOKBACK_DAYS, NH_VOL_BASE_DAYS) + 10
            if prices is None or len(prices) < min_len_needed:
                continue
            if volume is None or len(volume) < NH_VOL_BASE_DAYS + 1:
                continue

            ma10 = prices.rolling(NH_MA_FAST).mean()
            ma20 = prices.rolling(NH_MA_SLOW).mean()
            latest = float(prices.iloc[-1])
            m10, m20 = float(ma10.iloc[-1]), float(ma20.iloc[-1])
            if any(pd.isna(x) for x in (m10, m20)):
                continue

            # 布林位階（以MA10為中軌：MA10=0，上軌=+10，下軌=-10）
            std10 = float(prices.rolling(NH_MA_FAST).std().iloc[-1])
            upper10 = m10 + NH_BB_STD * std10
            lower10 = m10 - NH_BB_STD * std10
            band_u10 = upper10 - m10
            band_l10 = m10 - lower10
            if latest >= m10:
                bb_rank10 = (latest - m10) / band_u10 * 10 if band_u10 > 0 else 0.0
            else:
                bb_rank10 = (latest - m10) / band_l10 * 10 if band_l10 > 0 else 0.0

            # ② 股價>MA10>MA20 多頭排列
            if not (latest > m10 > m20):
                continue

            # ① 近2週內曾創一年新高
            roll_high = prices.rolling(NH_NEWHIGH_LOOKBACK_DAYS, min_periods=NH_NEWHIGH_LOOKBACK_DAYS).max()
            recent_close = prices.tail(NH_NEWHIGH_WINDOW_DAYS)
            recent_roll_high = roll_high.tail(NH_NEWHIGH_WINDOW_DAYS)
            hit_mask = (recent_close >= recent_roll_high * NH_NEWHIGH_TOLERANCE) & recent_roll_high.notna()
            if not bool(hit_mask.any()):
                continue
            hit_dates = hit_mask[hit_mask].index
            latest_hit_date = hit_dates.max()
            days_since_1y_high = int(prices.index.get_loc(prices.index[-1]) - prices.index.get_loc(latest_hit_date))

            # ③ 流動性：近一個月均量 > 300張
            vol_base = float(volume.iloc[-NH_VOL_BASE_DAYS:].mean())
            if vol_base / 1000 < NH_MIN_AVG_VOLUME_LOTS:
                continue

            # ④ 今日成交量 < 昨日成交量 × 0.7
            vol_today = float(volume.iloc[-1])
            vol_yesterday = float(volume.iloc[-2])
            if vol_yesterday <= 0 or vol_today >= vol_yesterday * NH_DAY_CONTRACT_RATIO:
                continue

            results.append({
                "code": code,
                "name": name,
                "market": info_map.get(code, {}).get("market", ""),
                "industry": info_map.get(code, {}).get("industry", "其他"),
                "price": round(latest, 2),
                "ma10": round(m10, 2), "ma20": round(m20, 2),
                "bb_rank10": round(bb_rank10, 2),
                "vol_base_k": round(vol_base / 1000, 1),
                "vol_today_k": round(vol_today / 1000, 1),
                "vol_yesterday_k": round(vol_yesterday / 1000, 1),
                "vol_day_ratio": round(vol_today / vol_yesterday, 2) if vol_yesterday else None,
                "days_since_1y_high": days_since_1y_high,
            })
        except Exception as e:
            log.debug("創高量縮掃描 %s 錯誤：%s", code, e)

    # 排序：越近期創高、量縮越明顯的優先
    results.sort(key=lambda x: (x["days_since_1y_high"], x["vol_day_ratio"] if x["vol_day_ratio"] is not None else 1))
    log.info("STEP 創高量縮掃描完成：%d 支符合條件", len(results))
    return results


def step_breakout(price_data: dict, stock_list: pd.DataFrame) -> dict:
    """
    盤整突破即時掃描。重用 price_data（不重新下載）。
    回傳結構：
      {
        "generated_at": "...",
        "params": {...},
        "versions": {
          "原始版": {"today_signals":[...], "recent_signals":[...]},
          ...
        }
      }
    """
    log.info("STEP 5e: 盤整突破即時掃描...")

    info_map = {}
    if stock_list is not None and not stock_list.empty and "code" in stock_list.columns:
        info_map = stock_list.set_index("code").to_dict("index")

    results_by_version = {v["name"]: [] for v in BREAKOUT_VERSIONS}

    for code, d in price_data.items():
        close = d.get("close")
        if close is None or len(close) < 60:
            continue
        # 用 close 的 index 為準，重新對齊 high/low/volume（各自獨立 dropna 過，長度可能不同）
        df = pd.DataFrame({"close": close})
        for col in ("high", "low", "volume"):
            s = d.get(col)
            if s is not None:
                df[col] = s
        df = df.dropna(subset=["close", "high", "low", "volume"])
        if len(df) < 60:
            continue

        meta = info_map.get(code, {})
        name = meta.get("name", "")
        market = meta.get("market", "")

        if _is_etf(code, name):
            continue

        for ver in BREAKOUT_VERSIONS:
            res = _scan_breakout_one(df, ver)
            if res:
                res["code"]   = code
                res["name"]   = name
                res["market"] = market
                results_by_version[ver["name"]].append(res)

    output = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "params": {
            "lookback_days":   BREAKOUT_LOOKBACK_DAYS,
            "max_extend_pct":  BREAKOUT_MAX_EXTEND_PCT,
            "avg_vol_days":    BREAKOUT_AVG_VOL_DAYS,
            "min_avg_vol":     BREAKOUT_MIN_AVG_VOL,
            "versions": BREAKOUT_VERSIONS,
        },
        "versions": {}
    }

    total_today = 0
    for vname, items in results_by_version.items():
        items_sorted = sorted(items, key=lambda x: x["extend_pct"])
        today_signals  = [x for x in items_sorted if x["is_today"]]
        recent_signals = [x for x in items_sorted if not x["is_today"]]
        total_today += len(today_signals)

        output["versions"][vname] = {
            "total_count":    len(items_sorted),
            "today_count":    len(today_signals),
            "recent_count":   len(recent_signals),
            "today_signals":  today_signals,
            "recent_signals": recent_signals,
        }

    log.info("  盤整突破掃描完成：今日訊號合計 %d 檔（跨 %d 組參數）",
              total_today, len(BREAKOUT_VERSIONS))

    return output


# ════════════════════════════════════════════════════════════
# STEP 6：合併輸出 dashboard.json
# ════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════
# Telegram 通知
# ════════════════════════════════════════════════════════════

def _tg_send(text: str):
    """發送 Telegram 訊息，失敗只 log 不中斷主程式"""
    if not TG_ENABLED:
        return
    # Telegram 單則訊息上限 4096 字元，超過自動截斷並加提示
    if len(text) > 4000:
        text = text[:3950] + "\n\n... （訊息過長已截斷，請開儀表板查看完整名單）"
    try:
        url  = f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage"
        data = json.dumps({
            "chat_id": TG_CHAT_ID,
            "text":    text,
        }).encode()
        req = urllib.request.Request(
            url, data=data,
            headers={"Content-Type": "application/json"})
        urllib.request.urlopen(req, timeout=10)
        log.info("Telegram 通知已發送")
    except Exception as e:
        log.warning("Telegram 通知失敗（不影響主程式）：%s", e)


def _esc(text) -> str:
    """純文字模式不需轉義，直接回傳字串"""
    return str(text)


def _fmt_stock(s: dict, extra: str = "") -> str:
    code = str(s.get("code", ""))
    name = str(s.get("name", ""))
    return f"  {code} {name}{extra}"


def _capital_tag(code: str, strategy: str, shares_map: dict) -> str:
    """回傳這支股票的股本(億元)標註文字，方便在 Telegram 訊息裡直接看到符不符合股本版甜蜜點"""
    cap = _capital_billion(code, shares_map)
    if cap is None:
        return "  股本:—"
    return f"  股本:{cap:.0f}億"


def _send_tg_all_modules(
    vcp_results, breakout, disposition, v8_scan, newhigh_contraction_scan,
    market_pivot, margin_data, elapsed_sec, margin_maintenance=None, inst_summary=None, fear_greed=None
):
    """自動發送的選股口徑對齊「虛擬持倉-股本版」：大盤轉折+融資餘額 → 盤整突破-嚴格版(股本10-30億)
    → 創高量縮(篩選+股本100-500億) → MA20斜率策略(股本30-100億) → 處置機會股(不設股本濾網)，共5則訊息。
    VCP-強勢VCP 因為不在回測前三名，股本版沒有這個類別，這裡也一併移除，避免自動發送跟實際會建倉的
    虛擬持倉股本版對不起來。股本濾網用 CAPITAL_STRATEGY_RANGES／capital_bucket_ok，跟 run_all.py
    其他地方（Step 5.9b）共用同一套規則，改甜蜜點只要改那邊，這裡會自動跟著變。"""
    if not TG_ENABLED:
        return

    today = datetime.now().strftime("%Y-%m-%d")
    header = f"📊 台股儀表板 {today}  ⏱{elapsed_sec//60}分{elapsed_sec%60}秒\n"

    shares_map = load_shares_cache()
    if not shares_map:
        log.warning("TG 通知：找不到 shares_cache.json，股本濾網本次無法套用，選股清單會是空的")

    messages = []

    # ── 0. 大盤轉折 + 融資餘額（純資訊，不套股本濾網）──────────────
    try:
        lines = [header, "📈 大盤總覽"]
        mp = market_pivot or {}
        daily = mp.get("daily") or []
        real = [d for d in daily if d.get("open") is not None]
        if real:
            last = real[-1]
            chg = last.get("change") or 0
            bull = chg >= 0
            pivot = last.get("pivot")
            above = (pivot is not None) and (last.get("close", 0) > pivot)
            lines.append(
                f"\n大盤：{last.get('close')}　{'+' if bull else ''}{chg}"
                f"（{'▲多方轉折之上' if above else '▼空方轉折之下' if pivot is not None else ''}"
                f" {pivot if pivot is not None else ''}）"
            )
        else:
            lines.append("\n大盤：（無轉折資料）")

        md = margin_data or {}
        if md.get("today") is not None:
            chg_amt = md.get("change_amt", 0)
            chg_pct = md.get("change_pct", 0)
            arrow = "🔻" if chg_amt < 0 else "🔺"
            lines.append(f"融資餘額：{md['today']:,} 張　{arrow}{chg_amt:+,}張（{chg_pct:+.2f}%）")
        else:
            lines.append("融資餘額：（無資料，可能是假日）")

        mm = margin_maintenance or {}
        if mm.get("ratio") is not None:
            ratio = mm["ratio"]
            note = "偏安全" if ratio >= 150 else ("中性偏緊" if ratio >= 120 else "偏緊注意斷頭")
            lines.append(f"融資維持率(估算)：{ratio:.2f}%（{note}）")

        bd = (inst_summary or {}).get("breadth") or {}
        if bd.get("total"):
            lines.append(f"漲跌家數：🔴{bd.get('up',0)}漲　🟢{bd.get('down',0)}跌　{bd.get('flat',0)}平")

        fg = fear_greed or {}
        if fg.get("score") is not None:
            lines.append(f"恐慌貪婪指數：{fg['score']:.1f}分（{fg.get('label','')}）")

        messages.append("\n".join(lines))
    except Exception as e:
        log.warning("TG 大盤總覽組裝失敗：%s", e)

    # ── 1. 盤整突破（僅嚴格版，股本要 10-30億）──────────────────
    try:
        lo, hi = CAPITAL_STRATEGY_RANGES["盤整突破-嚴格版"]
        bo = breakout or {}
        vd = bo.get("versions", {}).get("嚴格版", {})
        today_sigs  = [s for s in (vd.get("today_signals")  or []) if capital_bucket_ok(s["code"], "盤整突破-嚴格版", shares_map)]
        recent_sigs = [s for s in (vd.get("recent_signals") or []) if capital_bucket_ok(s["code"], "盤整突破-嚴格版", shares_map)]
        lines = [header, f"📐 盤整突破【嚴格版】（股本{lo}-{hi}億）"]
        if today_sigs or recent_sigs:
            if today_sigs:
                lines.append(f"\n🟢 今日剛突破（{len(today_sigs)} 檔）")
                for s in today_sigs:
                    lines.append(_fmt_stock(s,
                        f"  支撐{_esc(s.get('support',''))} 延伸{_esc(s.get('extend_pct',''))}% 量{_esc(s.get('vol_ratio',''))}x"
                        + _capital_tag(s["code"], "盤整突破-嚴格版", shares_map)))
            if recent_sigs:
                lines.append(f"\n🔵 近5日仍有效（{len(recent_sigs)} 檔）")
                for s in recent_sigs:
                    lines.append(_fmt_stock(s,
                        f"  {_esc(s.get('days_ago',''))}天前 支撐{_esc(s.get('support',''))} 延伸{_esc(s.get('extend_pct',''))}%"
                        + _capital_tag(s["code"], "盤整突破-嚴格版", shares_map)))
        else:
            lines.append("（今日無符合條件）")
        messages.append("\n".join(lines))
    except Exception as e:
        log.warning("TG 盤整突破組裝失敗：%s", e)

    # ── 2. 創高量縮（布林位階0~3 且 投本/外本比>0，股本要 100-500億）
    try:
        lo, hi = CAPITAL_STRATEGY_RANGES["創高量縮"]
        nhc = newhigh_contraction_scan or []
        nhc_filtered = [
            s for s in nhc
            if s.get("bb_rank10") is not None and 0 <= s["bb_rank10"] <= 3
            and (
                (s.get("trust_ratio_5d_chg") is not None and s["trust_ratio_5d_chg"] > 0)
                or (s.get("foreign_ratio_5d_chg") is not None and s["foreign_ratio_5d_chg"] > 0)
            )
            and capital_bucket_ok(s["code"], "創高量縮", shares_map)
        ]
        lines = [header, f"🆕 創高量縮（布林位階0~3 且 投本/外本比>0，股本{lo}-{hi}億，{len(nhc_filtered)} 檔）"]
        if nhc_filtered:
            for s in nhc_filtered:
                tr = s.get("trust_ratio_5d_chg")
                fr = s.get("foreign_ratio_5d_chg")
                parts = [f"  BB{s.get('bb_rank10','')}"]
                if tr is not None:
                    parts.append(f"  投本Δ{tr:+.3f}%")
                if fr is not None:
                    parts.append(f"  外本Δ{fr:+.3f}%")
                parts.append(_capital_tag(s["code"], "創高量縮", shares_map))
                lines.append(_fmt_stock(s, "".join(parts)))
        else:
            lines.append("（今日無符合條件）")
        messages.append("\n".join(lines))
    except Exception as e:
        log.warning("TG 創高量縮組裝失敗：%s", e)

    # ── 3. MA20斜率策略（v8，股本要 30-100億）───────────────────
    try:
        lo, hi = CAPITAL_STRATEGY_RANGES["MA20斜率策略"]
        v8 = [s for s in (v8_scan or []) if capital_bucket_ok(s["code"], "MA20斜率策略", shares_map)]
        lines = [header, f"🎯 MA20斜率策略（股本{lo}-{hi}億，{len(v8)} 檔）"]
        if v8:
            for s in v8:
                zone = "🟢低位" if "低位" in s.get("chip_zone","") else "🔵高位"
                lines.append(_fmt_stock(s,
                    f"  斜率+{s.get('ma20_slope','')}%  BB{s.get('bb_rank',''):+.1f}"
                    f"  籌碼{s.get('chip_60d',0):.2f}{zone}" + _capital_tag(s["code"], "MA20斜率策略", shares_map)))
        else:
            lines.append("（今日無符合條件）")
        messages.append("\n".join(lines))
    except Exception as e:
        log.warning("TG MA20斜率策略組裝失敗：%s", e)

    # ── 4. 處置機會股（月線斜率>1 且 布林位階<4，沒有回測資料，不設股本濾網）
    try:
        dp = disposition or {}
        disposal = dp.get("disposal", [])
        opp = [
            r for r in disposal
            if r.get("ma20_slope") is not None and r["ma20_slope"] > 1
            and r.get("bb_rank") is not None and r["bb_rank"] < 4
        ]
        lines = [header, f"⚠️ 處置機會股（月線斜率>1 且 布林位階<4，不設股本濾網，{len(opp)} 檔）"]
        if opp:
            for r in opp:
                lines.append(_fmt_stock(r,
                    f"  斜率+{r.get('ma20_slope','')}%  BB{r.get('bb_rank','')}  {r.get('mode','')}"))
        else:
            lines.append("（今日無符合條件）")
        messages.append("\n".join(lines))
    except Exception as e:
        log.warning("TG 處置機會股組裝失敗：%s", e)

    log.info("開始發送 Telegram 通知，共 %d 則...", len(messages))
    for i, msg in enumerate(messages):
        _tg_send(msg)
        if i < len(messages) - 1:
            time.sleep(0.5)
    log.info("Telegram 通知發送完畢")



# ════════════════════════════════════════════════════════════
# 每日選股歷史紀錄（供「出場邏輯回測工具」累積歷史 entries.csv）
# ════════════════════════════════════════════════════════════

def _clean_code(raw: str) -> str:
    """把「2330.TW」/「2330.TWO」/「2330」統一清成 4 碼代號（先去 .TWO 再去 .TW，順序不能反）"""
    return str(raw).replace(".TWO", "").replace(".TW", "").strip()


def _filter_new_streak_rows(rows_to_check: list, old_df, date_str: str, max_gap_days: int = 3) -> list:
    """
    只保留「訊號剛出現」的列：如果同一支股票同一個策略，在過去 max_gap_days 天內
    的既有紀錄裡已經出現過，代表訊號還在延續中，不重複記錄。
    用於 VCP 這類「型態可能連續好幾天都成立」的策略，避免同一段走勢被重複採樣進回測。
    """
    if old_df is None or old_df.empty or not rows_to_check:
        return rows_to_check
    strategies_involved = {r["strategy"] for r in rows_to_check}
    sub = old_df[old_df["strategy"].isin(strategies_involved)]
    if sub.empty:
        return rows_to_check

    lookup = {}  # (code,strategy) -> 最近一次出現日期(Timestamp)
    sub_dates = pd.to_datetime(sub["date"], errors="coerce")
    for (code, strategy), d in zip(zip(sub["code"], sub["strategy"]), sub_dates):
        if pd.isna(d):
            continue
        key = (code, strategy)
        if key not in lookup or d > lookup[key]:
            lookup[key] = d

    today_dt = pd.Timestamp(date_str)
    kept = []
    skipped = 0
    for r in rows_to_check:
        key = (r["code"], r["strategy"])
        last_dt = lookup.get(key)
        if last_dt is not None and (today_dt - last_dt).days <= max_gap_days:
            skipped += 1
            continue
        kept.append(r)
    if skipped:
        log.info("  streak去重：%d 筆訊號因延續中的舊訊號被跳過（%s）",
                  skipped, "、".join(sorted(strategies_involved)))
    return kept


def log_daily_entries(strong_stocks: dict, newhigh: list, insti_ma: list,
                       vcp_results: list, minervini_results: list,
                       canslim_results: list, v8_scan: list,
                       breakout: dict, newhigh_contraction_scan: list = None, date_str: str = None) -> int:
    """
    把今天各策略選到的股票，各自標上 strategy 名稱，
    append 進 ENTRIES_HISTORY_CSV（date,code,strategy,entry_price）。
    用 (date,code,strategy) 去重，同一天重複執行 run_all.py 不會產生重複列。
    VCP 額外用 _filter_new_streak_rows 做「訊號延續中不重複記錄」的過濾。
    回傳新增的列數。
    """
    date_str = date_str or TODAY_STR
    rows = []

    old_df = None
    if ENTRIES_HISTORY_CSV.exists():
        try:
            old_df = pd.read_csv(ENTRIES_HISTORY_CSV, dtype={"code": str})
        except Exception as e:
            log.warning("  讀取既有 entries_history.csv 失敗（不影響本次寫入）：%s", e)

    # 強勢股
    for s in (strong_stocks or {}).get("stocks", []):
        code = _clean_code(s.get("代號", ""))
        price = s.get("最新收盤")
        if code and price is not None:
            rows.append({"date": date_str, "code": code, "strategy": "強勢股", "entry_price": price})

    # 創新高
    for r in (newhigh or []):
        code, price = r.get("code"), r.get("price")
        if code and price is not None:
            rows.append({"date": date_str, "code": _clean_code(code), "strategy": "創新高", "entry_price": price})

    # 連買+MA
    for r in (insti_ma or []):
        code, price = r.get("code"), r.get("price")
        if code and price is not None:
            rows.append({"date": date_str, "code": _clean_code(code), "strategy": "連買+MA", "entry_price": price})

    # VCP（強勢VCP / 廣義VCP 分開記，方便之後獨立比較；訊號延續中不重複記錄）
    vcp_rows = []
    for r in (vcp_results or []):
        code, price = r.get("code"), r.get("price")
        tier = r.get("tier", "VCP")
        if code and price is not None:
            vcp_rows.append({"date": date_str, "code": _clean_code(code), "strategy": f"VCP-{tier}", "entry_price": price})
    vcp_rows = _filter_new_streak_rows(vcp_rows, old_df, date_str, max_gap_days=3)
    rows.extend(vcp_rows)

    # 馬克維尼亞 Trend Template — 只記「完美趨勢」8/8全過的
    for r in (minervini_results or []):
        if r.get("passed", 0) < MINERVINI_MIN_PASSED:
            continue
        code, price = r.get("code"), r.get("price")
        if code and price is not None:
            rows.append({"date": date_str, "code": _clean_code(code), "strategy": "Minervini", "entry_price": price})

    # CANSLIM — 只記全條件 7/7 或 6/7 的
    for r in (canslim_results or []):
        if r.get("passed", 0) < CANSLIM_MIN_PASSED:
            continue
        code, price = r.get("code"), r.get("price")
        if code and price is not None:
            rows.append({"date": date_str, "code": _clean_code(code), "strategy": "CANSLIM", "entry_price": price})

    # v8 策略選股
    for r in (v8_scan or []):
        code, price = r.get("code"), r.get("price")
        if code and price is not None:
            rows.append({"date": date_str, "code": _clean_code(code), "strategy": "v8策略", "entry_price": price})

    # 創高量縮
    for r in (newhigh_contraction_scan or []):
        code, price = r.get("code"), r.get("price")
        if code and price is not None:
            rows.append({"date": date_str, "code": _clean_code(code), "strategy": "創高量縮", "entry_price": price})

    # 盤整突破（只記「今天」真正觸發的訊號，不記 recent_signals，避免同一根訊號被記錄好幾天）
    for vname, vdata in (breakout or {}).get("versions", {}).items():
        for r in vdata.get("today_signals", []):
            code, price = r.get("code"), r.get("current_close")
            if code and price is not None:
                rows.append({"date": date_str, "code": _clean_code(code),
                             "strategy": f"盤整突破-{vname}", "entry_price": price})

    if not rows:
        log.info("  今日選股紀錄：無資料可寫入")
        return 0

    new_df = pd.DataFrame(rows)

    if old_df is not None and not old_df.empty:
        combined = pd.concat([old_df, new_df], ignore_index=True)
        combined = combined.drop_duplicates(subset=["date", "code", "strategy"], keep="last")
    else:
        combined = new_df

    combined = combined.sort_values(["date", "strategy", "code"])
    combined.to_csv(ENTRIES_HISTORY_CSV, index=False, encoding="utf-8-sig")

    added = len(new_df.drop_duplicates(subset=["date", "code", "strategy"]))
    log.info("  今日選股紀錄：新增/更新 %d 筆 → %s（累積 %d 筆）",
             added, ENTRIES_HISTORY_CSV, len(combined))
    return added


def step6_merge(analysis, radar, strong_stocks, newhigh, insti_ma, vcp_results=None, minervini_results=None, canslim_results=None, cb_result=None, overheat=None, breakout=None, disposition=None, v8_scan=None, newhigh_contraction_scan=None, market_pivot=None, margin_data=None, kline_data=None, buy_alerts=None, sell_alerts=None, positions=None, positions_capital=None, buy_alerts_capital=None, sell_alerts_capital=None, briefing_picks_capital=None, business_cycle=None, inst_summary=None, margin_maintenance=None, fear_greed=None, vix=None, twvix=None, recent_highs=None, pe_river=None) -> Path:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    is_empty_stocks = (
        (isinstance(strong_stocks, dict) and not strong_stocks.get("stocks")) or
        (isinstance(strong_stocks, dict) and strong_stocks.get("note"))
    )

    if is_empty_stocks:
        # 優先從個別備份檔繼承（比 dashboard.json 更可靠）
        def _load_json(path):
            try:
                with open(path, encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return None

        ss_backup  = _load_json(OUTPUT_DIR / "strong_stocks.json")
        nh_backup  = _load_json(OUTPUT_DIR / "newhigh.json")
        im_backup  = _load_json(OUTPUT_DIR / "insti_ma.json")

        if ss_backup and ss_backup.get("stocks"):
            log.warning("STEP 6: 強勢股為空，從 strong_stocks.json 繼承（%s，%d 支）",
                        ss_backup.get("generated_at","?"), len(ss_backup["stocks"]))
            strong_stocks = ss_backup
            newhigh       = nh_backup  if nh_backup  else newhigh
            insti_ma      = im_backup  if im_backup  else insti_ma
        else:
            # 備份也沒有，嘗試從舊的 dashboard.json 繼承
            out_path = OUTPUT_DIR / "dashboard.json"
            old_dash = _load_json(out_path) or {}
            if old_dash and old_dash.get("strong_stocks", {}).get("stocks"):
                log.warning("STEP 6: 強勢股為空，從 dashboard.json 繼承（%s）",
                            old_dash.get("date","?"))
                strong_stocks = old_dash.get("strong_stocks", strong_stocks)
                newhigh       = old_dash.get("newhigh",       newhigh)
                insti_ma      = old_dash.get("insti_ma",      insti_ma)
            else:
                log.warning("STEP 6: 強勢股為空，且無可用備份資料")

    if market_pivot is None:
        # 若本次跳過/失敗，繼承舊 dashboard.json 的資料，避免分頁瞬間變空
        try:
            with open(OUTPUT_DIR / "dashboard.json", encoding="utf-8") as f:
                old_dash2 = json.load(f)
            market_pivot = old_dash2.get("market_pivot")
        except Exception:
            market_pivot = None

    if margin_data is None:
        try:
            with open(OUTPUT_DIR / "dashboard.json", encoding="utf-8") as f:
                old_dash3 = json.load(f)
            margin_data = old_dash3.get("margin_data")
        except Exception:
            margin_data = None

    if business_cycle is None:
        # 景氣對策信號每月才更新一次，本次抓取失敗/跳過時繼承舊資料，
        # 不要讓分頁瞬間變空（跟 market_pivot/margin_data 同樣的作法）
        try:
            with open(OUTPUT_DIR / "dashboard.json", encoding="utf-8") as f:
                old_dash4 = json.load(f)
            business_cycle = old_dash4.get("business_cycle")
        except Exception:
            business_cycle = None

    if inst_summary is None:
        try:
            with open(OUTPUT_DIR / "dashboard.json", encoding="utf-8") as f:
                old_dash5 = json.load(f)
            inst_summary = old_dash5.get("inst_summary")
        except Exception:
            inst_summary = None

    if margin_maintenance is None:
        try:
            with open(OUTPUT_DIR / "dashboard.json", encoding="utf-8") as f:
                old_dash6 = json.load(f)
            margin_maintenance = old_dash6.get("margin_maintenance")
        except Exception:
            margin_maintenance = None

    if fear_greed is None:
        try:
            with open(OUTPUT_DIR / "dashboard.json", encoding="utf-8") as f:
                old_dash7 = json.load(f)
            fear_greed = old_dash7.get("fear_greed")
        except Exception:
            fear_greed = None

    if vix is None:
        try:
            with open(OUTPUT_DIR / "dashboard.json", encoding="utf-8") as f:
                old_dash8 = json.load(f)
            vix = old_dash8.get("vix")
        except Exception:
            vix = None

    if twvix is None:
        try:
            with open(OUTPUT_DIR / "dashboard.json", encoding="utf-8") as f:
                old_dash9 = json.load(f)
            twvix = old_dash9.get("twvix")
        except Exception:
            twvix = None

    if recent_highs is None:
        try:
            with open(OUTPUT_DIR / "dashboard.json", encoding="utf-8") as f:
                old_dash10 = json.load(f)
            recent_highs = old_dash10.get("recent_highs")
        except Exception:
            recent_highs = None

    if pe_river is None:
        try:
            with open(OUTPUT_DIR / "dashboard.json", encoding="utf-8") as f:
                old_dash11 = json.load(f)
            pe_river = old_dash11.get("pe_river")
        except Exception:
            pe_river = None

    dashboard = {
        "generated_at":   ts,
        "date":           TODAY_STR,
        "analysis_result": analysis,
        "radar_multi":    radar,
        "strong_stocks":  strong_stocks,
        "newhigh":        newhigh,
        "insti_ma":       insti_ma,
        "cb_result":      cb_result,
        "vcp_result":       vcp_results or [],
        "minervini_result": minervini_results or [],
        "canslim_result":   canslim_results or [],
        "overheat":         overheat or [],
        "breakout":         breakout or {},
        "disposition":      disposition or {},
        "v8_scan":          v8_scan or [],
        "newhigh_contraction_scan": newhigh_contraction_scan or [],
        "market_pivot":     market_pivot or {},
        "margin_data":      margin_data or {},
        "kline_data":       kline_data or {},
        "buy_alerts":       buy_alerts or [],
        "sell_alerts":      sell_alerts or [],
        "positions":        positions or [],
        "positions_capital": positions_capital or [],
        "buy_alerts_capital":  buy_alerts_capital or [],
        "sell_alerts_capital": sell_alerts_capital or [],
        "briefing_picks_capital": briefing_picks_capital or [],
        "business_cycle": business_cycle,
        "inst_summary": inst_summary,
        "margin_maintenance": margin_maintenance,
        "fear_greed": fear_greed,
        "vix": vix,
        "twvix": twvix,
        "recent_highs": recent_highs,
        "pe_river": pe_river,
    }
    out_path = OUTPUT_DIR / "dashboard.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(dashboard, f, ensure_ascii=False, indent=2)
    size_kb = out_path.stat().st_size / 1024
    log.info("STEP 6：dashboard.json 輸出完成（%.0f KB）", size_kb)

    # 個別備份（供緊急還原用）
    if isinstance(strong_stocks, dict) and strong_stocks.get("stocks"):
        with open(OUTPUT_DIR / "strong_stocks.json", "w", encoding="utf-8") as f:
            json.dump(strong_stocks, f, ensure_ascii=False)
    if newhigh:
        with open(OUTPUT_DIR / "newhigh.json", "w", encoding="utf-8") as f:
            json.dump(newhigh, f, ensure_ascii=False)
    if insti_ma:
        with open(OUTPUT_DIR / "insti_ma.json", "w", encoding="utf-8") as f:
            json.dump(insti_ma, f, ensure_ascii=False)
    if breakout and breakout.get("versions"):
        with open(OUTPUT_DIR / "breakout.json", "w", encoding="utf-8") as f:
            json.dump(breakout, f, ensure_ascii=False)

    return out_path


# ════════════════════════════════════════════════════════════
# 主程式
# ════════════════════════════════════════════════════════════

def run_scan_stage(args) -> dict:
    """執行選股掃描階段：Step1~6，輸出 dashboard.json，回傳所有結果供通知階段使用"""
    t0 = time.time()
    sep = "="*55
    log.info(sep)
    log.info("台股整合儀表板  %s", datetime.now().strftime("%Y-%m-%d %H:%M"))
    log.info(sep)

    # Step 1: 更新法人CSV
    _, inst_summary = step1_update_market_flow(skip=args.skip_flow)

    # Step 1b: 大盤月轉折更新（K線圖.html）
    market_pivot = step1b_pivot_update(skip=getattr(args, "skip_pivot", False))
    recent_highs = None
    if market_pivot and market_pivot.get("daily"):
        try:
            recent_highs = compute_recent_highs(market_pivot["daily"])
            if recent_highs:
                log.info("  [轉折] 近3月高點 %s(%s) 回撤%.2f%%／近6月高點 %s(%s) 回撤%.2f%%",
                          recent_highs.get("high_3m"), recent_highs.get("high_3m_date"),
                          recent_highs.get("drawdown_3m_pct", 0),
                          recent_highs.get("high_6m"), recent_highs.get("high_6m_date"),
                          recent_highs.get("drawdown_6m_pct", 0))
        except Exception as e:
            log.warning("  [轉折] 近3/6月高點計算失敗（不影響主流程）：%s", e)
            recent_highs = None

    # Step 1c: 融資餘額變化
    margin_data = step1c_margin_update(skip=getattr(args, "skip_margin", False))

    # Step 1e: 融資維持率估算（用 step1c 抓到的融資金額 + 逐股融資餘額 + 今日收盤價）
    margin_amount_today = margin_data.get("margin_amount_today") if margin_data else None
    margin_maintenance = step1e_margin_maintenance_ratio(
        margin_amount_today, skip=getattr(args, "skip_maintenance", False))

    # Step 1g: 恐慌貪婪指數（自製版，六項指標綜合）
    fear_greed = step1g_fear_greed_index(
        market_pivot, margin_data, margin_maintenance, inst_summary,
        skip=getattr(args, "skip_feargreed", False))

    # Step 1h: VIX恐慌指數（外部參考，不納入 fear_greed 計分）
    vix = step1h_vix(skip=getattr(args, "skip_vix", False))

    # Step 1i: 台指VIX（VIXTWN，台股原生恐慌指標，外部參考）
    twvix = step1i_twvix(skip=getattr(args, "skip_twvix", False))

    # Step 1j: 大盤本益比河流（每天自動累加一筆進 market_pe_history.csv）
    pe_river = step1j_market_pe_river(skip=getattr(args, "skip_pe_river", False))

    # Step 1d: 景氣對策信號（國發會，每月發布一次）
    business_cycle = step1d_business_cycle_signal(skip=getattr(args, "skip_cycle", False))

    # Step 2: 資金流向分析（快，不需 yfinance）
    analysis = step2_analysis()

    # Step 3: 多時框雷達（快，不需 yfinance）
    radar = step3_radar()

    # Step 4+5: 強勢股 + 創新高 + 連買MA（共用 yfinance）
    if not args.skip_yf:
        log.info("取得個股清單...")
        stock_list = _fetch_stock_list()
        log.info("下載股價歷史（%d 支）...", len(stock_list))
        price_data = _download_prices(stock_list, need_days=500)
    else:
        log.info("STEP 4+5 跳過（--skip-yf）")
        stock_list = pd.DataFrame(columns=["code","name","market","ticker"])
        price_data = {}

    strong_stocks, newhigh, insti_ma, vcp_results, minervini_results, canslim_results = step45_strong_and_screener(stock_list, price_data, args.skip_yf)

    # Step 5d: 過熱指標
    overheat = []
    if not args.skip_yf and price_data:
        overheat = step_overheat(price_data, strong_stocks, newhigh, insti_ma, vcp_results, stock_list)

    # ── 讀取投信/外資籌碼歷史 + 流通股數，計算「近5日投本比/外本比變化量」 ──
    # 優先讀根目錄（build_insti_history.py 預設輸出位置），fallback 到 output/ 子目錄（舊版路徑）
    insti_path_v8 = BASE_DIR / "insti_history.csv"
    if not insti_path_v8.exists():
        insti_path_v8 = INSTI_CSV   # output/insti_history.csv
    insti_df_v8 = pd.read_csv(insti_path_v8, dtype={"code": str, "date": str}) \
                  if insti_path_v8.exists() else pd.DataFrame()
    if insti_df_v8.empty:
        log.warning("  找不到 insti_history.csv，v8選股籌碼計算 + 投本比/外本比欄位都會跳過")
    shares_map = load_shares_cache()
    ratio5d_map = compute_ratio5d_map(insti_df_v8, shares_map)
    log.info("  投本比/外本比：%d 檔股票有可計算的近5日變化量", len(ratio5d_map))

    # Step 5e: 盤整突破即時掃描（重用 price_data，不重新下載）
    breakout = {}
    if not args.skip_yf and price_data:
        breakout = step_breakout(price_data, stock_list)
        for vdata in breakout.get("versions", {}).values():
            _enrich_ratio5d(vdata.get("today_signals", []), ratio5d_map)
            _enrich_ratio5d(vdata.get("recent_signals", []), ratio5d_map)

    # Step 5f: 處置股 + 注意股預警
    disposition = step_disposition(price_data if not args.skip_yf else None)
    _enrich_ratio5d(disposition.get("disposal", []), ratio5d_map)
    _enrich_ratio5d(disposition.get("warning", []), ratio5d_map)

    # Step 5g: v8 策略選股掃描（月線斜率>1% + BB位階<4 + chip_60d甜蜜區間）
    v8_scan = []
    if not args.skip_yf and price_data:
        log.info("STEP 5g: v8 策略選股掃描...")
        # 判斷大盤是否在 MA60 之上
        try:
            taiex_v8    = yf.download("^TWII", period="4mo", auto_adjust=True, progress=False)
            taiex_c     = taiex_v8["Close"].squeeze().dropna()
            if isinstance(taiex_c.columns if hasattr(taiex_c, 'columns') else [], pd.MultiIndex):
                taiex_c = taiex_c.iloc[:, 0]
            mkt_ma60    = float(taiex_c.tail(60).mean()) if len(taiex_c) >= 60 else None
            mkt_above   = (mkt_ma60 is not None and float(taiex_c.iloc[-1]) > mkt_ma60)
            log.info("  大盤 ^TWII：最新=%.0f  MA60=%.0f  在MA60%s",
                     float(taiex_c.iloc[-1]), mkt_ma60 or 0,
                     "之上✅" if mkt_above else "之下⛔")
        except Exception as e:
            log.warning("  大盤資料失敗：%s，預設允許進場", e)
            mkt_above = True

        _v8_info = stock_list.set_index("code").to_dict("index") if not stock_list.empty else {}
        v8_scan = step_v8_scan(price_data, insti_df_v8,
                               market_above_ma60=mkt_above, info_map=_v8_info)
        _enrich_ratio5d(v8_scan, ratio5d_map)
        log.info("STEP 5g 完成：v8選股 %d 支", len(v8_scan))
    else:
        _v8_info = {}

    # Step 5i: 創高量縮選股（近2週曾創一年新高+多頭排列+流動性+單日急速量縮）
    newhigh_contraction_scan = []
    if not args.skip_yf and price_data:
        log.info("STEP 5i: 創高量縮選股掃描...")
        newhigh_contraction_scan = step_newhigh_contraction_scan(price_data, _v8_info)
        _enrich_ratio5d(newhigh_contraction_scan, ratio5d_map)
        log.info("STEP 5i 完成：創高量縮選股 %d 支", len(newhigh_contraction_scan))

    # Step 5.5: 可轉債篩選
    cb_result = None
    try:
        import importlib.util, sys
        cb_path = Path("cb_screener.py")
        if cb_path.exists():
            log.info("STEP 5.5: 可轉債篩選...")
            spec = importlib.util.spec_from_file_location("cb_screener", cb_path)
            cb_mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(cb_mod)
            cb_result = cb_mod.run()
            log.info("STEP 5.5 完成：突破轉換價 %d 筆", cb_result.get("above_count",0))
        else:
            log.info("STEP 5.5 跳過（cb_screener.py 不在同一資料夾）")
    except Exception as e:
        log.warning("STEP 5.5 可轉債篩選失敗：%s", e)

    # Step 5.9: 為「每日快報」的5個選股分類截取個股K線資料（不重新抓取，從已下載的price_data切片）
    # 同時用完全相同的篩選條件組成 briefing_picks，餵給虛擬持倉追蹤（買進/賣出訊號）
    kline_data = {}
    briefing_picks = []
    try:
        briefing_codes = set()

        if breakout and breakout.get("versions", {}).get("嚴格版"):
            v = breakout["versions"]["嚴格版"]
            for s in (v.get("today_signals") or []) + (v.get("recent_signals") or []):
                briefing_codes.add(s["code"])
                briefing_picks.append({"code": s["code"], "name": s.get("name",""), "strategy": "盤整突破-嚴格版"})

        if vcp_results:
            for r in vcp_results:
                if r.get("tier") == "強勢VCP":
                    briefing_codes.add(r["code"])
                    briefing_picks.append({"code": r["code"], "name": r.get("name",""), "strategy": "VCP-強勢VCP"})

        if newhigh_contraction_scan:
            for r in newhigh_contraction_scan:
                briefing_codes.add(r["code"])
                bb = r.get("bb_rank10")
                tr = r.get("trust_ratio_5d_chg")
                fr = r.get("foreign_ratio_5d_chg")
                if bb is not None and 0 <= bb <= 3 and ((tr is not None and tr > 0) or (fr is not None and fr > 0)):
                    briefing_picks.append({"code": r["code"], "name": r.get("name",""), "strategy": "創高量縮"})

        if v8_scan:
            for r in v8_scan:
                briefing_codes.add(r["code"])
                briefing_picks.append({"code": r["code"], "name": r.get("name",""), "strategy": "MA20斜率策略"})

        if disposition and disposition.get("disposal"):
            for r in disposition["disposal"]:
                briefing_codes.add(r["code"])
                slope, bb = r.get("ma20_slope"), r.get("bb_rank")
                if slope is not None and slope > 1 and bb is not None and bb < 4:
                    briefing_picks.append({"code": r["code"], "name": r.get("name",""), "strategy": "處置預警-機會股"})

        if briefing_codes and price_data:
            kline_data = build_kline_slices(price_data, briefing_codes, days=90)
            log.info("STEP 5.9: 每日快報K線資料截取完成，%d / %d 檔有資料", len(kline_data), len(briefing_codes))
    except Exception as e:
        log.warning("STEP 5.9 K線資料截取失敗（不影響主流程）：%s", e)

    # Step 5.9b: 虛擬持倉－股本版的訊號 = 從上面同一批訊號中：
    #   (A) 回測前三名策略（創高量縮/MA20斜率策略/盤整突破-嚴格版），股本要落在該策略最佳級距內才進場
    #   (B) CAPITAL_UNFILTERED_STRATEGIES 裡的策略（目前是「處置預警-機會股」），沒有回測資料可用，
    #       Kevin 指定保留，不套股本濾網、無條件跟著原版一起進場
    briefing_picks_capital = []
    try:
        shares_map = load_shares_cache()
        if shares_map:
            # 一次性把原版虛擬持倉「現有」的持股（不管持有中或已出場）也依股本規則補進股本版，
            # 不用等它們剛好又出現在今天的訊號裡才會被抓到；已經補過的不會重複補。
            backfilled = backfill_positions_capital(shares_map)
            if backfilled:
                log.info("STEP 5.9b: 從原版虛擬持倉補齊 %d 筆符合股本規則的既有部位", backfilled)

            # 逐策略統計「篩選前有幾筆訊號」跟「股本濾網後留下幾筆」，方便在 log 裡判斷
            # 某個策略今天是「根本沒訊號」還是「有訊號但股本沒對到甜蜜點」
            diag = {}
            for pick in briefing_picks:
                strat = pick["strategy"]
                if strat not in CAPITAL_STRATEGY_RANGES and strat not in CAPITAL_UNFILTERED_STRATEGIES:
                    continue
                d = diag.setdefault(strat, {"raw": 0, "kept": 0})
                d["raw"] += 1
                if strat in CAPITAL_UNFILTERED_STRATEGIES:
                    d["kept"] += 1
                    briefing_picks_capital.append(pick)
                elif capital_bucket_ok(pick["code"], strat, shares_map):
                    d["kept"] += 1
                    briefing_picks_capital.append(pick)
            for strat, d in diag.items():
                rng = CAPITAL_STRATEGY_RANGES.get(strat)
                rng_str = f"（甜蜜點 {rng[0]}-{rng[1]}億）" if rng else "（不套股本濾網）"
                log.info("STEP 5.9b   %s%s：今天原始訊號 %d 筆 → 股本符合 %d 筆",
                          strat, rng_str, d["raw"], d["kept"])
            log.info("STEP 5.9b: 股本版訊號 %d 筆（原始 %d 筆篩選後）", len(briefing_picks_capital), len(briefing_picks))
        else:
            log.warning("STEP 5.9b 跳過：找不到 shares_cache.json，股本版虛擬持倉本次不會有新訊號")
    except Exception as e:
        log.warning("STEP 5.9b 股本版訊號篩選失敗（不影響主流程）：%s", e)

    # Step 5.10: 虛擬持倉更新 — 用跟回測相同的 exit_logic 規則，每天檢查買進/賣出訊號
    buy_alerts, sell_alerts, positions_snapshot = [], [], []
    buy_alerts_capital, sell_alerts_capital, positions_capital_snapshot = [], [], []
    try:
        if price_data:
            pos_df, buy_alerts, sell_alerts = update_positions(price_data, briefing_picks, TODAY_STR)
            positions_snapshot = json.loads(pos_df.to_json(orient="records", force_ascii=False))
            log.info("STEP 5.10: 虛擬持倉更新完成，新買進 %d 筆、新賣出 %d 筆", len(buy_alerts), len(sell_alerts))

            pos_df_cap, buy_alerts_capital, sell_alerts_capital = update_positions(
                price_data, briefing_picks_capital, TODAY_STR, csv_path=POSITIONS_CAPITAL_CSV)
            positions_capital_snapshot = json.loads(pos_df_cap.to_json(orient="records", force_ascii=False))
            log.info("STEP 5.10b: 股本版虛擬持倉更新完成，新買進 %d 筆、新賣出 %d 筆",
                      len(buy_alerts_capital), len(sell_alerts_capital))
        else:
            log.info("STEP 5.10 跳過（無 price_data，可能是 --skip-yf）")
            # --skip-yf 時仍嘗試讀回現有 positions.csv，讓儀表板不會突然變空
            try:
                positions_snapshot = json.loads(load_positions().to_json(orient="records", force_ascii=False))
            except Exception:
                positions_snapshot = []
            try:
                positions_capital_snapshot = json.loads(
                    load_positions(POSITIONS_CAPITAL_CSV).to_json(orient="records", force_ascii=False))
            except Exception:
                positions_capital_snapshot = []
    except Exception as e:
        log.warning("STEP 5.10 虛擬持倉更新失敗（不影響主流程）：%s", e)

    # Step 6: 合併輸出
    out = step6_merge(analysis, radar, strong_stocks, newhigh, insti_ma, vcp_results, minervini_results, canslim_results, cb_result, overheat, breakout, disposition, v8_scan, newhigh_contraction_scan, market_pivot, margin_data, kline_data, buy_alerts, sell_alerts, positions_snapshot, positions_capital_snapshot, buy_alerts_capital, sell_alerts_capital, briefing_picks_capital, business_cycle, inst_summary, margin_maintenance, fear_greed, vix, twvix, recent_highs, pe_river)

    # Step 6.5: 記錄今日各策略選股到 entries_history.csv（供出場邏輯回測工具累積歷史樣本）
    try:
        log_daily_entries(strong_stocks, newhigh, insti_ma, vcp_results,
                           minervini_results, canslim_results, v8_scan, breakout, newhigh_contraction_scan)
    except Exception as e:
        log.warning("每日選股歷史紀錄失敗（不影響主流程）：%s", e)

    elapsed = int(time.time() - t0)
    log.info(sep)
    log.info("完成！耗時 %d 分 %d 秒", elapsed//60, elapsed%60)
    log.info("輸出：%s", out.resolve())
    log.info(sep)
    print(f"\n✅ dashboard.json 已輸出至：{out.resolve()}")
    print(f"   載入儀表板：點右上角 [📂 dashboard.json] 按鈕")

    return {
        "analysis": analysis, "radar": radar, "strong_stocks": strong_stocks,
        "newhigh": newhigh, "insti_ma": insti_ma, "vcp_results": vcp_results,
        "minervini_results": minervini_results, "canslim_results": canslim_results,
        "cb_result": cb_result, "overheat": overheat, "breakout": breakout,
        "disposition": disposition, "v8_scan": v8_scan,
        "newhigh_contraction_scan": newhigh_contraction_scan, "elapsed": elapsed,
        "out": out, "buy_alerts": buy_alerts, "sell_alerts": sell_alerts,
        "buy_alerts_capital": buy_alerts_capital, "sell_alerts_capital": sell_alerts_capital,
        "market_pivot": market_pivot, "margin_data": margin_data,
        "margin_maintenance": margin_maintenance, "inst_summary": inst_summary,
        "fear_greed": fear_greed, "vix": vix, "twvix": twvix, "recent_highs": recent_highs, "pe_river": pe_river,
    }


def load_scan_data_for_notify() -> dict:
    """--stage notify 單獨執行時，從 output/dashboard.json 讀回上次掃描的結果"""
    dash_path = OUTPUT_DIR / "dashboard.json"
    if not dash_path.exists():
        log.error("找不到 %s，請先用 --stage scan（或 --stage all）跑過一次掃描", dash_path)
        sys.exit(1)
    with open(dash_path, encoding="utf-8") as f:
        dash = json.load(f)
    log.info("STAGE=notify：從 %s 載入資料（產生於 %s）", dash_path, dash.get("generated_at","?"))
    return {
        "analysis": dash.get("analysis_result"), "radar": dash.get("radar_multi"),
        "strong_stocks": dash.get("strong_stocks"), "newhigh": dash.get("newhigh"),
        "insti_ma": dash.get("insti_ma"), "vcp_results": dash.get("vcp_result"),
        "minervini_results": dash.get("minervini_result"),
        "canslim_results": dash.get("canslim_result"), "cb_result": dash.get("cb_result"),
        "overheat": dash.get("overheat"), "breakout": dash.get("breakout"),
        "disposition": dash.get("disposition"), "v8_scan": dash.get("v8_scan"),
        "newhigh_contraction_scan": dash.get("newhigh_contraction_scan"),
        "elapsed": 0, "out": dash_path,
        "buy_alerts": dash.get("buy_alerts") or [], "sell_alerts": dash.get("sell_alerts") or [],
        "buy_alerts_capital": dash.get("buy_alerts_capital") or [],
        "sell_alerts_capital": dash.get("sell_alerts_capital") or [],
        "market_pivot": dash.get("market_pivot"), "margin_data": dash.get("margin_data"),
        "margin_maintenance": dash.get("margin_maintenance"), "inst_summary": dash.get("inst_summary"),
        "fear_greed": dash.get("fear_greed"), "vix": dash.get("vix"), "twvix": dash.get("twvix"),
        "recent_highs": dash.get("recent_highs"),
        "pe_river": dash.get("pe_river"),
    }


def run_notify_stage(data: dict):
    """執行通知階段：把 run_scan_stage() 或 load_scan_data_for_notify() 的結果發送 Telegram"""
    vcp_results = data["vcp_results"]; breakout = data["breakout"]
    disposition = data["disposition"]; v8_scan = data["v8_scan"]
    newhigh_contraction_scan = data["newhigh_contraction_scan"]; elapsed = data["elapsed"]
    buy_alerts = data.get("buy_alerts") or []; sell_alerts = data.get("sell_alerts") or []
    buy_alerts_capital = data.get("buy_alerts_capital") or []
    sell_alerts_capital = data.get("sell_alerts_capital") or []
    market_pivot = data.get("market_pivot"); margin_data = data.get("margin_data")
    margin_maintenance = data.get("margin_maintenance"); inst_summary = data.get("inst_summary")
    fear_greed = data.get("fear_greed")

    # ── Telegram 通知：跟「每日快報」同一套精選內容（大盤總覽+5分類）──
    _send_tg_all_modules(
        vcp_results, breakout, disposition, v8_scan, newhigh_contraction_scan,
        market_pivot, margin_data, elapsed, margin_maintenance, inst_summary, fear_greed
    )

    # ── 虛擬持倉 買進/賣出 Telegram 通知（每日快報5分類，套用回測同一套 exit_logic）──
    _send_tg_position_alerts("💼 虛擬持倉異動", buy_alerts, sell_alerts)

    # ── 虛擬持倉－股本版 買進/賣出 Telegram 通知（前三名策略＋股本濾網／處置預警不設限）──
    _send_tg_position_alerts("🎯 虛擬持倉異動（股本版）", buy_alerts_capital, sell_alerts_capital)


def _send_tg_position_alerts(title: str, buy_alerts: list, sell_alerts: list):
    """共用：組裝＋發送一則「持倉異動」Telegram 通知，原版／股本版都呼叫這個，避免重複程式碼、
    也保證兩邊的訊息格式一致，好對照。沒有異動就不發送，不會產生空白通知洗版。"""
    if not TG_ENABLED or not (buy_alerts or sell_alerts):
        return
    try:
        today_str4 = datetime.now().strftime("%Y-%m-%d")
        lines = [f"📊 台股儀表板 {today_str4}\n", f"{title}\n"]

        if sell_alerts:
            lines.append(f"🔴 賣出（{len(sell_alerts)} 筆）")
            for s in sell_alerts[:20]:
                ret = s.get("ret_pct")
                ret_str = f"{ret:+.2f}%" if ret is not None else "—"
                delay_note = "（延遲偵測，非當日觸發）" if s.get("delayed") else ""
                lines.append(
                    f"  {s['code']} {s['name']}（{s['strategy']}）{delay_note}\n"
                    f"    進場 {s['entry_date']} @{s['entry_price']} → "
                    f"出場 {s['exit_date']} @{s['exit_price']}  {ret_str}\n"
                    f"    原因：{s['exit_reason']}"
                )
            lines.append("")

        if buy_alerts:
            lines.append(f"🟢 買進（{len(buy_alerts)} 筆）")
            for b in buy_alerts[:20]:
                lines.append(f"  {b['code']} {b['name']}（{b['strategy']}）@{b['entry_price']}")

        _tg_send("\n".join(lines))
        log.info("%s TG 通知已發送（買進%d、賣出%d）", title, len(buy_alerts), len(sell_alerts))
    except Exception as e:
        log.warning("%s TG 通知失敗：%s", title, e)




def main():
    parser = argparse.ArgumentParser(description="台股整合儀表板一鍵執行")
    parser.add_argument("--skip-flow", action="store_true", help="跳過 Step1 法人資料更新")
    parser.add_argument("--skip-pivot", action="store_true", help="跳過 Step1b 大盤月轉折更新（K線圖.html）")
    parser.add_argument("--skip-margin", action="store_true", help="跳過 Step1c 融資餘額更新")
    parser.add_argument("--skip-maintenance", action="store_true",
                         help="跳過 Step1e 融資維持率估算（注意：這是自己算出來的估算值，不是官方數字）")
    parser.add_argument("--skip-feargreed", action="store_true",
                         help="跳過 Step1g 恐慌貪婪指數計算（自製版，六項指標綜合，不是市場公認的標準指標）")
    parser.add_argument("--skip-vix", action="store_true",
                         help="跳過 Step1h VIX指數抓取（CBOE恐慌指數，外部參考用）")
    parser.add_argument("--skip-twvix", action="store_true",
                         help="跳過 Step1i 台指VIX抓取（VIXTWN，台灣期交所編製，台股原生恐慌指標）")
    parser.add_argument("--skip-pe-river", action="store_true",
                         help="跳過 Step1j 大盤本益比河流（每天累加market_pe_history.csv一筆）")
    parser.add_argument("--enable-cycle", action="store_true",
                         help="（相容用，Step1d 現在預設就會嘗試更新，不需要再加這個參數）")
    parser.add_argument("--skip-cycle", action="store_true",
                         help="跳過 Step1d 景氣對策信號更新（每月才更新一次，來源 index.ndc.gov.tw）")
    parser.add_argument("--skip-yf",   action="store_true", help="跳過 yfinance 下載（測試用）")
    parser.add_argument("--stage", choices=["all","scan","notify"], default="all",
                        help="all=選股+發送(預設，跟以前一樣) / scan=只選股不發送 / notify=只發送(讀取上次dashboard.json)")
    args = parser.parse_args()

    if args.stage in ("all", "scan"):
        data = run_scan_stage(args)
    else:
        data = load_scan_data_for_notify()

    if args.stage in ("all", "notify"):
        run_notify_stage(data)
    else:
        log.info("STAGE=scan：只執行選股，未發送通知（要發送請另外執行 --stage notify）")


if __name__ == "__main__":
    main()
